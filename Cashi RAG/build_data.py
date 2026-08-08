# -*- coding: utf-8 -*-
"""
build_data.py — Pipeline ÚNICO de datos del Dashboard Vista Territorio.

Reemplaza a: precompute_aggs.py, regenerar_jsons_desde_parquet.py,
             01_generar_jsons_con_variaciones.py, 02_generar_nvo_plan.py

Fuentes de verdad:
  r25_sem26.parquet   Real 2025
  r26_sem26.parquet   Real 2026
  p26_sem26.parquet   Plan 2026 (original)
  nvo_plan_2026.parquet  Nvo Plan 2026 (ancho, 53 sem, signo invertido -> *-1)
  Catálogo grupo cuentas.xlsx   (Cta Mayor -> ID_CONCEPTO_CUENTA_NIV3, junto a build_data.py)
  Catálogo de estructura.xlsx   (CECO_1    -> ID_CENTRO_COSTOS, junto a build_data.py)

Salidas (carpeta aggs/):
  _consolidado.parquet   tabla larga (Serie x semana x todas las dims) [intermedio reusable]
  <dim>.parquet          agregados dim x semana x 4 series (con columnas padre para drill)
  cierres.parquet        conteo de PDCs por Estatus / tipo movimiento / semana de cierre
  movimiento.parquet     Plan vs Nvo Plan por PDC (redistribución)
  cuenta_semana.parquet  cuenta x semana (para pestaña Detalle Cuenta)
  _meta.json             sem_min, sem_max_real, sem_max_plan, generado

Uso:
    python build_data.py              # todo (usa _consolidado si existe)
    python build_data.py --rebuild    # fuerza reconstruir _consolidado desde parquets
    python build_data.py --aggs-only  # solo regenera aggs desde _consolidado existente
"""
import argparse
import functools
import gc
import json
import sys
import time
from datetime import datetime
from pathlib import Path

import duckdb
import polars as pl

import semana

BASE = Path(__file__).parent.resolve()
OUT = BASE / "aggs"
OUT.mkdir(exist_ok=True)
DL = Path.home() / "Downloads"


def _fuente(prefijo, tabla, sem):   # sem: int o None (py3.9: sin `int | None`)
    """Parquet de la semana pedida; si no está, el más reciente en disco.
    Antes esto era el literal 'r26_sem26.parquet' y por eso la app seguía
    leyendo el corte viejo aunque se hubiera extraído uno nuevo."""
    p = BASE / semana.nombres(sem)[tabla]
    if not p.exists():
        alt = semana.parquet_mas_reciente(prefijo)
        if alt is None:
            raise FileNotFoundError(
                f"no hay parquet de {tabla}: falta {p.name} y no existe ningún "
                f"{prefijo}_sem*.parquet. Corre primero: python extraer_sql.py")
        print(f"[aviso] {p.name} no existe; uso {alt.name}", flush=True)
        p = alt
    return p.as_posix()


SEM = None      # lo fija main() con --semana; None = semana EKT vigente
R25 = R26 = P26 = None    # se resuelven en _resolver_fuentes()


def _resolver_fuentes(sem=None):
    global R25, R26, P26
    R25 = _fuente("r25", "Real 2025", sem)
    R26 = _fuente("r26", "Real 2026", sem)
    P26 = _fuente("p26", "Plan 2026", sem)
    print(f"[fuentes] {Path(R25).name} · {Path(R26).name} · {Path(P26).name}", flush=True)


NVO = BASE / "nvo_plan_2026.parquet"
CONS = OUT / "_consolidado.parquet"

CAT_CUENTAS = BASE / "Catálogo grupo cuentas.xlsx"
CAT_ESTRUCT = BASE / "Catálogo de estructura.xlsx"
# Forecast por Cta Mayor (2026-07-24, reemplaza el forecast_ceco.parquet SOLO
# en las columnas Forecast/vs Forecast de tablas/árboles que se puedan cruzar
# por ID_CONCEPTO_CUENTA_NIV3 — Resumen, Div/Terr jerarquía Contable, Detalle
# Cuenta. La pestaña "Plan + Real" NO se toca, sigue con forecast_ceco.parquet.
CAT_FORECAST_CUENTA = BASE / "FCST VTA RAPIDA.xlsx"
# hoja '4.Detalle PDC': Eco (col 3) -> Nombre (col 4); da nombre a los CECOs sin PDC
# 2026-07-27: antes apuntaba a un nombre de semana fijo (S25) y se quedaba
# atrás cada vez que Planeación soltaba un archivo nuevo (ya vamos en S29) ->
# se detecta el .xlsx con el número de semana más alto en la carpeta, igual
# que semana.parquet_mas_reciente() hace con los parquets.
CAT_CONCIL_DIR = (Path.home() / "OneDrive - Onuris Tenant"
                   / "Archivos de Karlo De Jesus Juarez Ramirez - Planeación Financiera"
                   / "99. Catálogos" / "3. Conciliación PDC")


def _concil_mas_reciente() -> Path:
    import re
    cands = []
    if CAT_CONCIL_DIR.exists():
        for p in CAT_CONCIL_DIR.glob("Conciliación PDC_*_S*.xlsx"):
            m = re.search(r"_S(\d+)\.xlsx$", p.name)
            if m:
                cands.append((int(m.group(1)), p))
    if not cands:
        raise FileNotFoundError(
            f"no encontré ningún 'Conciliación PDC_*_S<sem>.xlsx' en {CAT_CONCIL_DIR}")
    return max(cands)[1]


CAT_CONCIL = _concil_mas_reciente()
# Eco PDV (id) -> Conciliación (nombre del PDC oficial); da el ID que la app
# antepone como "ID · Nombre" en cat_PDC / cat_Conciliacion. Junto a los demás
# catálogos en BASE (no en la raíz de Downloads) por la misma razón que
# CAT_CUENTAS/CAT_ESTRUCT: 2026-07-21 encontramos que un catálogo en la raíz
# de Downloads nunca refleja las correcciones que el usuario guarda en la
# carpeta del dashboard (ver CLAUDE_errores_build.md).
CAT_PUNTOS = BASE / "Puntos de contacto.xlsx"
# Pestaña Cierres (2026-07-24, reemplaza cat_Agrupa3 por completo): categorías
# de "Proyecto" (Mismos/Nuevos/Baja/Transformaciones/Transformaciones
# Matriz/Aperturas Modulo de Préstamos/Aperturas) por PDC, desde el Master
# Cierres de Planeación. Se reemplaza manualmente cada semana con el mismo
# nombre de archivo (decisión del usuario), copiado a la carpeta del
# dashboard por la misma razón que los demás catálogos (ver nota arriba).
CAT_CIERRES_EXP = BASE / "Seguimiento Expansion.xlsx"

# ---------------------------------------------------------------- dimensiones
# alias -> columna en el consolidado
DIMS = {
    "division": "cat_Direccion_Division",
    "territorio": "cat_Subdireccion_Territorio",
    "zona": "cat_Subdireccion_Zona",
    "region": "cat_Subdireccion_Region",
    "pdc": "cat_PDC",
    "formato": "cat_Formato",
    "grupo_cuentas": "cat_Grupo_de_Cuentas",
    "cuentas": "cat_Cuentas",
    "naturaleza": "cat_Naturaleza",
    "agrupador_reales": "cat_Agrupador_Reales",
    "agrupa1": "cat_Agrupa1",
    "agrupa2": "cat_Agrupa2",
    "agrupa3": "cat_Agrupa3",
    "agrupador": "cat_Agrupador",
    "agrupadores": "cat_Agrupadores",
    "clasificacion2": "cat_Clasificacion_2",
    "conciliacion": "cat_Conciliacion",
    "segmento1": "cat_Segmento1",
    "segmento2": "cat_Segmento2",
    "estatus": "cat_Estatus",
    "trimestre": "TRIMESTRE",
    "global": None,
}

# columnas padre extra por dimensión (habilitan drill-down sin fusionar homónimos)
HIER = {
    "division": ["cat_Agrupa1", "cat_Prefijo"],
    "territorio": ["cat_Agrupa1", "cat_Direccion_Division"],
    "zona": ["cat_Agrupa1", "cat_Direccion_Division", "cat_Subdireccion_Territorio"],
    "region": ["cat_Agrupa1", "cat_Direccion_Division", "cat_Subdireccion_Territorio", "cat_Subdireccion_Zona"],
    "pdc": ["cat_Agrupa1", "cat_Grupo_de_Cuentas", "cat_Direccion_Division", "cat_Subdireccion_Territorio",
            "cat_Subdireccion_Zona", "cat_Subdireccion_Region"],
    "cuentas": ["cat_Agrupa1", "cat_Grupo_de_Cuentas"],
    "agrupa2": ["cat_Agrupa1"],
    "agrupa3": ["cat_Agrupa1", "cat_Agrupa2"],
}

# (columna Excel estructura, alias en consolidado)
COLS_ESTRUCT = [
    ("Prefijo", "cat_Prefijo"),
    ("Dirección / División", "cat_Direccion_Division"),
    ("Subdirección / Territorio", "cat_Subdireccion_Territorio"),
    ("Subdirección / Zona", "cat_Subdireccion_Zona"),
    ("Subdirección / Región", "cat_Subdireccion_Region"),
    ("PDC", "cat_PDC"),
    ("Formato", "cat_Formato"),
    ("Agrupa 1", "cat_Agrupa1"),
    ("Agrupa 2", "cat_Agrupa2"),
    ("Agrupa 3", "cat_Agrupa3"),
    ("Agrupador", "cat_Agrupador"),
    ("Agrupadores", "cat_Agrupadores"),
    ("Clasificacion_2", "cat_Clasificacion_2"),
    ("Conciliación", "cat_Conciliacion"),
    ("Segmento 1", "cat_Segmento1"),
    ("Segmento 2", "cat_Segmento2"),
    ("Estatus", "cat_Estatus"),
    ("Subtipo", "cat_Subtipo"),
    ("Aperturas", "cat_Aperturas"),
    ("PDV Cierres - Transformaciones - Aperturas", "cat_Mov_PDV"),
    ("Semana de Cierre", "cat_Semana_Cierre"),
]
COLS_CUENTAS = [
    ("Grupo de Cuentas", "cat_Grupo_de_Cuentas"),
    ("Cuentas", "cat_Cuentas"),
    ("Naturaleza", "cat_Naturaleza"),
    ("Agrupador Reales", "cat_Agrupador_Reales"),
    ("Denom.Cta. Mayor", "cat_CtaMayor_Nombre"),
    ("PosPre", "cat_PosPre"),
    ("Denom.PosPre", "cat_Denom_PosPre"),
]

MEDIDAS = ["Real_2025", "Plan_2026", "Real_2026", "Nvo_Plan_2026"]

# ------------------------------------------------------- cuentas duplicadas (split 1:2)
# El catálogo de cuentas trae 20 "Cta Mayor" duplicadas (llaves repetidas, problema
# conocido). De esas, estas 12 tienen un PosPre DISTINTO entre sus 2 filas duplicadas:
# es una relación real 1:2 (misma cuenta, mismo "Denom.Cta. Mayor", reparte a 2
# posiciones presupuestales) y NO dato sucio -> decisión del usuario: conservar AMBAS
# filas (split) en vez de deduplicar con keep="first". El consolidado de movimientos
# no dice a cuál PosPre pertenece cada movimiento (solo agrega por Cta Mayor), así
# que NO se debe dividir el monto: cada movimiento se muestra completo (100%) bajo
# AMBAS combinaciones Grupo+PosPre (fan-out real del join), con una nota en la UI
# aclarando que es el mismo monto duplicado por falta de detalle de PosPre en el
# origen -> decisión del usuario 2026-07-20 (antes se partía 50/50, estaba mal:
# nadie pidió prorratear, solo mostrar el detalle). Para 2 de las 12 (6010100047 y
# 6250100014) el "Grupo de Cuentas" TAMBIÉN cambia entre filas duplicadas -> "split
# completo": se respeta la combinación (Grupo de Cuentas, PosPre) de cada fila como
# combinación válida distinta, cada una con el monto completo.
CUENTAS_SPLIT_POSPRE = [
    "1140999006", "1140999025", "1140999041", "1150090107", "6010100047",
    "6010600004", "6200501009", "6200600017", "6210100006", "6210119001",
    "6250100004", "6250100014",
]


def log(msg):
    print(f"[{time.strftime('%H:%M:%S')}] {msg}", flush=True)


@functools.lru_cache(maxsize=None)
def read_xlsx(path: Path) -> pl.DataFrame:
    """polars viejo/nuevo: fuerza todo a string (infer_schema_length=0).
    Cacheado: CAT_CUENTAS se lee 5 veces y CAT_ESTRUCT 2 veces por corrida
    completa (build_consolidado, ceco_nombres, build_orden_gpo,
    build_responsables_gpo, build_forecast_ceco/except) -> sin caché es el
    mismo .xlsx parseado desde cero cada vez. Seguro: cada caller aplica sus
    propias .rename()/.with_columns() (inmutables en polars) sobre el
    resultado, nunca lo muta in-place."""
    for kw in ({"infer_schema_length": 0},
               {"read_csv_options": {"infer_schema_length": 0}},
               {"read_options": {"infer_schema_length": 0}}):
        try:
            return pl.read_excel(path, **kw)
        except TypeError:
            continue
    return pl.read_excel(path)


def split_cuentas_especiales(cta: pl.DataFrame) -> pl.DataFrame:
    """Separa las 12 Cta Mayor de CUENTAS_SPLIT_POSPRE (ver comentario junto a la
    constante) del resto del catálogo antes de deduplicar. Para esas 12, conserva
    TODAS sus filas duplicadas (2 c/u, combinación Grupo de Cuentas + PosPre
    distinta); para el resto aplica el .unique(keep="first") de siempre.
    Devuelve el catálogo completo (especiales + resto) con las filas especiales
    marcadas en la columna booleana '_es_split', que build_consolidado usa para
    identificar qué filas vienen de fan-out 1:2 (cada una conserva el monto
    COMPLETO del movimiento, no se divide -> ver comentario en CUENTAS_SPLIT_POSPRE)."""
    especiales = (cta.filter(pl.col("ID_CONCEPTO_CUENTA_NIV3").is_in(CUENTAS_SPLIT_POSPRE))
                     .with_columns(pl.lit(True).alias("_es_split")))
    resto = (cta.filter(~pl.col("ID_CONCEPTO_CUENTA_NIV3").is_in(CUENTAS_SPLIT_POSPRE))
                .unique(subset=["ID_CONCEPTO_CUENTA_NIV3"], keep="first")
                .with_columns(pl.lit(False).alias("_es_split")))
    n_esp = especiales.select(pl.col("ID_CONCEPTO_CUENTA_NIV3").n_unique()).item()
    log(f"  cuentas especiales (split PosPre): {n_esp}/{len(CUENTAS_SPLIT_POSPRE)} "
        f"encontradas en catálogo -> {especiales.height} filas conservadas")
    return pl.concat([especiales, resto], how="vertical_relaxed")


def nvo_plan_long() -> pl.DataFrame:
    """Ancho (53 cols) -> largo. Signo invertido para igualar a las otras series."""
    df = pl.read_parquet(NVO)
    sem_cols = [c for c in df.columns if c.startswith("Nvo Plan Sem ")]
    if not sem_cols:
        raise RuntimeError("nvo_plan_2026.parquet no tiene columnas 'Nvo Plan Sem N'")
    melt = df.melt if hasattr(df, "melt") else df.unpivot
    long = melt(id_vars=["Ce.gestor", "Id Concepto Cuenta Niv3"],
                value_vars=sem_cols, variable_name="st", value_name="m")
    return long.with_columns([
        pl.col("st").str.replace("Nvo Plan Sem ", "").cast(pl.Int64).alias("SEMANA_EKT"),
        (pl.col("m").cast(pl.Float64, strict=False) * -1).alias("monto"),
        pl.col("Ce.gestor").cast(pl.Utf8).str.strip_chars().alias("ID_CENTRO_COSTOS"),
        pl.col("Id Concepto Cuenta Niv3").cast(pl.Utf8).str.strip_chars().alias("ID_CONCEPTO_CUENTA_NIV3"),
    ]).select(["ID_CENTRO_COSTOS", "ID_CONCEPTO_CUENTA_NIV3", "SEMANA_EKT", "monto"])


# ---------------------------------------------------------------- consolidado
def build_consolidado(con: duckdb.DuckDBPyConnection):
    log("Leyendo catálogos...")
    cta = read_xlsx(CAT_CUENTAS)
    est = read_xlsx(CAT_ESTRUCT)
    if "Cta Mayor" in cta.columns:
        cta = cta.rename({"Cta Mayor": "ID_CONCEPTO_CUENTA_NIV3"})
    if "CECO_1" in est.columns:
        est = est.rename({"CECO_1": "ID_CENTRO_COSTOS"})
    cta = cta.with_columns(pl.col("ID_CONCEPTO_CUENTA_NIV3").cast(pl.Utf8, strict=False).str.strip_chars())
    est = est.with_columns(pl.col("ID_CENTRO_COSTOS").cast(pl.Utf8, strict=False).str.strip_chars())

    # el catálogo de estructura tiene CECOs repetidos -> deduplicar o el join explota
    n0 = est.height
    est = est.unique(subset=["ID_CENTRO_COSTOS"], keep="first")
    if est.height != n0:
        log(f"  estructura: {n0:,} -> {est.height:,} filas (deduplicado por CECO)")
    n0 = cta.height
    cta = split_cuentas_especiales(cta)
    if cta.height != n0:
        log(f"  cuentas: {n0:,} -> {cta.height:,} filas (deduplicado por Cta Mayor, "
            f"12 cuentas especiales conservan sus 2 filas -> ver CUENTAS_SPLIT_POSPRE)")

    log("Preparando Nvo Plan (unpivot 53 semanas)...")
    nvo = nvo_plan_long()
    log(f"  nvo plan: {nvo.height:,} filas")

    con.register("cat_c", cta.to_arrow())
    con.register("cat_e", est.to_arrow())
    con.register("nvo", nvo.to_arrow())

    log("Uniendo 4 series...")
    con.execute(f"""
        CREATE OR REPLACE TABLE du AS
          SELECT ID_CENTRO_COSTOS, ID_CONCEPTO_CUENTA_NIV3, SEMANA_EKT, 'R25' AS Serie,
                 CAST("SUM_MONTO_MNX_CONS_(SUMA)" AS DOUBLE) AS monto FROM read_parquet('{R25}')
          UNION ALL SELECT ID_CENTRO_COSTOS, ID_CONCEPTO_CUENTA_NIV3, SEMANA_EKT, 'P26',
                 CAST("SUM_MONTO_MNX_CONS_(SUMA)" AS DOUBLE) FROM read_parquet('{P26}')
          UNION ALL SELECT ID_CENTRO_COSTOS, ID_CONCEPTO_CUENTA_NIV3, SEMANA_EKT, 'R26',
                 CAST("SUM_MONTO_MNX_CONS_(SUMA)" AS DOUBLE) FROM read_parquet('{R26}')
          UNION ALL SELECT ID_CENTRO_COSTOS, ID_CONCEPTO_CUENTA_NIV3, SEMANA_EKT, 'NVO', monto FROM nvo
    """)
    del nvo
    gc.collect()

    ec, cc = set(est.columns), set(cta.columns)
    def ce(n, a):
        return f'COALESCE(NULLIF(TRIM(CAST(e."{n}" AS VARCHAR)), \'\'), \'(Sin dato)\') AS {a}' if n in ec else f"'(Sin dato)' AS {a}"
    def cca(n, a):
        return f'COALESCE(NULLIF(TRIM(CAST(c."{n}" AS VARCHAR)), \'\'), \'(Sin dato)\') AS {a}' if n in cc else f"'(Sin dato)' AS {a}"

    faltan = [n for n, _ in COLS_ESTRUCT if n not in ec] + [n for n, _ in COLS_CUENTAS if n not in cc]
    if faltan:
        log(f"  AVISO columnas ausentes en catálogos: {faltan}")

    sel = ",\n            ".join([cca(n, a) for n, a in COLS_CUENTAS] +
                                 [ce(n, a) for n, a in COLS_ESTRUCT])

    log("Construyendo consolidado (join + persistir)...")
    con.execute(f"""
        CREATE TABLE consolidado_raw AS
        SELECT
            CAST(d.SEMANA_EKT AS INTEGER) AS sem,
            d.Serie,
            -- 12 cuentas de CUENTAS_SPLIT_POSPRE tienen 2 filas en cat_c (Grupo+PosPre
            -- distinto) -> el join hace fan-out 1:2 y el movimiento aparece 2 veces,
            -- una por cada PosPre, cada una con el monto COMPLETO (no se divide: ver
            -- comentario en CUENTAS_SPLIT_POSPRE -> decisión del usuario 2026-07-20,
            -- el detalle por PosPre debe mostrar el monto real, la UI aclara que es
            -- el mismo movimiento duplicado por falta de detalle en el origen).
            d.monto AS monto,
            d.ID_CENTRO_COSTOS,
            d.ID_CONCEPTO_CUENTA_NIV3,
            {sel},
            CASE WHEN CAST(d.SEMANA_EKT AS INTEGER)<=13 THEN 'Q1'
                 WHEN CAST(d.SEMANA_EKT AS INTEGER)<=26 THEN 'Q2'
                 WHEN CAST(d.SEMANA_EKT AS INTEGER)<=39 THEN 'Q3'
                 ELSE 'Q4' END AS TRIMESTRE
        FROM du d
        LEFT JOIN cat_c c ON CAST(d.ID_CONCEPTO_CUENTA_NIV3 AS VARCHAR)=CAST(c.ID_CONCEPTO_CUENTA_NIV3 AS VARCHAR)
        LEFT JOIN cat_e e ON CAST(d.ID_CENTRO_COSTOS AS VARCHAR)=CAST(e.ID_CENTRO_COSTOS AS VARCHAR)
    """)
    con.execute("DROP TABLE du")
    gc.collect()

    n = con.execute("SELECT COUNT(*) FROM consolidado_raw").fetchone()[0]
    log(f"  consolidado: {n:,} filas")

    log(f"Persistiendo {CONS.name} (para futuras corridas rápidas)...")
    con.execute(f"COPY consolidado_raw TO '{CONS.as_posix()}' (FORMAT PARQUET, COMPRESSION ZSTD)")
    log(f"  {CONS.name}: {CONS.stat().st_size/1e6:.0f} MB")


def detect_cortes(con) -> dict:
    """Semana de corte del Real 2026: última con volumen >= 50% del promedio de las 8 primeras."""
    stats = con.execute(
        "SELECT sem, COUNT(*) n FROM consolidado WHERE Serie='R26' AND sem IS NOT NULL GROUP BY 1 ORDER BY 1"
    ).fetchall()
    if not stats:
        raise RuntimeError("no hay filas Serie='R26'")
    base = sum(n for _, n in stats[:8]) / max(len(stats[:8]), 1)
    smr = 1
    for s, n in stats:
        if n >= 0.5 * base:
            smr = int(s)
        else:
            break
    smp = con.execute("SELECT MAX(sem) FROM consolidado WHERE Serie IN ('P26','NVO')").fetchone()[0] or 53
    return {"sem_min": 1, "sem_max_real": smr, "sem_max_plan": int(smp),
            "generado": datetime.now().isoformat(timespec="seconds")}


# ---------------------------------------------------------------- agregados
def ceco_nombres() -> "pl.DataFrame":
    """{ID CECO -> nombre} para etiquetar los CECOs sin PDC como '<id> · <nombre>'.
    Fuente principal: 'Descripción' del Catálogo de estructura (ahí viven los
    CECOs administrativos/staff). Complemento: hoja '4.Detalle PDC' del
    Conciliación PDC (Eco -> Nombre, por si aparece algún eco de PDV)."""
    nombres = {}
    try:
        est = read_xlsx(CAT_ESTRUCT)
        if "CECO_1" in est.columns:
            est = est.rename({"CECO_1": "ID_CENTRO_COSTOS"})
        for r in (est.select(["ID_CENTRO_COSTOS", "Descripción"]).drop_nulls()
                     .with_columns(pl.col("ID_CENTRO_COSTOS").cast(pl.Utf8, strict=False)
                                     .str.strip_chars()).to_dicts()):
            nombres.setdefault(r["ID_CENTRO_COSTOS"], str(r["Descripción"]).strip())
    except Exception as e:
        log(f"  AVISO: estructura sin Descripción de CECO ({e})")
    try:
        from openpyxl import load_workbook
        ws = load_workbook(CAT_CONCIL, read_only=True)["4.Detalle PDC"]
        en_datos = False
        for row in ws.iter_rows(values_only=True):
            if not en_datos:
                en_datos = row[2] == "Eco" and row[3] == "Nombre"
                continue
            if row[2] is not None and row[3]:
                nombres.setdefault(str(row[2]).strip(), str(row[3]).strip())
    except Exception as e:
        log(f"  AVISO: sin hoja 4 del Conciliación PDC ({e})")
    log(f"  ceco_nombres: {len(nombres):,} cecos con nombre")
    return pl.DataFrame({"eco": list(nombres.keys()), "nombre": list(nombres.values())},
                        schema={"eco": pl.Utf8, "nombre": pl.Utf8})


# cat_Grupo_de_Cuentas va en TODOS los aggs (baja cardinalidad, 843 valores).
# Subtipo/Cta Mayor solo en los aggs chicos usados por Resumen/Trimestres +
# Cierres: en pdc.parquet (2.6M filas, 2,930 ctas mayor distintas) inflarían
# el tamaño varias veces -> ahí el filtro global no aplica (decisión del usuario).
FILTROS_LIGEROS = ["cat_Subtipo", "ID_CONCEPTO_CUENTA_NIV3", "cat_CtaMayor_Nombre"]
ALGGS_CON_FILTROS_LIGEROS = {"global", "division", "grupo_cuentas", "cuentas",
                             "trimestre", "conciliacion"}
# cat_Cuentas (alta cardinalidad) solo en los aggs de Puntos de Contacto: al
# meterla en conciliacion/cierres_det el archivo se disparaba de 41MB a 718MB
# y de 18MB a 122MB -> decisión del usuario: filtro global de Cuenta solo
# aplica en Resumen/Semanal/Trimestres/Cierres-cards (aggs chicos) + PDC.
ALGGS_CON_CUENTAS = {"pdc", "conciliacion"}


# columnas extra SOLO del agg 'cuentas': cat_PosPre/cat_Denom_PosPre, propagación
# análoga a cat_Grupo_de_Cuentas. Las 12 cuentas de CUENTAS_SPLIT_POSPRE llegan aquí
# ya duplicadas (fan-out 1:2 desde 'consolidado', monto completo en cada fila); al
# agrupar por PosPre aparecen con sus 2 combinaciones Grupo+PosPre, cada una con el
# monto real (no partido) — la UI debe aclarar que es el mismo movimiento repetido
# por falta de detalle de PosPre en el origen.
COLS_EXTRA_AGG = {"cuentas": ["cat_PosPre", "cat_Denom_PosPre"]}


def build_aggs(con, meta, resume=False):
    log("Generando agregados por dimensión...")
    for alias, col in DIMS.items():
        fp = (OUT / f"{alias}.parquet").as_posix()
        if resume and Path(fp).exists():
            n = con.execute(f"SELECT COUNT(*) FROM read_parquet('{fp}')").fetchone()[0]
            log(f"  {alias:18s} {n:>8,} filas (ya existe, saltando)")
            continue
        gcols = ([col] if col else []) + HIER.get(alias, [])
        if "cat_Grupo_de_Cuentas" not in gcols:
            gcols.append("cat_Grupo_de_Cuentas")
        if "cat_Agrupa1" not in gcols:
            gcols.append("cat_Agrupa1")
        if alias in ALGGS_CON_CUENTAS and "cat_Cuentas" not in gcols:
            gcols.append("cat_Cuentas")
        if alias in ALGGS_CON_FILTROS_LIGEROS:
            for fg in FILTROS_LIGEROS:
                if fg not in gcols:
                    gcols.append(fg)
        for extra in COLS_EXTRA_AGG.get(alias, []):
            if extra not in gcols:
                gcols.append(extra)
        gsql = "".join(f'"{c}", ' for c in gcols)
        sql = f"""
            SELECT {gsql} sem,
                SUM(CASE WHEN Serie='R25' THEN monto ELSE 0 END) AS Real_2025,
                SUM(CASE WHEN Serie='P26' THEN monto ELSE 0 END) AS Plan_2026,
                SUM(CASE WHEN Serie='R26' THEN monto ELSE 0 END) AS Real_2026,
                SUM(CASE WHEN Serie='NVO' THEN monto ELSE 0 END) AS Nvo_Plan_2026
            FROM consolidado GROUP BY {gsql} sem
        """
        con.execute(f"COPY ({sql}) TO '{fp}' (FORMAT PARQUET)")
        n = con.execute(f"SELECT COUNT(*) FROM read_parquet('{fp}')").fetchone()[0]
        log(f"  {alias:18s} {n:>8,} filas")


def build_pdc_pospre(con):
    """2026-07-24: agregada cat_Subdireccion_Zona (orden Territorio->Zona->
    Región, igual que HIER_TERR en app.py) — faltaba desde el diseño original
    del 22-jul (esa versión solo tenía 3 vistas, sin "Zona"); la vista "Zona"
    de Detalle Cuenta en app.py pedía la columna y el árbol tronaba
    (ColumnNotFoundError: cat_Subdireccion_Zona).
    2026-07-27: agregada cat_Agrupador_Reales — reemplaza a cat_PosPre_Full
    como última rama del árbol de Detalle Cuenta (pedido del usuario). Es
    1:1 con cat_Cuentas (viene del catálogo de cuentas, no del movimiento),
    así que no fragmenta el join ni cambia los totales."""
    log("Generando pdc_pospre.parquet...")
    gcols = ["cat_Agrupa1", "cat_Grupo_de_Cuentas", "cat_Cuentas", "cat_Agrupador_Reales",
              "cat_Direccion_Division", "cat_Subdireccion_Territorio",
              "cat_Subdireccion_Zona", "cat_Subdireccion_Region",
              "cat_PosPre_Full"]
    gsql = "".join(f'"{c}", ' for c in gcols)
    fp = (OUT / "pdc_pospre.parquet").as_posix()
    con.execute(f"""
        COPY (
            SELECT {gsql} sem,
                SUM(CASE WHEN Serie='R25' THEN monto ELSE 0 END) AS Real_2025,
                SUM(CASE WHEN Serie='P26' THEN monto ELSE 0 END) AS Plan_2026,
                SUM(CASE WHEN Serie='R26' THEN monto ELSE 0 END) AS Real_2026,
                SUM(CASE WHEN Serie='NVO' THEN monto ELSE 0 END) AS Nvo_Plan_2026
            FROM (
                SELECT *,
                    CASE WHEN cat_PosPre = '(Sin dato)' THEN '(Sin dato)'
                         ELSE cat_PosPre || ' · ' || cat_Denom_PosPre END AS cat_PosPre_Full
                FROM consolidado
            ) GROUP BY {gsql} sem
        ) TO '{fp}' (FORMAT PARQUET, COMPRESSION ZSTD)
    """)
    n = con.execute(f"SELECT COUNT(*) FROM read_parquet('{fp}')").fetchone()[0]
    log(f"  pdc_pospre         {n:>8,} filas")


def build_cierres_arbol(con, resume=False):
    """2026-07-29: agg para el árbol de la pestaña Cierres (Grupo de Cuenta ->
    Cuenta -> Agrupador de Reales -> PDC, pedido del usuario). Por PDC en vez
    de por territorio (a diferencia de pdc_pospre) para poder filtrar por los
    ECO de la tabla de Seguimiento Expansión y expandir hasta el PDC exacto.
    --resume lo salta si ya existe, igual que build_aggs, para no regenerar
    aggs grandes que no cambiaron con cada corrida chica."""
    fp = (OUT / "cierres_arbol.parquet").as_posix()
    if resume and Path(fp).exists():
        n = con.execute(f"SELECT COUNT(*) FROM read_parquet('{fp}')").fetchone()[0]
        log(f"  cierres_arbol      {n:>8,} filas (ya existe, saltando)")
        return
    log("Generando cierres_arbol.parquet...")
    gcols = ["cat_Grupo_de_Cuentas", "cat_Cuentas", "cat_Agrupador_Reales", "cat_PDC"]
    gsql = "".join(f'"{c}", ' for c in gcols)
    con.execute(f"""
        COPY (
            SELECT {gsql} sem,
                SUM(CASE WHEN Serie='R25' THEN monto ELSE 0 END) AS Real_2025,
                SUM(CASE WHEN Serie='P26' THEN monto ELSE 0 END) AS Plan_2026,
                SUM(CASE WHEN Serie='R26' THEN monto ELSE 0 END) AS Real_2026,
                SUM(CASE WHEN Serie='NVO' THEN monto ELSE 0 END) AS Nvo_Plan_2026
            FROM consolidado GROUP BY {gsql} sem
        ) TO '{fp}' (FORMAT PARQUET, COMPRESSION ZSTD)
    """)
    n = con.execute(f"SELECT COUNT(*) FROM read_parquet('{fp}')").fetchone()[0]
    log(f"  cierres_arbol      {n:>8,} filas")


_SUMS = """
                SUM(CASE WHEN Serie='R25' THEN monto ELSE 0 END) AS Real_2025,
                SUM(CASE WHEN Serie='P26' THEN monto ELSE 0 END) AS Plan_2026,
                SUM(CASE WHEN Serie='R26' THEN monto ELSE 0 END) AS Real_2026,
                SUM(CASE WHEN Serie='NVO' THEN monto ELSE 0 END) AS Nvo_Plan_2026"""


def build_orden_gpo():
    """Mapa Grupo de Cuentas -> 'Orden presentación Gpo. Cta' del catálogo.
    La app ordena con él las tablas/gráficas de grupos (1..n; sin orden -> al final)."""
    log("Generando orden_gpo.json...")
    try:
        cta = read_xlsx(CAT_CUENTAS)
    except Exception as e:
        log(f"  AVISO: no pude leer {CAT_CUENTAS.name}: {e}")
        return
    col = "Orden presentación Gpo. Cta"
    if col not in cta.columns or "Grupo de Cuentas" not in cta.columns:
        log(f"  AVISO: catálogo sin columna '{col}'")
        return
    # el catálogo trae duplicados (p.ej. 'Gastos de Operación' con 4 y 998) -> min
    m = (cta.select(["Grupo de Cuentas", col])
            .with_columns(pl.col(col).cast(pl.Int64, strict=False))
            .drop_nulls()
            .group_by("Grupo de Cuentas").agg(pl.col(col).min()))
    orden = {r["Grupo de Cuentas"]: r[col] for r in m.to_dicts()}
    (OUT / "orden_gpo.json").write_text(json.dumps(orden, ensure_ascii=False, indent=2),
                                        encoding="utf-8")
    log(f"  orden_gpo.json: {len(orden)} grupos")



def build_responsables_gpo():
    """Genera Responsable Interno -> grupos desde Catálogo grupo cuentas.xlsx.
    Se guarda aparte para no inflar ni obligar a reconstruir los parquets."""
    log("Generando responsables_gpo.json...")
    try:
        cta = read_xlsx(CAT_CUENTAS)
    except Exception as e:
        log(f"  AVISO: no pude leer {CAT_CUENTAS.name}: {e}")
        return

    def normalizar(nombre):
        import unicodedata
        texto = unicodedata.normalize("NFKD", str(nombre))
        return "".join(ch for ch in texto if not unicodedata.combining(ch)).casefold().strip()

    columnas = {normalizar(c): c for c in cta.columns}
    col_resp = columnas.get("responsable interno")
    col_gpo = columnas.get("grupo de cuentas")
    if not col_resp or not col_gpo:
        log(f"  AVISO: catálogo sin Responsable Interno o Grupo de Cuentas. Columnas: {cta.columns}")
        return

    pares = (
        cta.select([col_resp, col_gpo])
        .with_columns([
            pl.col(col_resp).cast(pl.Utf8, strict=False).str.strip_chars(),
            pl.col(col_gpo).cast(pl.Utf8, strict=False).str.strip_chars(),
        ])
        .filter(
            pl.col(col_resp).is_not_null()
            & pl.col(col_gpo).is_not_null()
            & (pl.col(col_resp) != "")
            & (pl.col(col_gpo) != "")
        )
        .unique()
    )

    responsables = {}
    for fila in pares.to_dicts():
        responsable = fila[col_resp]
        grupo = fila[col_gpo]
        responsables.setdefault(responsable, set()).add(grupo)

    salida = {k: sorted(v) for k, v in sorted(responsables.items())}
    (OUT / "responsables_gpo.json").write_text(
        json.dumps(salida, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    log(f"  responsables_gpo.json: {len(salida):,} responsables · {len(pares):,} relaciones")

def build_pdc_ids():
    """Mapa nombre de PDC (cat_PDC / cat_Conciliacion) -> Eco PDV (id). La app
    antepone este id como 'ID · Nombre'.
    Fuente principal: 'Puntos de contacto.xlsx' (catálogo oficial vigente).
    Complemento: 'Catálogo de estructura.xlsx' (columnas PDC / Eco PDV / Eco) —
    los PDC Cerrados/Cancelados (ej. distribuidores Italika dados de baja) ya
    no aparecen en Puntos de contacto pero sí conservan su Eco en la
    estructura; sin este complemento salían sin id en los árboles/tablas
    (pedido del usuario 2026-07-27, catálogo de estructura recién actualizado)."""
    log("Generando pdc_ids.json...")
    ids: dict = {}

    if CAT_PUNTOS.exists():
        try:
            df = read_xlsx(CAT_PUNTOS)
            if df.width < 2:
                log(f"  AVISO: {CAT_PUNTOS.name} sin las columnas esperadas")
            else:
                df = df[:, :2]
                df.columns = ["eco", "nombre"]
                # '0'/'Sin Agrupar' no son un PDC oficial; el catálogo trae duplicados
                # de nombre donde una fila es el Eco PDV real y la otra es relleno con
                # eco=0 (ej. 'Mega Cd Juarez Azt' -> filas eco=0 Y eco=144) -> se descarta
                # eco=0 ANTES de group_by, si no min() siempre se queda con el 0 inválido.
                m = (df.filter(pl.col("eco").is_not_null() & pl.col("nombre").is_not_null()
                               & (pl.col("nombre") != "Sin Agrupar"))
                       .with_columns(pl.col("eco").cast(pl.Int64, strict=False))
                       .drop_nulls()
                       .filter(pl.col("eco") != 0)
                       .group_by("nombre").agg(pl.col("eco").min()))
                ids = {r["nombre"]: r["eco"] for r in m.to_dicts()}
        except Exception as e:
            log(f"  AVISO: no pude leer {CAT_PUNTOS.name}: {e}")
    else:
        log(f"  AVISO: no existe {CAT_PUNTOS.name}")

    n_oficiales = len(ids)
    try:
        est = read_xlsx(CAT_ESTRUCT)
        col_eco = "Eco PDV" if "Eco PDV" in est.columns else "Eco"
        if "PDC" in est.columns and col_eco in est.columns:
            m2 = (est.select([pl.col("PDC").alias("nombre"), pl.col(col_eco).alias("eco")])
                     .filter(pl.col("nombre").is_not_null() & pl.col("eco").is_not_null()
                             & (pl.col("nombre") != "Sin Agrupar") & (pl.col("nombre") != "0"))
                     .with_columns(pl.col("eco").cast(pl.Utf8).str.strip_chars()
                                     .cast(pl.Int64, strict=False))
                     .drop_nulls()
                     .filter(pl.col("eco") != 0)
                     .group_by("nombre").agg(pl.col("eco").min()))
            for r in m2.to_dicts():
                ids.setdefault(r["nombre"], r["eco"])
        else:
            log(f"  AVISO: {CAT_ESTRUCT.name} sin columnas PDC/{col_eco}")
    except Exception as e:
        log(f"  AVISO: no pude complementar pdc_ids con {CAT_ESTRUCT.name} ({e})")

    (OUT / "pdc_ids.json").write_text(json.dumps(ids, ensure_ascii=False, indent=2),
                                      encoding="utf-8")
    log(f"  pdc_ids.json: {len(ids):,} PDCs con id ({n_oficiales:,} oficiales + "
        f"{len(ids) - n_oficiales:,} de estructura)")


def build_pdc_cecos(con):
    """Mapa cat_PDC -> lista de ID_CENTRO_COSTOS (CECO) que caen bajo ese PDC,
    para el buscador por CECO de la pestaña Detalle PDC. JSON ligero aparte en
    vez de meter ID_CENTRO_COSTOS en el agg 'pdc': esa columna tiene 21,056
    valores únicos contra 4,115 de cat_PDC (varios PDC genéricos como '0' o
    'Cerrados y Cancelados' agrupan cientos de CECOs) -> agregarla al agg
    numérico multiplicaría filas ~5x (7.8M -> ~40M), igual que pasó con
    pdc_pospre. Aquí solo se guarda la relación de texto, sin duplicar montos."""
    log("Generando pdc_cecos.json...")
    filas = con.execute("""
        SELECT DISTINCT cat_PDC, CAST(ID_CENTRO_COSTOS AS VARCHAR) AS ceco
        FROM consolidado
        WHERE ID_CENTRO_COSTOS IS NOT NULL
    """).fetchall()
    mapa: dict = {}
    for pdc, ceco in filas:
        mapa.setdefault(pdc, []).append(ceco)
    (OUT / "pdc_cecos.json").write_text(json.dumps(mapa, ensure_ascii=False), encoding="utf-8")
    log(f"  pdc_cecos.json: {len(mapa):,} PDC -> {sum(len(v) for v in mapa.values()):,} CECOs")


def build_cierres(con):
    """Pestaña Cierres (estructura del dashboard HTML de referencia):
    - cierres_det: Tipo (Agrupa3) → GPO → Territorio → División x sem (árbol);
      conserva Región/Subtipo/Cta Mayor para los filtros de la pestaña. Sin
      cat_Cuentas (alta cardinalidad, ya no se usa en el árbol -> 122MB->18MB).
    - cierres_pdc: Tipo → PDC (con División/Territorio y GPO para el filtro global) x sem
    La app los lee con scan_parquet."""
    log("Generando cierres_det / cierres_pdc...")
    fp = (OUT / "cierres_det.parquet").as_posix()
    con.execute(f"""
        COPY (
            SELECT cat_Agrupa3, cat_Grupo_de_Cuentas,
                   cat_Direccion_Division, cat_Subdireccion_Territorio,
                   cat_Subdireccion_Region, cat_Subtipo, ID_CONCEPTO_CUENTA_NIV3,
                   cat_CtaMayor_Nombre, sem,{_SUMS}
            FROM consolidado GROUP BY 1,2,3,4,5,6,7,8,9
        ) TO '{fp}' (FORMAT PARQUET, COMPRESSION ZSTD)
    """)
    n = con.execute(f"SELECT COUNT(*) FROM read_parquet('{fp}')").fetchone()[0]
    log(f"  cierres_det        {n:>8,} filas")
    fp = (OUT / "cierres_pdc.parquet").as_posix()
    con.execute(f"""
        COPY (
            SELECT cat_Agrupa3, cat_Grupo_de_Cuentas, cat_PDC,
                   cat_Direccion_Division, cat_Subdireccion_Territorio, sem,{_SUMS}
            FROM consolidado GROUP BY 1,2,3,4,5,6
        ) TO '{fp}' (FORMAT PARQUET, COMPRESSION ZSTD)
    """)
    n = con.execute(f"SELECT COUNT(*) FROM read_parquet('{fp}')").fetchone()[0]
    log(f"  cierres_pdc        {n:>8,} filas")


def build_cierres_expansion(con):
    """2026-07-24: pestaña Cierres REEMPLAZA cat_Agrupa3 por las categorías de
    'Proyecto' del Master Cierres de Expansión (Mismos/Nuevos/Baja/
    Transformaciones/Transformaciones Matriz/Aperturas Modulo de
    Préstamos/Aperturas) — decisión del usuario: reemplazo total, ya no
    Operativo Normal/Sin Movimiento/Otros/Cerrados y Cancelados. Todo depende
    de este Excel — pedido explícito del usuario, SIN cruzar contra
    pdc_ids.json (el intento inicial via 'Eco' solo matcheaba 45% de los PDC,
    'Mismos'/'Nuevos' casi no cruzaban por ser aperturas recientes aún no
    catalogadas). Se usa directo la columna 'PDC' (nombre) del propio Excel,
    cruzada por NOMBRE contra consolidado.cat_PDC.

    El Excel (CAT_CIERRES_EXP, hoja 'Seguimiento semanal') solo trae PDC
    (nombre, col 'PDC') + categoría (col 'Proyecto'), SIN monto — el gasto se
    cruza con lo que el dashboard ya calcula por PDC (consolidado).
    Salida: aggs/cierres_expansion.parquet — mismo shape que cierres_det.parquet
    (Categoria en vez de cat_Agrupa3) para poder reusar arbol_multicol()."""
    log("Generando cierres_expansion.parquet...")
    if not CAT_CIERRES_EXP.exists():
        log(f"  AVISO: no existe {CAT_CIERRES_EXP.name} — se omite cierres_expansion.parquet")
        return

    import openpyxl
    wb = openpyxl.load_workbook(CAT_CIERRES_EXP, read_only=True, data_only=True)
    ws = wb["Seguimiento semanal"]
    # fila 21 = encabezados (Proyecto/Especialidad/.../Eco/PDC/...), datos desde la 22.
    rows = list(ws.iter_rows(min_row=22, values_only=True))
    wb.close()
    proyectos, pdcs = [], []
    for r in rows:
        proyecto, pdc_nombre = r[1], r[6]
        if proyecto is None or pdc_nombre is None:
            continue
        proyectos.append(str(proyecto).strip())
        pdcs.append(str(pdc_nombre).strip())
    exp = pl.DataFrame({"Categoria": proyectos, "cat_PDC": pdcs})
    log(f"  Seguimiento Expansión: {exp.height:,} PDCs (columna 'PDC' del Excel, sin cruzar catálogo)")

    con.register("cierres_exp_map", exp.to_arrow())
    fp = (OUT / "cierres_expansion.parquet").as_posix()
    # LEFT JOIN desde el Excel (fuente maestra, "todo depende de este Excel" —
    # pedido explícito): un PDC del plan de expansión que aún no tiene
    # movimiento en el sistema transaccional (apertura muy reciente) igual
    # aparece, con Real/Plan/Nvo Plan en 0 en vez de desaparecer del listado.
    con.execute(f"""
        COPY (
            SELECT m.Categoria, m.cat_PDC,
                   COALESCE(c.cat_Grupo_de_Cuentas, '(Sin dato)') AS cat_Grupo_de_Cuentas,
                   COALESCE(c.cat_Direccion_Division, '(Sin dato)') AS cat_Direccion_Division,
                   COALESCE(c.cat_Subdireccion_Territorio, '(Sin dato)') AS cat_Subdireccion_Territorio,
                   COALESCE(c.cat_Subdireccion_Region, '(Sin dato)') AS cat_Subdireccion_Region,
                   c.cat_Subtipo, c.ID_CONCEPTO_CUENTA_NIV3, c.cat_CtaMayor_Nombre,
                   COALESCE(c.sem, 1) AS sem,
                   COALESCE(SUM(CASE WHEN c.Serie='R25' THEN c.monto ELSE 0 END), 0.0) AS Real_2025,
                   COALESCE(SUM(CASE WHEN c.Serie='P26' THEN c.monto ELSE 0 END), 0.0) AS Plan_2026,
                   COALESCE(SUM(CASE WHEN c.Serie='R26' THEN c.monto ELSE 0 END), 0.0) AS Real_2026,
                   COALESCE(SUM(CASE WHEN c.Serie='NVO' THEN c.monto ELSE 0 END), 0.0) AS Nvo_Plan_2026
            FROM cierres_exp_map m
            LEFT JOIN consolidado c ON c.cat_PDC = m.cat_PDC
            GROUP BY 1,2,3,4,5,6,7,8,9,10
        ) TO '{fp}' (FORMAT PARQUET, COMPRESSION ZSTD)
    """)
    n = con.execute(f"SELECT COUNT(*) FROM read_parquet('{fp}')").fetchone()[0]
    n_pdc = con.execute(f"SELECT COUNT(DISTINCT cat_PDC) FROM read_parquet('{fp}')").fetchone()[0]
    n_con_gasto = con.execute(
        f"SELECT COUNT(DISTINCT cat_PDC) FROM read_parquet('{fp}') WHERE Real_2026 <> 0 OR Plan_2026 <> 0"
    ).fetchone()[0]
    log(f"  cierres_expansion  {n:>8,} filas · {n_pdc:,} PDCs del Excel · "
        f"{n_con_gasto:,} con gasto cruzado, {n_pdc - n_con_gasto:,} en $0 (sin match en consolidado)")


def build_movimiento(con):
    """Pestaña Movimiento: redistribución Plan Original vs Nvo Plan por PDC (año completo)."""
    log("Generando movimiento (Plan vs Nvo Plan)...")
    fp = (OUT / "movimiento.parquet").as_posix()
    con.execute(f"""
        COPY (
            SELECT cat_Direccion_Division, cat_Subdireccion_Territorio, cat_PDC,
                   cat_Grupo_de_Cuentas,
                SUM(CASE WHEN Serie='P26' THEN monto ELSE 0 END) AS Plan_Original,
                SUM(CASE WHEN Serie='NVO' THEN monto ELSE 0 END) AS Plan_Nuevo,
                SUM(CASE WHEN Serie='NVO' THEN monto ELSE 0 END)
              - SUM(CASE WHEN Serie='P26' THEN monto ELSE 0 END) AS Delta
            FROM consolidado
            GROUP BY 1,2,3,4
            HAVING SUM(CASE WHEN Serie='P26' THEN monto ELSE 0 END) <> 0
                OR SUM(CASE WHEN Serie='NVO' THEN monto ELSE 0 END) <> 0
        ) TO '{fp}' (FORMAT PARQUET)
    """)
    n = con.execute(f"SELECT COUNT(*) FROM read_parquet('{fp}')").fetchone()[0]
    log(f"  movimiento         {n:>8,} filas")


FORECAST_CSV = DL / "Hoja 3_data (1).csv"

def build_forecast_ceco(con):
    """Pestaña Forecast: cruza el Forecast (CSV externo, ImporteLocal x CECO x
    CuentaMayor x semana) con el Real 2026 que ya tenemos en 'consolidado'.
    Se agrega por GRUPO DE CUENTAS + CUENTA (no por CECO — decisión del
    usuario, la tabla/gráficas de la pestaña son por cuenta) cruzando
    CuentaMayor del CSV con ID_CONCEPTO_CUENTA_NIV3 del catálogo de cuentas
    (~93% de match). Conserva División/Territorio/CECO para filtros futuros.
    Salida: aggs/forecast_ceco.parquet.
    Si el CSV no está, no falla el pipeline (pestaña opcional)."""
    if not FORECAST_CSV.exists():
        log(f"  AVISO: no existe {FORECAST_CSV.name} — se omite forecast_ceco.parquet")
        return
    log("Generando forecast_ceco.parquet...")
    # ImporteLocal viene con coma decimal (formato ES) -> string, luego a float.
    fc = (pl.scan_csv(FORECAST_CSV, separator=";", dtypes={"ImporteLocal": pl.Utf8})
          .select([
              pl.col("ID_CECO").cast(pl.Utf8).str.strip_chars().alias("ID_CENTRO_COSTOS"),
              pl.col("CuentaMayor").cast(pl.Utf8).str.strip_chars().alias("ID_CONCEPTO_CUENTA_NIV3"),
              pl.col("Semana").cast(pl.Int64).alias("sem"),
              pl.col("ImporteLocal").str.replace(",", ".").cast(pl.Float64).alias("Forecast"),
          ])
          .group_by(["ID_CENTRO_COSTOS", "ID_CONCEPTO_CUENTA_NIV3", "sem"])
          .agg(pl.col("Forecast").sum())
          .collect())
    log(f"  forecast CSV: {fc.height:,} filas CECO x cuenta x semana")

    con.register("fc", fc.to_arrow())
    # cat_c puede no estar registrada (se salta build_consolidado si se usa
    # el _consolidado.parquet en caché, p.ej. con --aggs-only), o estar registrada
    # como la tabla cruda del catálogo sin alias cat_* (queda así tras correr
    # build_consolidado en la misma sesión, ver split_cuentas_especiales) -> en
    # ambos casos se reconstruye aquí la vista con los alias que este bloque espera.
    # Bug preexistente (no introducido por el split de PosPre): el try original solo
    # comprobaba que "cat_c" existiera como tabla, no que tuviera cat_Grupo_de_Cuentas.
    # Además, tras --rebuild "cat_c" ya existe como tabla registrada por Arrow
    # (build_consolidado, con columnas crudas sin alias) y CREATE OR REPLACE VIEW
    # NO logra reemplazar una tabla registrada por con.register() con el mismo
    # nombre (DuckDB deja ambas coexistir y sigue resolviendo a la tabla vieja) ->
    # hay que des-registrarla primero o la vista de abajo nunca se ve.
    try:
        con.execute("SELECT cat_Grupo_de_Cuentas FROM cat_c LIMIT 1")
    except Exception:
        cta = read_xlsx(CAT_CUENTAS)
        if "Cta Mayor" in cta.columns:
            cta = cta.rename({"Cta Mayor": "ID_CONCEPTO_CUENTA_NIV3"})
        cta = cta.with_columns(pl.col("ID_CONCEPTO_CUENTA_NIV3").cast(pl.Utf8, strict=False).str.strip_chars())
        cta = cta.unique(subset=["ID_CONCEPTO_CUENTA_NIV3"], keep="first")
        con.register("cat_c_raw", cta.to_arrow())
        try:
            con.unregister("cat_c")
        except Exception:
            pass
        con.execute("""
            CREATE OR REPLACE VIEW cat_c AS
            SELECT ID_CONCEPTO_CUENTA_NIV3,
                   COALESCE(NULLIF(TRIM(CAST("Grupo de Cuentas" AS VARCHAR)), ''), '(Sin dato)') AS cat_Grupo_de_Cuentas,
                   COALESCE(NULLIF(TRIM(CAST("Cuentas" AS VARCHAR)), ''), '(Sin dato)') AS cat_Cuentas
            FROM cat_c_raw
        """)
    # Salida a nivel Grupo de Cuentas x Cuenta x semana (colapsa CECO — la
    # tabla/gráficas de la pestaña son por cuenta, no por centro de costos).
    fp = (OUT / "forecast_ceco.parquet").as_posix()
    con.execute(f"""
        COPY (
            WITH real AS (
                SELECT ID_CONCEPTO_CUENTA_NIV3, sem,
                       SUM(CASE WHEN Serie='R26' THEN monto ELSE 0 END) AS Real_2026
                FROM consolidado
                GROUP BY ID_CONCEPTO_CUENTA_NIV3, sem
            ),
            fc_agg AS (
                SELECT ID_CONCEPTO_CUENTA_NIV3, sem, SUM(Forecast) AS Forecast
                FROM fc
                GROUP BY ID_CONCEPTO_CUENTA_NIV3, sem
            ),
            joined AS (
                SELECT
                    COALESCE(f.ID_CONCEPTO_CUENTA_NIV3, r.ID_CONCEPTO_CUENTA_NIV3) AS ID_CONCEPTO_CUENTA_NIV3,
                    COALESCE(f.sem, r.sem) AS sem,
                    COALESCE(r.Real_2026, 0.0) AS Real_2026,
                    COALESCE(f.Forecast, 0.0) AS Forecast
                FROM fc_agg f
                FULL OUTER JOIN real r
                  ON f.ID_CONCEPTO_CUENTA_NIV3 = r.ID_CONCEPTO_CUENTA_NIV3 AND f.sem = r.sem
            )
            SELECT
                COALESCE(c.cat_Grupo_de_Cuentas, 'Sin agrupar') AS cat_Grupo_de_Cuentas,
                COALESCE(c.cat_Cuentas, 'Sin agrupar') AS cat_Cuentas,
                j.sem,
                SUM(j.Real_2026) AS Real_2026, SUM(j.Forecast) AS Forecast
            FROM joined j
            LEFT JOIN cat_c c ON j.ID_CONCEPTO_CUENTA_NIV3 = c.ID_CONCEPTO_CUENTA_NIV3
            GROUP BY 1, 2, 3
        ) TO '{fp}' (FORMAT PARQUET, COMPRESSION ZSTD)
    """)
    n = con.execute(f"SELECT COUNT(*) FROM read_parquet('{fp}')").fetchone()[0]
    n_cta = con.execute(f"SELECT COUNT(DISTINCT cat_Cuentas) FROM read_parquet('{fp}')").fetchone()[0]
    log(f"  forecast_ceco      {n:>8,} filas · {n_cta:,} cuentas")


def build_forecast_cuenta():
    """2026-07-24: nueva fuente de Forecast por Cta Mayor (FCST VTA RAPIDA.xlsx),
    reemplaza la fórmula "Real + Plan restante" SOLO en vistas que agrupan por
    cuenta (Resumen, Div/Terr contable, Detalle Cuenta) — ver nota junto a
    CAT_FORECAST_CUENTA. NO toca forecast_ceco.parquet (pestaña Plan + Real).

    El Excel viene ANCHO: CECO_N7 | TipoOrigen | CuentaMayor | semanas 1-53.
    Pedido explícito del usuario: ignorar CECO_N7 y TipoOrigen, agrupar y sumar
    todo por CuentaMayor (ya no se ve por Centro de Costos). Encabezados reales
    en la fila 2 (fila 1 solo trae 'Semana' en una celda suelta).
    Salida: aggs/forecast_cuenta.parquet, columnas
    ID_CONCEPTO_CUENTA_NIV3 (Cta Mayor) x sem x Forecast — SIN Grupo de
    Cuentas/Cuenta (el cruce con el catálogo vive en app.py, igual que
    Real/Plan/Nvo Plan, para no duplicar la lógica de homologación de
    mayúsculas que ya aplica load_all() a todos los aggs)."""
    if not CAT_FORECAST_CUENTA.exists():
        log(f"  AVISO: no existe {CAT_FORECAST_CUENTA.name} — se omite forecast_cuenta.parquet")
        return
    log("Generando forecast_cuenta.parquet...")
    # fila 1 del Excel trae solo 'Semana' suelta; el encabezado real (CECO_N7,
    # TipoOrigen, CuentaMayor, 1..53) está en la fila 2 -> skip_rows salta la 1ª.
    raw = pl.read_excel(CAT_FORECAST_CUENTA, sheet_name="Sheet 1",
                        read_csv_options={"skip_rows": 1})
    sem_cols = [c for c in raw.columns if str(c).strip().isdigit()]
    if not sem_cols:
        raise RuntimeError("FCST VTA RAPIDA.xlsx: no se encontraron columnas de semana (1-53)")

    # Buscar columna CuentaMayor (puede venir como 'CuentaMayor' o variante)
    cuenta_col = None
    for col in raw.columns:
        if "cuenta" in str(col).lower() and "mayor" in str(col).lower():
            cuenta_col = col
            break

    if not cuenta_col:
        log(f"  AVISO: no se encontró columna CuentaMayor. Columnas disponibles: {raw.columns}")
        log(f"  Saltando forecast_cuenta.parquet")
        return

    melt = raw.melt if hasattr(raw, "melt") else raw.unpivot
    long = melt(id_vars=[cuenta_col], value_vars=sem_cols,
               variable_name="sem", value_name="Forecast")
    out = (long
           .filter(pl.col(cuenta_col).is_not_null())
           .with_columns([
               pl.col(cuenta_col).cast(pl.Utf8).str.strip_chars().alias("ID_CONCEPTO_CUENTA_NIV3"),
               pl.col("sem").cast(pl.Int64),
               pl.col("Forecast").cast(pl.Float64, strict=False).fill_null(0.0),
           ])
           .group_by(["ID_CONCEPTO_CUENTA_NIV3", "sem"])
           .agg(pl.col("Forecast").sum())
           .sort(["ID_CONCEPTO_CUENTA_NIV3", "sem"]))
    fp = OUT / "forecast_cuenta.parquet"
    out.write_parquet(fp, compression="zstd")
    n_cta = out.select(pl.col("ID_CONCEPTO_CUENTA_NIV3").n_unique()).item()
    log(f"  forecast_cuenta    {out.height:>8,} filas · {n_cta:,} cuentas mayor")


RIESGOS_DOCX = BASE / "Riesgos y Oportunidades.docx"

def build_riesgos():
    """Lee 'Riesgos y Oportunidades.docx' (en la carpeta del dashboard, lo
    edita Planeación en Word) y lo vuelca a aggs/riesgos.json. Formato del
    Word: 4 párrafos título abren cada sección — 'RIESGOS' / 'OPORTUNIDADES'
    (bloques YTD, los que muestra la pestaña Riesgos & Oport. del dashboard) y
    'RIESGOS SEMANAL' / 'OPORTUNIDADES SEMANAL' (bloques semanales, los usa
    generar_correo_canales.py) — cada párrafo debajo de un título es una
    viñeta. Las 2 secciones semanales son opcionales (si no están en el Word,
    quedan como listas vacías). Sin dependencias: un .docx es un ZIP con
    word/document.xml (se parsea con la librería estándar)."""
    log("Generando riesgos.json...")
    if not RIESGOS_DOCX.exists():
        log(f"  AVISO: no existe {RIESGOS_DOCX.name} — corre _crear_word_riesgos.py")
        return
    import re
    import zipfile
    try:
        with zipfile.ZipFile(RIESGOS_DOCX) as z:
            xml = z.read("word/document.xml").decode("utf-8", errors="replace")
    except Exception as e:
        log(f"  AVISO: no pude leer {RIESGOS_DOCX.name}: {e}")
        return
    # párrafos = <w:p>...</w:p>; texto = concatenación de sus <w:t>
    parrafos = []
    for p in re.findall(r"<w:p[ >].*?</w:p>|<w:p/>", xml, flags=re.S):
        t = "".join(re.findall(r"<w:t[^>]*>(.*?)</w:t>", p, flags=re.S))
        t = (t.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
              .replace("&quot;", '"').replace("&#39;", "'").strip())
        if t:
            parrafos.append(t)
    secciones = {"RIESGOS": "riesgos", "OPORTUNIDADES": "oportunidades",
                 "RIESGOS SEMANAL": "riesgos_semanal", "OPORTUNIDADES SEMANAL": "oportunidades_semanal"}
    out = {"riesgos": [], "oportunidades": [], "riesgos_semanal": [], "oportunidades_semanal": [],
           "actualizado": datetime.fromtimestamp(RIESGOS_DOCX.stat().st_mtime)
                          .strftime("%Y-%m-%d %H:%M")}
    seccion = None
    for t in parrafos:
        clave = t.strip().upper().rstrip(":")
        if clave in secciones:
            seccion = secciones[clave]
        elif seccion:
            out[seccion].append(t)
    (OUT / "riesgos.json").write_text(json.dumps(out, ensure_ascii=False, indent=2),
                                      encoding="utf-8")
    log(f"  riesgos.json: {len(out['riesgos'])} riesgos · {len(out['oportunidades'])} oportunidades"
        f" · {len(out['riesgos_semanal'])} riesgos sem · {len(out['oportunidades_semanal'])} oport. sem")


# ---------------------------------------------------------------- validación
def validar(con, meta):
    log("Validando...")
    smr = meta["sem_max_real"]
    ok = True

    g = pl.read_parquet(OUT / "global.parquet")
    tot = {m: g.filter(pl.col("sem") <= smr).select(pl.col(m).sum()).item() for m in MEDIDAS}
    log(f"  Acumulado 1..{smr}:")
    for m, v in tot.items():
        log(f"    {m:16s} {v/1e6:>14,.1f} M")

    for alias in ["division", "grupo_cuentas", "pdc", "trimestre"]:
        d = pl.read_parquet(OUT / f"{alias}.parquet").filter(pl.col("sem") <= smr)
        for m in MEDIDAS:
            a, b = tot[m], d.select(pl.col(m).sum()).item()
            if abs(a - b) > max(1.0, abs(a) * 1e-9):
                log(f"    [FALLA] {alias}.{m}: {b/1e6:,.1f}M != global {a/1e6:,.1f}M")
                ok = False
    if ok:
        log("  [OK] todas las dimensiones suman igual que global")

    if tot["Real_2026"] <= 0 or tot["Plan_2026"] <= 0:
        log("  [FALLA] Real_2026 o Plan_2026 no positivo (revisar signos)")
        ok = False
    nv = g.select(pl.col("Nvo_Plan_2026").sum()).item()
    if nv <= 0:
        log(f"  [FALLA] Nvo Plan total {nv/1e6:,.1f}M debe ser positivo (signo *-1)")
        ok = False
    else:
        log(f"  [OK] Nvo Plan 53 sem = {nv/1e6:,.1f} M (signo correcto)")

    r26 = g.filter(pl.col("sem") > smr).select(pl.col("Real_2026").sum()).item()
    if abs(r26) > 1:
        log(f"  [AVISO] hay Real_2026 = {r26/1e6:,.1f}M despues del corte sem {smr}")
    return ok


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rebuild", action="store_true", help="reconstruir _consolidado desde los parquets fuente")
    ap.add_argument("--aggs-only", action="store_true", help="solo regenerar aggs desde _consolidado existente")
    ap.add_argument("--semana", type=int, default=None,
                    help="semana EKT de los parquet a leer (por defecto, la vigente)")
    ap.add_argument("--resume", action="store_true", help="continuar desde donde se paró (salta consolidado si ya existe)")
    args = ap.parse_args()

    t0 = time.time()
    con = duckdb.connect(str(BASE / "_build.duckdb"))
    # máquina con 31GB RAM / 22 cores (2026-07-21): threads=6/memory=6GB era
    # config conservadora que dejaba la mayoría de recursos sin usar durante
    # las etapas pesadas (join del consolidado, conciliacion, pdc_pospre) ->
    # subido a 16/16GB, dejando margen para el SO y otras apps corriendo en
    # paralelo (streamlit, etc.) sin arriesgar ahogar la máquina.
    con.execute("PRAGMA threads=16")
    con.execute("PRAGMA memory_limit='16GB'")
    con.execute("PRAGMA temp_directory='_ddb_tmp'")

    try:
        _resolver_fuentes(args.semana)   # necesario también para comparar mtimes contra CONS
        fuente_mas_nueva = CONS.exists() and max(
            Path(p).stat().st_mtime for p in (R25, R26, P26)) > CONS.stat().st_mtime
        if fuente_mas_nueva and not args.rebuild and not args.aggs_only and not args.resume:
            log(f"Aviso: hay parquets fuente más nuevos que {CONS.name}; forzando --rebuild automático.")
        # --resume prioriza usar caché incluso si --rebuild se pasó (resume = "continúa donde se paró")
        usar_cache = (args.resume and CONS.exists()) or (CONS.exists() and not args.rebuild and not fuente_mas_nueva)

        con.execute("DROP VIEW IF EXISTS consolidado_raw")
        con.execute("DROP TABLE IF EXISTS consolidado_raw")

        if usar_cache:
            log(f"Usando {CONS.name} existente ({CONS.stat().st_size/1e6:.0f} MB). --rebuild para recrear.")
            con.execute(f"CREATE OR REPLACE VIEW consolidado_raw AS SELECT * FROM read_parquet('{CONS.as_posix()}')")
        else:
            if args.aggs_only:
                sys.exit(f"ERROR: --aggs-only requiere {CONS.name}, que no existe. Corre sin esa bandera.")
            build_consolidado(con)

        # El catálogo trae '0' (texto) en varias columnas de la jerarquía:
        #  - PDC = '0' fusionaba miles de CECOs en un renglón -> se sustituye por
        #    'CECO <ID_CENTRO_COSTOS>' (o 'Sin agrupar' si tampoco hay CECO).
        #  - Territorio/Zona/Región = '0' son niveles de agrupación (un solo '0'
        #    cubre muchos CECOs, no se puede mapear a UN ceco) -> 'Sin agrupar'.
        con.register("ceco_nom", ceco_nombres().to_arrow())
        con.execute("""
            CREATE OR REPLACE VIEW consolidado AS
            SELECT r.* REPLACE (
                CASE WHEN r.cat_PDC IN ('0', '(Sin dato)') OR r.cat_PDC IS NULL
                     THEN COALESCE(CAST(r.ID_CENTRO_COSTOS AS VARCHAR) || ' · ' || n.nombre,
                                   'CECO ' || CAST(r.ID_CENTRO_COSTOS AS VARCHAR), 'Sin agrupar')
                     ELSE r.cat_PDC END AS cat_PDC,
                CASE WHEN cat_Subdireccion_Territorio = '0'
                     THEN 'Sin agrupar' ELSE cat_Subdireccion_Territorio END AS cat_Subdireccion_Territorio,
                CASE WHEN cat_Subdireccion_Zona = '0'
                     THEN 'Sin agrupar' ELSE cat_Subdireccion_Zona END AS cat_Subdireccion_Zona,
                CASE WHEN cat_Subdireccion_Region = '0'
                     THEN 'Sin agrupar' ELSE cat_Subdireccion_Region END AS cat_Subdireccion_Region)
            FROM consolidado_raw r
            LEFT JOIN ceco_nom n ON CAST(r.ID_CENTRO_COSTOS AS VARCHAR) = n.eco
        """)

        meta = detect_cortes(con)
        log(f"Cortes: sem_max_real={meta['sem_max_real']} sem_max_plan={meta['sem_max_plan']}")

        build_aggs(con, meta, resume=args.resume)
        build_cierres(con)
        build_cierres_expansion(con)
        build_movimiento(con)
        build_forecast_ceco(con)
        try:
            build_forecast_cuenta()
        except Exception as e:
            log(f"  [AVISO] build_forecast_cuenta falló (Excel FORECAST mal formado): {e}")
        build_pdc_pospre(con)
        build_cierres_arbol(con, resume=args.resume)
        build_pdc_cecos(con)
        build_orden_gpo()
        build_responsables_gpo()
        build_pdc_ids()
        build_riesgos()

        (OUT / "_meta.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")
        ok = validar(con, meta)
    finally:
        con.close()
        for p in (BASE / "_build.duckdb", BASE / "_build.duckdb.wal"):
            try:
                p.unlink()
            except OSError:
                pass

    log(f"{'[OK] LISTO' if ok else '[!] TERMINÓ CON FALLAS'} en {(time.time()-t0)/60:.1f} min -> {OUT}")
    sys.exit(0 if ok else 1)


if __name__ == "__main__":
    main()


