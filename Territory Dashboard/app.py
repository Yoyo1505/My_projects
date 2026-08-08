# -*- coding: utf-8 -*-
"""
Dashboard Vista Territorio — Streamlit (demo pública)

Lee agregados en aggs/. Si no hay datos, genera demo sintética.
Acceso abierto: sin login ni branding corporativo.

Correr:
    streamlit run app.py
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import plotly.graph_objects as go
import polars as pl
import streamlit as st

BASE = Path(__file__).parent.resolve()
AGGS = BASE / "aggs"
USERS_FILE = BASE / "usuarios.json"

MEDIDAS = ["Real_2025", "Plan_2026", "Real_2026", "Nvo_Plan_2026"]
# Rangos de semana fijos por trimestre (año EKT de 53 semanas: Q4 absorbe la
# semana 53 extra). Usado por la pestaña Trimestres y por la tabla Trimestral
# de Temporalidad — un solo lugar para no divergir entre las dos.
QRANGO_TRI = {"Q1": (1, 13), "Q2": (14, 26), "Q3": (27, 39), "Q4": (40, 53), "FY": (1, 53)}
COLOR = {"Real_2025": "#94a3b8", "Real_2026": "#38bdf8",
         "Plan_2026": "#a78bfa", "Nvo_Plan_2026": "#34d399",
         "Forecast_Cierre": "#fbbf24"}
NOMBRE = {"Real_2025": "Real 2025", "Real_2026": "Real 2026",
          "Plan_2026": "Plan 2026", "Nvo_Plan_2026": "Nvo Plan",
          "Forecast_Cierre": "Forecast"}

# alias -> (etiqueta, columna)
DIMS = {
    "division": ("División", "cat_Direccion_Division"),
    "territorio": ("Territorio", "cat_Subdireccion_Territorio"),
    "region": ("Región", "cat_Subdireccion_Region"),
    "zona": ("Zona", "cat_Subdireccion_Zona"),
    "pdc": ("PDC", "cat_PDC"),
    "grupo_cuentas": ("Grupo de Cuentas", "cat_Grupo_de_Cuentas"),
    "cuentas": ("Cuentas", "cat_Cuentas"),
    "pospre": ("PosPre", "cat_PosPre"),
    "formato": ("Formato", "cat_Formato"),
    "naturaleza": ("Naturaleza", "cat_Naturaleza"),
    "agrupador_reales": ("Agrupador Reales", "cat_Agrupador_Reales"),
    "agrupa1": ("Agrupa 1", "cat_Agrupa1"),
    "agrupa2": ("Agrupa 2", "cat_Agrupa2"),
    "agrupa3": ("Agrupa 3", "cat_Agrupa3"),
    "agrupador": ("Agrupador", "cat_Agrupador"),
    "clasificacion2": ("Clasificación 2", "cat_Clasificacion_2"),
    "segmento1": ("Segmento 1", "cat_Segmento1"),
    "segmento2": ("Segmento 2", "cat_Segmento2"),
    "estatus": ("Estatus", "cat_Estatus"),
    "global": ("Global", None),
}
# columnas padre (para no fusionar homónimos al agrupar)
# Jerarquía territorial: División → Territorio → Región → Zona → PDC
# (Región antes que Zona — pedido del usuario 2026-07-15).
PARENTS = {
    "division": ["cat_Agrupa1"],
    "territorio": ["cat_Agrupa1", "cat_Direccion_Division"],
    "zona": ["cat_Agrupa1", "cat_Direccion_Division", "cat_Subdireccion_Territorio"],
    "region": ["cat_Agrupa1", "cat_Direccion_Division", "cat_Subdireccion_Territorio", "cat_Subdireccion_Zona"],
    "pdc": ["cat_Agrupa1", "cat_Direccion_Division", "cat_Subdireccion_Territorio",
            "cat_Subdireccion_Zona", "cat_Subdireccion_Region"],
    "grupo_cuentas": ["cat_Agrupa1"],
    "cuentas": ["cat_Agrupa1", "cat_Grupo_de_Cuentas"],
    "pospre": ["cat_Agrupa1", "cat_Grupo_de_Cuentas", "cat_Cuentas"],
    "agrupa2": ["cat_Agrupa1"],
    "agrupa3": ["cat_Agrupa1", "cat_Agrupa2"],
}
HIER_TERR = ["agrupa1", "division", "territorio", "zona", "region", "pdc"]
HIER_CTA = ["agrupa1", "grupo_cuentas", "cuentas", "pospre"]

# ---------------------------------------------------------- calendario EKT
# Semanas de 7 días. Una semana que cruza dos meses reparte su monto entre ellos
# en proporción a los días que le tocan a cada uno (prorrateo diario).
#   Verificado con el usuario: sem 27 de 2026 = 28-jun a 4-jul  (3 días junio, 4 julio).
#   -> semana 1 de 2026 arranca el domingo 28-dic-2025.
#   2025 tuvo 52 semanas -> su semana 1 arranca el domingo 29-dic-2024.
ANCLA = {2025: date(2024, 12, 29), 2026: date(2025, 12, 28)}
ANIO = {"Real_2025": 2025, "Real_2026": 2026, "Plan_2026": 2026, "Nvo_Plan_2026": 2026}

NOM_MES = {1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
           7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"}

@st.cache_data(show_spinner=False)
def pesos_mes(anio: int) -> dict:
    """{semana: {mes: fracción de la semana que cae en ese mes}}.
    Solo cuenta los meses del propio año: los días que caen en el año vecino
    se descartan (decisión del usuario: enero recibe solo sus 3/7)."""
    ini1 = ANCLA[anio]
    out = {}
    for s in range(1, 54):
        ini = ini1 + timedelta(days=7 * (s - 1))
        dias = defaultdict(int)
        for i in range(7):
            d = ini + timedelta(days=i)
            if d.year == anio:
                dias[d.month] += 1
        if dias:
            out[s] = {m: n / 7 for m, n in dias.items()}
    return out

@st.cache_data(show_spinner=False)
def rango_mes(anio: int, mes: int) -> tuple:
    """Primera y última semana EKT que tocan ese mes (para el filtro del sidebar)."""
    p = pesos_mes(anio)
    sems = [s for s, d in p.items() if mes in d]
    return (min(sems), max(sems)) if sems else (1, 1)

def semanal_a_mensual(w: pd.DataFrame) -> pd.DataFrame:
    """Reparte cada serie semanal a meses con prorrateo diario.
    Cada serie usa el calendario de SU año (Real 2025 vs las de 2026)."""
    filas = {m: {k: 0.0 for k in MEDIDAS} for m in range(1, 13)}
    vistos = {m: {k: False for k in MEDIDAS} for m in range(1, 13)}
    for medida in MEDIDAS:
        p = pesos_mes(ANIO[medida])
        for _, r in w.iterrows():
            s = int(r["sem"])
            v = r[medida]
            if s not in p or pd.isna(v):
                continue
            for mes, frac in p[s].items():
                filas[mes][medida] += float(v) * frac
                vistos[mes][medida] = True
    out = []
    for m in range(1, 13):
        rec = {"mes": m, "Mes": NOM_MES[m]}
        for k in MEDIDAS:
            rec[k] = filas[m][k] if vistos[m][k] else None
        out.append(rec)
    return pd.DataFrame(out)

st.set_page_config(page_title="Vista Territorio",
                   page_icon="◈", layout="wide", initial_sidebar_state="expanded")

st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=JetBrains+Mono:wght@500;700&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', system-ui, sans-serif; }

.stApp {
  background: radial-gradient(1200px 600px at 10% -10%, #1e293b 0%, #0b1220 45%, #070b14 100%);
  color: #e2e8f0;
}
[data-testid="stHeader"] { background: rgba(7,11,20,.6); }

section[data-testid="stSidebar"] {
  background: linear-gradient(180deg, #111827 0%, #0f172a 100%) !important;
  border-right: 1px solid rgba(148,163,184,.12);
}
section[data-testid="stSidebar"] * { color: #e2e8f0 !important; }
section[data-testid="stSidebar"] .stCaption,
section[data-testid="stSidebar"] label { color: #94a3b8 !important; }

[data-testid="stMetric"] {
  background: linear-gradient(145deg, rgba(30,41,59,.9), rgba(15,23,42,.95));
  border: 1px solid rgba(56,189,248,.18);
  border-radius: 16px;
  padding: 14px 16px;
  box-shadow: 0 8px 24px rgba(0,0,0,.25);
}
[data-testid="stMetricValue"] {
  font-size: 1.5rem; font-family: 'JetBrains Mono', monospace; color: #f8fafc !important;
}
[data-testid="stMetricLabel"] p { font-size: .78rem; color: #94a3b8 !important; }

h1, h2, h3, h4, h5, h6 { color: #f1f5f9 !important; letter-spacing: -0.02em; }
.stMarkdown p, .stCaption, label { color: #cbd5e1 !important; }
hr { border-color: rgba(148,163,184,.15) !important; }

[data-testid="stDataFrame"] {
  border-radius: 12px; overflow: hidden;
  border: 1px solid rgba(148,163,184,.12);
}

.vt-hero {
  position: relative; border-radius: 20px; padding: 22px 26px; margin: 0 0 18px 0;
  background:
    linear-gradient(135deg, rgba(56,189,248,.14), rgba(167,139,250,.12) 45%, rgba(52,211,153,.08)),
    linear-gradient(180deg, rgba(15,23,42,.95), rgba(2,6,23,.9));
  border: 1px solid rgba(148,163,184,.16);
  box-shadow: 0 20px 50px rgba(0,0,0,.35); overflow: hidden;
}
.vt-hero-kicker {
  font-size:.72rem; letter-spacing:.18em; font-weight:700;
  color:#7dd3fc; text-transform:uppercase; margin-bottom:6px;
}
.vt-hero-title {
  font-size:1.85rem; font-weight:700; color:#f8fafc; line-height:1.15; margin:0 0 6px 0;
}
.vt-hero-sub { font-size:.95rem; color:#94a3b8; margin:0; }
.vt-badge {
  display:inline-block; margin-top:10px; padding:4px 10px; border-radius:999px;
  font-size:.72rem; font-weight:600; letter-spacing:.04em;
  background:rgba(52,211,153,.12); color:#6ee7b7; border:1px solid rgba(52,211,153,.25);
}
.vt-side-brand {
  padding: 6px 2px 14px 2px; margin-bottom: 8px;
  border-bottom: 1px solid rgba(148,163,184,.12);
}
.vt-side-brand .t {
  font-size:1.05rem; font-weight:700; color:#f8fafc !important; letter-spacing:-.02em;
}
.vt-side-brand .s { font-size:.75rem; color:#64748b !important; margin-top:2px; }

</style>""", unsafe_allow_html=True)

# =========================================================== datos
def _version_datos() -> str:
    """Huella de los agregados. Si cambia, el caché se invalida solo."""
    m = AGGS / "_meta.json"
    return f"{m.stat().st_mtime_ns}" if m.exists() else "0"

def _homologar(dfs: list) -> dict:
    """El catálogo trae el mismo valor con distinto uso de mayúsculas
    ('Gastos De Operación' vs 'Gastos de Operación') y eso duplica categorías.
    Canónico = la variante con menos mayúsculas (la bien escrita en español).
    Devuelve {columna: {variante: canónico}} mirando TODOS los aggs a la vez,
    para que cada parquet quede homologado igual."""
    variantes: dict = {}
    for df in dfs:
        for c in df.columns:
            if c.startswith("cat_") and df.schema[c] == pl.Utf8:
                for v in df[c].unique().to_list():
                    if v is not None:
                        variantes.setdefault(c, {}).setdefault(v.lower(), set()).add(v)
    mapa = {}
    for c, grupos in variantes.items():
        for vs in grupos.values():
            if len(vs) > 1:
                canon = min(vs, key=lambda s: (sum(ch.isupper() for ch in s), s))
                for v in vs:
                    if v != canon:
                        mapa.setdefault(c, {})[v] = canon
    return mapa

def _remplazar(col: str, m: dict):
    """replace (polars >= 0.20) con fallback a map_dict (0.19)."""
    e = pl.col(col)
    return (e.replace(m) if hasattr(e, "replace")
            else e.map_dict(m, default=e)).alias(col)

def _aplicar_homologacion(df: pl.DataFrame, mapa: dict) -> pl.DataFrame:
    for c, m in mapa.items():
        if c in df.columns:
            df = df.with_columns(_remplazar(c, m))
    return df

# parquets grandes que NO se leen al arrancar: se cargan (y homologan) la
# primera vez que una pestaña los pide. cierres_det (2026-07-24: ya no se usa,
# la pestaña Cierres pasó a cierres_expansion.parquet — chico, se carga
# directo, ver load_all) quedó fuera de esta lista.
# 2026-07-27 (performance): "conciliacion" también sale. Pesa 815 MB y ninguna
# vista lo consulta desde que se quitó la pestaña Movimiento; dejarlo como
# diferido permitía que un filtro accidental lo cargara entero a RAM y
# congelara el dashboard. El archivo sigue en aggs/, solo no se abre.
# 2026-07-28: confirmado que conciliacion es 815MB (62% de aggs). Agregado a DIFERIDOS.
_DIFERIDOS_AGG = {"pdc", "conciliacion"}
_DIFERIDOS_EXTRA = {"cierres_arbol"}

class _AggsDiferidos(dict):
    """dict de aggs con carga bajo demanda de los parquet grandes. El mapa de
    homologación se calcula solo con los aggs chicos: las categorías cat_*
    provienen de los mismos catálogos, así que las variantes de mayúsculas ya
    están representadas ahí; al cargar un diferido se le aplica el mismo mapa."""
    def __init__(self, pendientes: set, mapa: dict):
        super().__init__()
        self._pend, self._mapa = set(pendientes), mapa
    def __missing__(self, k):
        if k not in self._pend:
            raise KeyError(k)
        fp = AGGS / f"{k}.parquet"
        if not fp.exists():
            raise KeyError(k)
        df = _aplicar_homologacion(pl.read_parquet(fp), self._mapa)
        self[k] = df
        return df
    def get(self, k, default=None):
        try:
            return self[k]
        except KeyError:
            return default
    def __contains__(self, k):
        return dict.__contains__(self, k) or (k in self._pend and (AGGS / f"{k}.parquet").exists())

def _ensure_demo_data():
    """Si no hay agregados, genera datos sintéticos para demo pública."""
    meta_p = AGGS / "_meta.json"
    if meta_p.exists():
        return
    try:
        from seed_demo_data import build as _seed
        _seed()
    except Exception as e:
        st.error(f"No hay datos en `aggs/` y falló la demo: {e}")
        st.stop()

@st.cache_resource(show_spinner="Cargando agregados…")
def load_all(version: str):
    """`version` no se usa dentro: existe para que Streamlit recargue
    cuando build_data.py reescribe aggs/."""
    _ensure_demo_data()
    meta_p = AGGS / "_meta.json"
    if not meta_p.exists():
        st.error(f"Falta `{meta_p}`.\n\nCorre: `python seed_demo_data.py`")
        st.stop()
    meta = json.loads(meta_p.read_text(encoding="utf-8"))
    pi = AGGS / "pdc_ids.json"
    meta["pdc_ids"] = json.loads(pi.read_text(encoding="utf-8")) if pi.exists() else {}
    pcc = AGGS / "pdc_cecos.json"
    meta["pdc_cecos"] = json.loads(pcc.read_text(encoding="utf-8")) if pcc.exists() else {}
    og = AGGS / "orden_gpo.json"
    meta["orden_gpo"] = json.loads(og.read_text(encoding="utf-8")) if og.exists() else {}
    rg = AGGS / "responsables_gpo.json"
    meta["responsables_gpo"] = json.loads(rg.read_text(encoding="utf-8")) if rg.exists() else {}
    chicos = {}
    for alias in DIMS:
        if alias in _DIFERIDOS_AGG:
            continue
        fp = AGGS / f"{alias}.parquet"
        if fp.exists():
            chicos[alias] = pl.read_parquet(fp)
    chicos_extra = {}
    # cierres_expansion (2026-07-24): chico (~145 PDCs), se carga entero, no
    # lazy — a diferencia de cierres_pdc/pdc_pospre. cierres_pdc y pdc_pospre
    # se leen lazy (scan_parquet).
    for name in ("forecast_ceco", "cierres_expansion"):
        fp = AGGS / f"{name}.parquet"
        if fp.exists():
            chicos_extra[name] = pl.read_parquet(fp)
    # forecast_cuenta: ID_CONCEPTO_CUENTA_NIV3 x sem x Forecast (2026-07-24).
    # NO pasa por homologación de mayúsculas — es un ID exacto, no una
    # categoría de texto; se cruza directo por match de string en compute().
    fc_cta_p = AGGS / "forecast_cuenta.parquet"
    forecast_cuenta = pl.read_parquet(fc_cta_p) if fc_cta_p.exists() else None
    # homologar variantes de mayúsculas (catálogo sucio) con los aggs chicos
    mapa = _homologar(list(chicos.values()) + list(chicos_extra.values()))
    agg = _AggsDiferidos(_DIFERIDOS_AGG, mapa)
    agg.update({k: _aplicar_homologacion(v, mapa) for k, v in chicos.items()})
    extra = _AggsDiferidos(_DIFERIDOS_EXTRA, mapa)
    extra.update({k: _aplicar_homologacion(v, mapa) for k, v in chicos_extra.items()})
    if forecast_cuenta is not None:
        extra["forecast_cuenta"] = forecast_cuenta   # sin homologar, ver nota arriba
    meta["homologados"] = mapa
    return meta, agg, extra

VERSION = _version_datos()
META, AGG, EXTRA = load_all(VERSION)
SMR, SMP = META["sem_max_real"], META["sem_max_plan"]
# nombre oficial de PDC -> Eco PDV (id) — catálogo del Excel
PDC_IDS = META.get("pdc_ids", {})
# cat_PDC -> lista de ID_CENTRO_COSTOS (CECO) bajo ese PDC; para buscar por CECO.
PDC_CECOS = META.get("pdc_cecos", {})
_CECO_A_PDC = {ceco: pdc for pdc, cecos in PDC_CECOS.items() for ceco in cecos}
# Grupo de Cuentas -> nº de "Orden presentación Gpo. Cta" del catálogo Excel.
# El json trae ambas variantes de mayúsculas; tras homologar el dato solo
# queda la canónica, así que se re-mapean las claves con el mismo mapa.
_ORDEN_GPO_RAW = META.get("orden_gpo", {})
_map_gc = META.get("homologados", {}).get("cat_Grupo_de_Cuentas", {})
ORDEN_GPO = {_map_gc.get(k, k): v for k, v in _ORDEN_GPO_RAW.items()}
_RESP_GPO_RAW = META.get("responsables_gpo", {})
RESPONSABLES_GPO = {
    responsable: tuple(sorted({_map_gc.get(grupo, grupo) for grupo in grupos}))
    for responsable, grupos in _RESP_GPO_RAW.items()
}

def _orden_gpo_key(nombre) -> tuple:
    """Clave de sort: primero el nº de ORDEN_GPO (grupos sin orden, al final)."""
    return (ORDEN_GPO.get(nombre, 999), str(nombre))

# Orden fijo de División pedido por el usuario (2026-07-22): Norte, Centro,
# Sur, Concentrador, Cerrados y Cancelados, Corporativo, Expansión, Fuerza
# Proactiva, Otros — en ese orden exacto. Cualquier valor de división que no
# esté en este mapa (ej. 'Direccion De Geografia Nacional', 'Geografía
# Historia', '(Sin dato)') cae al final, orden alfabético entre ellos.
ORDEN_DIV = {
    "Division Norte": 0, "Division Centro": 1, "Division Sur": 2,
    "Concentrador": 3, "Cerrados y Cancelados": 4, "Corporativo": 5,
    "Expansión": 6, "Fuerza Proactiva": 7, "Otros": 8,
}

def _orden_div_key(nombre) -> tuple:
    return (ORDEN_DIV.get(nombre, 999), str(nombre))

def con_id(nombre: str) -> str:
    """'Mega Cd Juarez Azt' -> '4821 · Mega Cd Juarez Azt'. Limpia IDs 0, NaN, o vacíos."""
    if nombre is None or pd.isna(nombre) or str(nombre).strip() in ("", "0", "0.0", "nan", "NaN", "None"):
        return "Sin agrupar"
    nombre_str = str(nombre).strip()
    pid = PDC_IDS.get(nombre_str)
    if pid is None or pd.isna(pid) or str(pid).strip() in ("", "0", "0.0", "nan", "NaN", "None"):
        return nombre_str
    try:
        fval = float(pid)
        if fval.is_integer() and fval > 0:
            return f"{int(fval)} · {nombre_str}"
    except (ValueError, TypeError):
        pass
    return f"{pid} · {nombre_str}"

# Logo deshabilitado en la versión pública del repositorio
def logo_b64() -> str:
    return ""

# ---------------------------------------------------- actualización semanal
ESTADO = BASE / "estado.json"

def leer_estado() -> dict | None:
    if not ESTADO.exists():
        return None
    try:
        return json.loads(ESTADO.read_text(encoding="utf-8"))
    except Exception:
        return None

def lanzar_actualizacion():
    """No-op in public demo (SQL pipeline removed)."""
    return None

@st.cache_data(ttl=3600, show_spinner=False)
def hay_datos_nuevos(_v: str) -> str:
    """SQL checks disabled in public demo."""
    return "desconocido"

# =========================================================== cálculo
def _medidas(d: dict, total_r26: float = 0.0, total_r25: float = 0.0, total_p26: float = 0.0,
            total_nvo: float = 0.0, forecast_cta: float | None = None) -> dict:
    """forecast_cta: si no es None, sustituye la fórmula Real+Plan_Restante
    por el Forecast cruzado desde FCST VTA RAPIDA.xlsx (2026-07-24, solo en
    vistas por cuenta — ver _forecast_cruzado_por_cuenta)."""
    r25, p26, nvo, r26 = d["Real_2025"], d["Plan_2026"], d["Nvo_Plan_2026"], d["Real_2026"]
    prest = d.get("Plan_Restante", 0.0)
    vs_aa, vs_plan, vs_nvo = r26 - r25, r26 - p26, r26 - nvo
    fc = forecast_cta if forecast_cta is not None else r26 + prest
    vs_fc = r26 - fc
    return {
        "Real_2025": r25,
        "Pct_r25": (r25 / total_r25 * 100) if total_r25 else 0.0,
        "Plan_2026": p26,
        "Pct_p26": (p26 / total_p26 * 100) if total_p26 else 0.0,
        "Nvo_Plan_2026": nvo,
        "Pct_nvo": (nvo / total_nvo * 100) if total_nvo else 0.0,
        "Real_2026": r26,
        "Pct_r26": (r26 / total_r26 * 100) if total_r26 else 0.0,
        "Cumplimiento_Plan": (r26 / p26 * 100) if p26 else 0.0,
        "YoY_2026vs2025": (vs_aa / abs(r25) * 100) if r25 else 0.0,
        "Pct_del_Total": (r26 / total_r26 * 100) if total_r26 else 0.0,
        "Vs_AA_Abs": vs_aa, "Vs_AA_Pct": (vs_aa / abs(r25) * 100) if r25 else 0.0,
        "Etiqueta_AA": "Ahorro" if vs_aa <= 0 else "Gasto extra",
        "Vs_Plan_Abs": vs_plan, "Vs_Plan_Pct": (vs_plan / abs(p26) * 100) if p26 else 0.0,
        "Etiqueta_Plan": "Bajo plan" if vs_plan <= 0 else "Sobre plan",
        "Vs_NvoPlan_Abs": vs_nvo, "Vs_NvoPlan_Pct": (vs_nvo / abs(nvo) * 100) if nvo else 0.0,
        "Forecast_Cierre": fc, "Forecast_Cumpl_Pct": (fc / p26 * 100) if p26 else 0.0,
        "Vs_Forecast_Abs": vs_fc, "Vs_Forecast_Pct": (vs_fc / abs(fc) * 100) if fc else 0.0,
    }

def _peso_expr(medida: str, mes, lo: int, hi: int, sems: tuple = None):
    """Peso de cada semana para una medida."""
    if mes is None and sems is None:
        return pl.when((pl.col("sem") >= lo) & (pl.col("sem") <= hi)).then(1.0).otherwise(0.0)
    if sems is not None:
        return pl.when(pl.col("sem").is_in(list(sems)) & (pl.col("sem") <= hi)).then(1.0).otherwise(0.0)
    meses = (mes,) if isinstance(mes, int) else tuple(mes)
    p = pesos_mes(ANIO[medida])
    mapa = {s: sum(d.get(m, 0.0) for m in meses) for s, d in p.items()}
    expr = pl.lit(0.0)
    for s, f in mapa.items():
        if f > 0:
            expr = pl.when(pl.col("sem") == s).then(f).otherwise(expr)
    return expr

# alias de compute() donde SÍ se puede cruzar Forecast por ID_CONCEPTO_CUENTA_NIV3
# (2026-07-24, pedido del usuario: "solo se aplica en las que puedas cruzar
# por ID CONCEPTO CUENTA" — el resto de vistas, territoriales/PDC, siguen con
# la fórmula Real + Plan restante, sin excepción).
# 2026-07-28: "division" agregado — division.parquet SÍ trae
# ID_CONCEPTO_CUENTA_NIV3 (está en FILTROS_LIGEROS/ALGGS_CON_FILTROS_LIGEROS
# de build_data.py), así que puede cruzar igual que grupo_cuentas/cuentas. Sin
# esto, la gráfica "Agrupador acumulado" de Resumen usaba el Forecast viejo
# (Real + Plan Restante de TODAS las semanas futuras del año, ignorando el
# rango de semanas del sidebar) — reportado por el usuario como "no se ajustó
# al nuevo forecast / no toma las semanas".
_ALIAS_FORECAST_CUENTA = {"global", "grupo_cuentas", "cuentas", "division"}

def _forecast_cruzado_por_cuenta(df: pl.DataFrame, gcols: list, si: int, sf: int,
                                 individual: bool, mes=None, sems: tuple = None) -> dict | None:
    """Suma el Forecast de FCST VTA RAPIDA.xlsx (aggs/forecast_cuenta.parquet)
    por cada combinación de gcols, cruzando por ID_CONCEPTO_CUENTA_NIV3 —
    mismo prorrateo de semana/mes que Plan 2026. Devuelve {tuple(gcols): monto}
    o None si no hay fuente/columna para cruzar.

    2026-07-28: fix de fan-out — cuando gcols NO es "por cuenta" (ej. División),
    una misma cuenta vive bajo varios gcols (90% de las cuentas están en más de
    una división, hasta 11 de 12) y el join simple duplicaba el Forecast
    COMPLETO de la cuenta en cada grupo (marcaba ~3.5x de más, reportado por el
    usuario con datos reales: Division Centro salía en 15,027M de forecast vs
    4,210M de Real/Plan). Ahora, si hay fan-out, el Forecast de cada cuenta se
    PRORRATEA entre sus grupos según el peso de Real_2026 de cada uno dentro de
    esa cuenta (mismo criterio que "de dónde sacas el número" — reparto real,
    no duplicado). Pedido explícito del usuario: "viene por ID_CONCEPTO_CUENTA,
    solo usa las que tenga y saca el forecast" — se prorratea, no se descarta."""
    fcta = EXTRA.get("forecast_cuenta") if "EXTRA" in globals() else None
    if fcta is None or "ID_CONCEPTO_CUENTA_NIV3" not in df.columns:
        return None
    ids = df.select("ID_CONCEPTO_CUENTA_NIV3", *[c for c in gcols if c in df.columns]).unique()
    lo = sf if individual else si
    w = _peso_expr("Plan_2026", mes, lo, sf, sems)
    # Forecast por CUENTA, agregado sobre fcta directo (sin fan-out todavía) —
    # el join con ids se hace DESPUÉS de agrupar, nunca antes, si no
    # group_by(gcols) sobre un cruce ya fan-out sigue sumando copias.
    fc_cta = (fcta.with_columns((pl.col("Forecast") * w).alias("_fc"))
                  .group_by("ID_CONCEPTO_CUENTA_NIV3").agg(pl.col("_fc").sum()))
    cruce = fc_cta.join(ids, on="ID_CONCEPTO_CUENTA_NIV3", how="inner")
    if cruce.height == 0:
        return {}

    fan_out = (ids.group_by("ID_CONCEPTO_CUENTA_NIV3").agg(pl.count().alias("_n"))
                  .filter(pl.col("_n") > 1).height > 0)
    if fan_out and gcols:
        # peso de cada grupo dentro de su cuenta = su Real_2026 / Real_2026 total de la cuenta
        w26 = _peso_expr("Real_2026", mes, lo, min(sf, SMR), sems)
        peso_real = (df.with_columns((pl.col("Real_2026") * w26).alias("_r26"))
                       .group_by(["ID_CONCEPTO_CUENTA_NIV3"] + gcols).agg(pl.col("_r26").sum()))
        tot_cta = peso_real.group_by("ID_CONCEPTO_CUENTA_NIV3").agg(
            pl.col("_r26").abs().sum().alias("_r26_tot_cta"))
        peso_real = peso_real.join(tot_cta, on="ID_CONCEPTO_CUENTA_NIV3", how="left")
        n_grupos = peso_real.group_by("ID_CONCEPTO_CUENTA_NIV3").agg(pl.count().alias("_n"))
        peso_real = peso_real.join(n_grupos, on="ID_CONCEPTO_CUENTA_NIV3", how="left").with_columns(
            pl.when(pl.col("_r26_tot_cta") > 0)
              .then(pl.col("_r26").abs() / pl.col("_r26_tot_cta"))
              .otherwise(1.0 / pl.col("_n"))  # sin gasto real -> reparto equitativo
              .alias("_peso_grupo"))
        # join por CUENTA + gcols juntos (no solo por cuenta): "cruce" ya trae
        # una fila por cuenta×grupo (del join con "ids"), igual que
        # "peso_real" — unir solo por cuenta multiplicaría cada fila de
        # "cruce" contra TODOS los grupos que comparten esa cuenta de nuevo.
        cruce = (cruce.join(peso_real.select(["ID_CONCEPTO_CUENTA_NIV3"] + gcols + ["_peso_grupo"]),
                            on=["ID_CONCEPTO_CUENTA_NIV3"] + gcols, how="inner")
                      .with_columns((pl.col("_fc") * pl.col("_peso_grupo")).alias("_fc")))

    g = cruce.group_by(gcols).agg(pl.col("_fc").sum().alias("Forecast_Cruzado")) if gcols \
        else cruce.select(pl.col("_fc").sum().alias("Forecast_Cruzado"))
    out = {}
    for r in g.to_dicts():
        clave = tuple(r[c] for c in gcols) if gcols else ()
        out[clave] = r["Forecast_Cruzado"]
    return out

_FUENTE_ALT = {"region": "pdc", "zona": "pdc"}

@st.cache_data(show_spinner=False)
def compute(alias: str, si: int, sf: int, individual: bool, filtros: tuple = (),
            mes=None, sems: tuple = None, excluir_padres: tuple = ()) -> pd.DataFrame:
    """Agrega una dimensión en el rango (o en un mes, con prorrateo diario).
    Agrupa por PADRES + hija: no fusiona homónimos.
    `excluir_padres`: columnas padre a NO usar para agrupar — para árboles que
    omiten un nivel intermedio y donde ese padre solo fragmentaría las filas."""
    df = AGG.get(_FUENTE_ALT.get(alias, alias))
    if df is None:
        return pd.DataFrame()
    col = DIMS[alias][1]
    parents = [p for p in PARENTS.get(alias, [])
               if p in df.columns and p not in excluir_padres]
    gcols = parents + ([col] if col else [])

    for k, v in filtros:
        if k in df.columns:
            df = df.filter(pl.col(k).is_in(list(v)) if isinstance(v, tuple)
                           else pl.col(k) == v)
    if df.height == 0:
        return pd.DataFrame()

    lo = sf if individual else si
    real_hi = min(sf, SMR)

    if mes is None and sems is None:
        w25 = w_p26 = w_nvo = _peso_expr("Real_2025", None, lo, sf)
        w26 = _peso_expr("Real_2026", None, lo, real_hi)
    elif sems is not None:
        w25 = w_p26 = w_nvo = _peso_expr("Real_2025", None, lo, sf, sems)
        w26 = _peso_expr("Real_2026", None, lo, real_hi, sems)
    else:
        w25 = _peso_expr("Real_2025", mes, lo, sf)
        w_p26 = _peso_expr("Plan_2026", mes, lo, sf)
        w_nvo = _peso_expr("Nvo_Plan_2026", mes, lo, sf)
        w26 = _peso_expr("Real_2026", mes, lo, sf) * \
            pl.when(pl.col("sem") <= SMR).then(1.0).otherwise(0.0)

    df = df.with_columns([
        (pl.col("Real_2025") * w25).alias("_r25"),
        (pl.col("Plan_2026") * w_p26).alias("_p26"),
        (pl.col("Nvo_Plan_2026") * w_nvo).alias("_nvo"),
        (pl.col("Real_2026") * w26).alias("_r26"),
        pl.when(pl.col("sem") > SMR).then(pl.col("Plan_2026")).otherwise(0.0).alias("_prest"),
    ])
    e = [pl.col("_r25").sum().alias("Real_2025"), pl.col("_p26").sum().alias("Plan_2026"),
         pl.col("_nvo").sum().alias("Nvo_Plan_2026"), pl.col("_r26").sum().alias("Real_2026"),
         pl.col("_prest").sum().alias("Plan_Restante")]
    g = df.group_by(gcols).agg(e) if gcols else df.select(e)

    fc_cruzado = (_forecast_cruzado_por_cuenta(df, gcols, si, sf, individual, mes, sems)
                 if alias in _ALIAS_FORECAST_CUENTA else None)

    rows = g.to_dicts()
    tot_r25 = sum(r["Real_2025"] for r in rows)
    tot_p26 = sum(r["Plan_2026"] for r in rows)
    tot_nvo = sum(r["Nvo_Plan_2026"] for r in rows)
    tot_r26 = sum(r["Real_2026"] for r in rows)
    out = []
    for r in rows:
        fc_r = (fc_cruzado.get(tuple(r[c] for c in gcols) if gcols else (), 0.0)
               if fc_cruzado is not None else None)
        out.append({**{c: r.get(c) for c in gcols}, **_medidas(r, tot_r26, tot_r25, tot_p26, tot_nvo, fc_r)})
    res = pd.DataFrame(out)
    if gcols and not res.empty:
        res = res[res[MEDIDAS].abs().sum(axis=1) > 0.5]
        res = _ordenar_resultado(res, gcols)
    return res.reset_index(drop=True)

def _ordenar_resultado(res: pd.DataFrame, gcols: list) -> pd.DataFrame:
    if gcols == ["cat_Grupo_de_Cuentas"]:
        return res.reindex(res["cat_Grupo_de_Cuentas"].map(_orden_gpo_key)
                           .sort_values(kind="stable").index)
    if gcols and gcols[0] == "cat_Direccion_Division":
        return res.reindex(res["cat_Direccion_Division"].map(_orden_div_key)
                           .sort_values(kind="stable").index)
    return res.reindex(res["Real_2026"].abs().sort_values(ascending=False).index)

def _agregar_df(df: pl.DataFrame, gcols: list, si: int, sf: int, individual: bool,
                mes=None, sems: tuple = None) -> pd.DataFrame:
    if df.height == 0:
        return pd.DataFrame()
    lo = sf if individual else si
    real_hi = min(sf, SMR)
    if mes is None and sems is None:
        w25 = w_p26 = w_nvo = _peso_expr("Real_2025", None, lo, sf)
        w26 = _peso_expr("Real_2026", None, lo, real_hi)
    elif sems is not None:
        w25 = w_p26 = w_nvo = _peso_expr("Real_2025", None, lo, sf, sems)
        w26 = _peso_expr("Real_2026", None, lo, real_hi, sems)
    else:
        w25 = _peso_expr("Real_2025", mes, lo, sf)
        w_p26 = _peso_expr("Plan_2026", mes, lo, sf)
        w_nvo = _peso_expr("Nvo_Plan_2026", mes, lo, sf)
        w26 = _peso_expr("Real_2026", mes, lo, sf) * \
            pl.when(pl.col("sem") <= SMR).then(1.0).otherwise(0.0)

    df = df.with_columns([
        (pl.col("Real_2025") * w25).alias("_r25"),
        (pl.col("Plan_2026") * w_p26).alias("_p26"),
        (pl.col("Nvo_Plan_2026") * w_nvo).alias("_nvo"),
        (pl.col("Real_2026") * w26).alias("_r26"),
        pl.when(pl.col("sem") > SMR).then(pl.col("Plan_2026")).otherwise(0.0).alias("_prest"),
    ])
    e = [pl.col("_r25").sum().alias("Real_2025"), pl.col("_p26").sum().alias("Plan_2026"),
         pl.col("_nvo").sum().alias("Nvo_Plan_2026"), pl.col("_r26").sum().alias("Real_2026"),
         pl.col("_prest").sum().alias("Plan_Restante")]
    g = df.group_by(gcols).agg(e) if gcols else df.select(e)

    # Forecast por Cta Mayor (2026-07-24): mismo cruce que compute(), activado
    # aquí por columna en vez de alias porque _agregar_df no recibe alias —
    # solo si el df trae ID_CONCEPTO_CUENTA_NIV3 y gcols agrupa por cuenta.
    fc_cruzado = (_forecast_cruzado_por_cuenta(df, gcols, si, sf, individual, mes, sems)
                 if ("ID_CONCEPTO_CUENTA_NIV3" in df.columns
                     and ("cat_Grupo_de_Cuentas" in gcols or "cat_Cuentas" in gcols)) else None)

    rows = g.to_dicts()
    tot_r25 = sum(r["Real_2025"] for r in rows)
    tot_p26 = sum(r["Plan_2026"] for r in rows)
    tot_nvo = sum(r["Nvo_Plan_2026"] for r in rows)
    tot_r26 = sum(r["Real_2026"] for r in rows)
    out = []
    for r in rows:
        fc_r = (fc_cruzado.get(tuple(r[c] for c in gcols) if gcols else (), 0.0)
               if fc_cruzado is not None else None)
        out.append({**{c: r.get(c) for c in gcols}, **_medidas(r, tot_r26, tot_r25, tot_p26, tot_nvo, fc_r)})
    res = pd.DataFrame(out)
    if gcols and not res.empty:
        res = res[res[MEDIDAS].abs().sum(axis=1) > 0.5]
        res = _ordenar_resultado(res, gcols)
    return res.reset_index(drop=True)

@st.cache_data(show_spinner="Armando árbol…")
def arbol_multicol(fuente: str, cols: tuple, si: int, sf: int, ind: bool,
                   filtros: tuple = (), mes=None, sems: tuple = None,
                   inc_fc: bool = False) -> str:
    """Árbol expandible multinivel sobre columnas que cruzan dimensiones
    (p.ej. cat_Grupo_de_Cuentas → cat_Cuentas → cat_Direccion_Division →
    cat_Subdireccion_Territorio, todas presentes en el mismo agg 'fuente').
    Cada nivel acumula las columnas anteriores como padres, igual que PARENTS
    hace en compute(). Reusa _arbol_html / _fila_arbol."""
    df = AGG.get(fuente, EXTRA.get(fuente))
    if df is None:
        return ""
    for k, v in filtros:
        if k in df.columns:
            df = df.filter(pl.col(k).is_in(list(v)) if isinstance(v, tuple)
                           else pl.col(k) == v)
    niveles = []
    for i, c in enumerate(cols):
        gcols = list(cols[:i + 1])
        d = _agregar_df(df, gcols, si, sf, ind, mes, sems)
        niveles.append((d, c))
    return _arbol_html(niveles, inc_fc=inc_fc)

@st.cache_data(show_spinner=False)
def cuentas_desde_pdc(si: int, sf: int, ind: bool, filtros: tuple = (),
                      mes=None, sems: tuple = None) -> pd.DataFrame:
    """Como compute('cuentas'), pero sobre pdc.parquet: necesario cuando el
    filtro incluye cat_PDC, columna que el agg 'cuentas' no trae."""
    df = AGG["pdc"]
    for k, v in filtros:
        if k in df.columns:
            df = df.filter(pl.col(k).is_in(list(v)) if isinstance(v, tuple)
                           else pl.col(k) == v)
    return _agregar_df(df, ["cat_Grupo_de_Cuentas", "cat_Cuentas"],
                       si, sf, ind, mes, sems)

@st.cache_data(show_spinner=False)
def weekly(alias: str = "global", filtros: tuple = ()) -> pd.DataFrame:
    df = AGG.get(alias, AGG["global"])
    for k, v in filtros:
        if k in df.columns:
            df = df.filter(pl.col(k).is_in(list(v)) if isinstance(v, tuple)
                           else pl.col(k) == v)
    d = df.group_by("sem").agg([pl.col(m).sum() for m in MEDIDAS]).sort("sem").to_pandas()
    d.loc[d["sem"] > SMR, "Real_2026"] = None
    return d

@st.cache_data(show_spinner="Calculando cierres…")
def cierres_rango(si: int, sf: int, individual: bool, mes=None, sems: tuple = None,
                  filtros: tuple = ()) -> pd.DataFrame:
    """Agrega cierres_expansion.parquet (Categoria del Excel Seguimiento
    Expansión x PDC) al rango elegido, con la misma lógica de
    pesos/prorrateo que compute() (2026-07-24: reemplaza cierres_pdc.parquet
    / cat_Agrupa3 por completo)."""
    fp = AGGS / "cierres_expansion.parquet"
    if not fp.exists():
        return pd.DataFrame()
    df = pl.scan_parquet(fp)
    mapa = META.get("homologados", {})
    for c, m in mapa.items():
        if c in df.columns:
            df = df.with_columns(_remplazar(c, m))
    for k, v in filtros:
        if k in df.columns:
            df = df.filter(pl.col(k).is_in(list(v)) if isinstance(v, tuple)
                           else pl.col(k) == v)
    lo = sf if individual else si
    real_hi = min(sf, SMR)
    if mes is None and sems is None:
        w25 = wp = wn = _peso_expr("Real_2025", None, lo, sf)
        w26 = _peso_expr("Real_2026", None, lo, real_hi)
    elif sems is not None:
        w25 = wp = wn = _peso_expr("Real_2025", None, lo, sf, sems)
        w26 = _peso_expr("Real_2026", None, lo, real_hi, sems)
    else:
        w25 = _peso_expr("Real_2025", mes, lo, sf)
        wp = _peso_expr("Plan_2026", mes, lo, sf)
        wn = _peso_expr("Nvo_Plan_2026", mes, lo, sf)
        w26 = _peso_expr("Real_2026", mes, lo, sf) * \
            pl.when(pl.col("sem") <= SMR).then(1.0).otherwise(0.0)
    gcols = ["Categoria", "cat_PDC", "cat_Direccion_Division", "cat_Subdireccion_Territorio"]
    return (df.with_columns([
                (pl.col("Real_2025") * w25).alias("_r25"),
                (pl.col("Plan_2026") * wp).alias("_p26"),
                (pl.col("Nvo_Plan_2026") * wn).alias("_nvo"),
                (pl.col("Real_2026") * w26).alias("_r26"),
                pl.when(pl.col("sem") > SMR).then(pl.col("Plan_2026")).otherwise(0.0).alias("_prest")])
              .group_by(gcols)
              .agg([pl.col("_r25").sum().alias("Real_2025"),
                    pl.col("_p26").sum().alias("Plan_2026"),
                    pl.col("_nvo").sum().alias("Nvo_Plan_2026"),
                    pl.col("_r26").sum().alias("Real_2026"),
                    pl.col("_prest").sum().alias("Plan_Restante")])
              .collect().to_pandas())

# =========================================================== formato / charts
def M(v) -> str:
    """TODAS las cifras del dashboard van en millones de pesos (MDP) — pedido
    explícito del usuario (2026-07-27): nada de MMDP ni de K, ni siquiera en
    los drill-downs. Una sola escala en todo el reporte para poder comparar
    cualquier cifra de un vistazo."""
    if v is None or pd.isna(v):
        return "—"
    return f"{v/1e6:,.1f} MDP"

def P(v) -> str:
    return "—" if v is None or pd.isna(v) else f"{v:+.1f}%"

# --- formato contable: todo es gasto. Positivo = gasto -> rojo y entre
# paréntesis; negativo = a favor -> carbón, sin signo.
ROJO, CARBON = "#fb7185", "#e2e8f0"

def MC(v) -> str:
    if v is None or pd.isna(v):
        return "—"
    s = M(abs(v))
    return f"({s})" if v > 0 else s

def PC(v) -> str:
    if v is None or pd.isna(v):
        return "—"
    return f"({abs(v):,.1f}%)" if v > 0 else f"{abs(v):,.1f}%"

def _css_gasto(v) -> str:
    if v is None or pd.isna(v):
        return f"color:{CARBON}"
    return f"color:{ROJO};font-weight:600" if v > 0 else f"color:{CARBON}"

def _layout(fig, titulo, h):
    """Título arriba, leyenda debajo de él, gráfica con aire suficiente."""
    fig.update_layout(
        height=h, template="plotly_dark",
        title=dict(text=titulo, x=0, xanchor="left", y=0.97, yanchor="top",
                   font=dict(size=20, color="#e2e8f0", family="DM Sans")) if titulo else None,
        margin=dict(l=10, r=10, t=96 if titulo else 40, b=10),
        legend=dict(orientation="h", y=1.0, yanchor="bottom", x=0, xanchor="left",
                    font=dict(color="#cbd5e1")),
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(15,23,42,0.35)",
        font=dict(color="#cbd5e1", family="DM Sans"),
        xaxis=dict(gridcolor="rgba(148,163,184,.12)", zerolinecolor="rgba(148,163,184,.2)"),
        yaxis=dict(gridcolor="rgba(148,163,184,.12)", zerolinecolor="rgba(148,163,184,.2)"))
    return fig

_CHART_N = {"i": 0}

def chart(fig, key=None):
    """st.plotly_chart con key única (evita StreamlitDuplicateElementId) y
    sin kwargs que Plotly interprete como config (evita el warning de deprecación).
    Nota: esta versión de Streamlit instalada NO tiene el parámetro `width` en
    plotly_chart (solo st.dataframe lo tiene) — usar use_container_width aquí."""
    if key is None:
        _CHART_N["i"] += 1
        key = f"chart_{_CHART_N['i']}"
    st.plotly_chart(fig, use_container_width=True, key=key)
    nota_filtro()

def _desc_filtros() -> str:
    """Texto legible del filtro global activo, p.ej.
    'Grupo de Cuentas: Renta, Luz · Cuenta: 2 seleccionadas'. '' si no hay filtro."""
    etiquetas = {"cat_Grupo_de_Cuentas": "Grupo de Cuentas", "cat_Cuentas": "Cuenta",
                 "cat_Subtipo": "Subtipo", "cat_Agrupa1": "Agrupa 1",
                 "cat_PDC": "PDC / Centro de Costos", "cat_PosPre_Full": "PosPre",
                 "cat_Direccion_Division": "División",
                 "cat_Subdireccion_Territorio": "Territorio", "Categoria": "Categoría"}
    partes = []
    for k, v in st.session_state.get("_filtros_vista", ()):
        if k not in etiquetas:      # ID_CONCEPTO_CUENTA_NIV3 es interno, no se muestra
            continue
        vals = list(v) if isinstance(v, tuple) else [v]
        detalle = ", ".join(str(x) for x in vals) if len(vals) <= 3 else f"{len(vals)} seleccionados"
        partes.append(f"{etiquetas[k]}: {detalle}")
    return " · ".join(partes)

def aviso_sin_subtipo(donde: str = "esta vista"):
    """Avisa cuando hay un Subtipo elegido pero la fuente de la vista no trae
    la columna (pdc / pdc_pospre) — el filtro se ignora y hay que decirlo."""
    if st.session_state.get("fg_subtipo"):
        st.warning(f"El filtro de **Subtipo** no aplica en {donde}: esta vista sale "
                   "de la fuente de Puntos de Contacto, que no incluye esa columna. "
                   "Los totales de aquí abarcan **todos** los subtipos.")

def nota_filtro():
    """Aviso bajo cada gráfica: si hay filtro activo, deja claro que lo que se
    ve es solo ese pedazo del universo (pedido del usuario 2026-07-27)."""
    txt = _desc_filtros()
    if txt:
        st.caption(f"⚠️ Vista filtrada — solo se está mostrando: **{txt}**. "
                   "Quita el filtro global del menú lateral para ver el total.")

PLANES = ("Plan_2026", "Nvo_Plan_2026", "Forecast_Cierre")
# 2026-07-27: series por defecto de las GRÁFICAS (barras/lineas) cuando el
# caller no especifica `series=` — a diferencia de MEDIDAS (motor de cálculo,
# no se toca), esta lista ya no incluye Nvo Plan porque se ocultó de todas
# las vistas en pantalla.
MEDIDAS_VISIBLES = ["Real_2025", "Plan_2026", "Real_2026"]

def barras(df: pd.DataFrame, xcol: str, series=None, titulo="", n=14, h=430):
    """Regla global del dashboard: los REALES van como barras (gris 2025, rojo
    claro 2026) y los PLANES como línea punteada encima (rojo oscuro plan viejo,
    azul nvo plan). Aplica en todas las secciones."""
    series = series or MEDIDAS_VISIBLES
    d = df.head(n)
    fig = go.Figure()
    for m in series:
        if m in PLANES:
            fig.add_scatter(name=NOMBRE.get(m, m), x=d[xcol], y=d[m], mode="lines+markers",
                            line=dict(color=COLOR.get(m), width=2.5, dash="dash"),
                            marker=dict(size=7), connectgaps=False,
                            hovertemplate="%{x}<br>%{y:,.0f}<extra></extra>")
        else:
            fig.add_bar(name=NOMBRE.get(m, m), x=d[xcol], y=d[m],
                        marker_color=COLOR.get(m), hovertemplate="%{x}<br>%{y:,.0f}<extra></extra>")
    fig.update_layout(barmode="group", hovermode="x unified")
    fig.update_xaxes(tickangle=-35)
    return _layout(fig, titulo, h)

def lineas(d: pd.DataFrame, series=None, titulo="", h=410, xcol="sem", xtitle="Semana"):
    """Reales (2025 o 2026) como barras; todo lo demás (Plan, Nvo Plan,
    Forecast) como línea punteada — regla fija en TODAS las pestañas."""
    series = series or MEDIDAS_VISIBLES
    fig = go.Figure()
    for m in series:
        if m in ("Real_2025", "Real_2026"):
            fig.add_bar(name=NOMBRE.get(m, m), x=d[xcol], y=d[m],
                        marker_color=COLOR.get(m), opacity=0.9,
                        hovertemplate="%{x}<br>%{y:,.0f}<extra></extra>")
        else:
            fig.add_scatter(name=NOMBRE.get(m, m), x=d[xcol], y=d[m], mode="lines+markers",
                            line=dict(color=COLOR.get(m), width=2.5, dash="dash"),
                            marker=dict(size=5), connectgaps=False,
                            hovertemplate="%{x}<br>%{y:,.0f}<extra></extra>")
    fig.update_layout(barmode="group", hovermode="x unified")
    # 2026-07-27: el usuario pidió ver TODAS las semanas/meses en el eje, no
    # solo los ticks espaciados que pone Plotly por defecto (0, 10, 20…).
    # tickmode="array" + tickvals=todo el eje fuerza una etiqueta por punto;
    # se achica y rota la letra para que quepan sin amontonarse.
    fig.update_xaxes(title=xtitle, tickmode="array", tickvals=d[xcol].tolist(),
                     tickangle=-45, tickfont=dict(size=10))
    return _layout(fig, titulo, h)

def _fila_total_tabla(df: pd.DataFrame, label_cols: list[str], orden: list[str]) -> dict:
    """Suma el DF crudo y arma dict fila TOTAL para anexar a fmt_tabla()."""
    r25 = df.get("Real_2025", pd.Series(dtype=float)).sum()
    p26 = df.get("Plan_2026", pd.Series(dtype=float)).sum()
    nvo = df.get("Nvo_Plan_2026", pd.Series(dtype=float)).sum()
    r26 = df.get("Real_2026", pd.Series(dtype=float)).sum()
    fc = df.get("Forecast_Cierre", pd.Series(dtype=float)).sum()
    vs_nvo, vs_plan, vs_aa, vs_fc = r26 - nvo, r26 - p26, r26 - r25, r26 - fc
    row = {c: "" for c in label_cols}
    row[label_cols[0]] = "TOTAL"
    row.update({"Real_2025": r25, "Pct_r25": 100.0,
                "Plan_2026": p26, "Pct_p26": 100.0,
                "Nvo_Plan_2026": nvo, "Pct_nvo": 100.0,
                "Real_2026": r26, "Pct_r26": 100.0,
                "Vs_NvoPlan_Abs": vs_nvo, "Vs_NvoPlan_Pct": (vs_nvo / abs(nvo) * 100) if nvo else 0.0,
                "Vs_Plan_Abs": vs_plan, "Vs_Plan_Pct": (vs_plan / abs(p26) * 100) if p26 else 0.0,
                "Vs_AA_Abs": vs_aa, "Vs_AA_Pct": (vs_aa / abs(r25) * 100) if r25 else 0.0,
                "Forecast_Cierre": fc,
                "Vs_Forecast_Abs": vs_fc, "Vs_Forecast_Pct": (vs_fc / abs(fc) * 100) if fc else 0.0,
                "Cumplimiento_Plan": (r26 / p26 * 100) if p26 else 0.0,
                "Pct_del_Total": 100.0})
    return row

# ORDEN OFICIAL DE COLUMNAS (2026-07-27, pedido del usuario) — el mismo en
# TODAS las tablas y TODOS los drill-downs, con o sin filtro:
#   Real 2025 | % del Total | Plan 2026 | % del Total | Nvo Plan | % del Total |
#   Real 2026 | % del Total | vs Nvo Plan | vs Nvo Plan % | vs Plan | vs Plan % |
#   vs AA | vs AA % | Forecast | vs Forecast | vs Forecast %
# Nota: 'Forecast' va DESPUÉS de vs AA % y ANTES de vs Forecast — por eso las
# vs de Forecast se excluyen de _VS_COMBINADAS y se anexan aparte.
_VS_COMBINADAS = [("Vs_NvoPlan_Abs", "Vs_NvoPlan_Pct", "vs Nvo Plan", "vs Nvo Plan %"),
                  ("Vs_Plan_Abs", "Vs_Plan_Pct", "vs Plan", "vs Plan %"),
                  ("Vs_AA_Abs", "Vs_AA_Pct", "vs AA", "vs AA %"),
                  ("Vs_Forecast_Abs", "Vs_Forecast_Pct", "vs Forecast", "vs Forecast %")]
# las 3 primeras siempre van; el bloque de Forecast solo si inc_fc
_VS_SIN_FC = _VS_COMBINADAS[:3]

# 2026-07-27: Nvo Plan se OCULTA de todas las vistas en pantalla (tablas y
# árboles) — sigue calculado en el motor (compute/_agregar_df/_medidas, no se
# toca) y sigue disponible completo en los CSV de descargar(), que trabaja
# sobre el DataFrame crudo, no sobre lo que se muestra. Es "solo para análisis
# posterior con el dashboard", como pidió el usuario.
_COLS_NVO_PLAN = ("Nvo_Plan_2026", "Pct_nvo", "Vs_NvoPlan_Abs", "Vs_NvoPlan_Pct")

def fmt_tabla(df: pd.DataFrame, label_cols: list[str], index_first: bool = False, total: bool = False, inc_fc: bool = False):
    """Pandas Styler con formato contable:
    Real 2025 | % del Total | Plan 2026 | % del Total | Nvo Plan | % del Total |
    Real 2026 (amarillo) | % del Total | vs Nvo Plan | vs Nvo Plan % | vs Plan |
    vs Plan % | vs AA | vs AA % | Forecast | vs Forecast | vs Forecast %.
    Nvo Plan y sus columnas derivadas NO se muestran (ver _COLS_NVO_PLAN)."""
    orden = ["Real_2025", "Pct_r25", "Plan_2026", "Pct_p26",
             "Nvo_Plan_2026", "Pct_nvo", "Real_2026", "Pct_r26"]
    # vs Nvo Plan / vs Plan / vs AA primero; luego (si aplica) Forecast y sus vs
    vs_cols = [c for par in _VS_SIN_FC for c in par[:2]]
    cols = label_cols + [c for c in orden if c in df.columns] + [c for c in vs_cols if c in df.columns]
    if inc_fc:
        if "Forecast_Cierre" in df.columns:
            cols.append("Forecast_Cierre")
        cols += [c for c in ("Vs_Forecast_Abs", "Vs_Forecast_Pct") if c in df.columns]
    cols = [c for c in cols if c not in _COLS_NVO_PLAN]
    d = df[[c for c in cols if c in df.columns]].copy()
    # el PDC siempre se muestra como "ID · Nombre"; si el caller ya aplicó
    # con_id() (Detalle PDC), no se vuelve a aplicar para no duplicar el ID.
    if "cat_PDC" in d:
        d["cat_PDC"] = d["cat_PDC"].map(
            lambda s: s if isinstance(s, str) and re.match(r"^\d+ · ", s) else con_id(s))
    fila_tot = None
    if total and not df.empty:
        fila_tot = {k: v for k, v in _fila_total_tabla(df, label_cols, orden).items() if k in d.columns}
        d = pd.concat([d, pd.DataFrame([fila_tot])], ignore_index=True)
    css = pd.DataFrame("", index=d.index, columns=d.columns)
    for m in ["Real_2025", "Plan_2026", "Nvo_Plan_2026", "Forecast_Cierre"]:
        if m in d:
            css[m] = d[m].map(_css_gasto)
            d[m] = d[m].map(MC)
    if "Real_2026" in d:
        css["Real_2026"] = d["Real_2026"].map(_css_gasto) + ";background-color:#fffde7;font-weight:600"
        d["Real_2026"] = d["Real_2026"].map(MC)
    if "Pct_r26" in d:
        css["Pct_r26"] = "background-color:#fffde7;color:#2b2b2b;"

    def _css_var_text(v):
        if v is None or pd.isna(v) or v == 0:
            return "color:#2b2b2b;"
        if v > 0:
            return "color:#78281f;font-weight:600;"
        return "color:#145a32;font-weight:600;"

    for abs_col, pct_col, _et_abs, _et_pct in _VS_COMBINADAS:
        if abs_col in d and pct_col in d:
            clr = d[abs_col].map(_css_var_text)
            css[abs_col] = clr
            css[pct_col] = clr
            d[abs_col] = d[abs_col].map(MC)
            d[pct_col] = d[pct_col].map(PC)
    for p in ("Pct_r25", "Pct_p26", "Pct_nvo", "Pct_r26", "Cumplimiento_Plan", "Pct_del_Total"):
        if p in d:
            d[p] = d[p].map(lambda v: f"{v:,.1f}%" if v is not None and not pd.isna(v) else "—")
    if fila_tot is not None:
        css.loc[d.index[-1], :] = (css.loc[d.index[-1], :]
                                   + ";background-color:#eef0f3;border-top:2px solid #b8bfc8;font-weight:700")
    ren = {c: DIMS[a][0] for a in DIMS for c in [DIMS[a][1]] if c}
    vs_ren = {abs_col: et_abs for abs_col, pct_col, et_abs, et_pct in _VS_COMBINADAS}
    vs_ren.update({pct_col: et_pct for abs_col, pct_col, et_abs, et_pct in _VS_COMBINADAS})
    ren.update({
        "Real_2025": "Real 2025", "Pct_r25": "% del Total",
        "Plan_2026": "Plan 2026", "Pct_p26": "% del Total ",
        "Nvo_Plan_2026": "Nvo Plan", "Pct_nvo": "% del Total  ",
        "Real_2026": "Real 2026", "Pct_r26": "% del Total   ",
        "Forecast_Cierre": "Forecast",
        **vs_ren
    })
    d = d.rename(columns=ren)
    css.columns = d.columns
    # NOTA (2026-07-24): st.dataframe (Streamlit 1.50) dibuja el header sobre
    # <canvas> vía glide-data-grid — un Styler.set_table_styles() aquí NUNCA
    # se aplica (solo serviría con st.table, que este proyecto no usa para
    # tablas con toolbar). El encabezado en negritas/carbón se controla desde
    # .streamlit/config.toml (textColor + dataframeHeaderBackgroundColor).
    return d.style.apply(lambda _: css, axis=None)

_ARBOL_CSS = """<style>
div[data-testid="stTable"] thead tr th,
.stTable thead tr th {
    font-weight: 700 !important;
    color: #e2e8f0 !important;
    font-size: 0.84rem !important;
}

.vt-tree-wrap {
    overflow-x: auto; overflow-y: auto; max-height: 640px;
    border: 1px solid rgba(148,163,184,.16);
    border-radius: 14px;
    background: rgba(15,23,42,.72);
    margin-bottom: 1rem;
    box-shadow: 0 12px 32px rgba(0,0,0,.28);
}
.vt-tree {
    display: table; width: 100%; border-collapse: collapse;
    font-size: 13.5px; font-family: 'DM Sans', system-ui, sans-serif;
    color: #e2e8f0;
}
.vt-row {
    display: grid;
    grid-template-columns: 340px repeat(10, minmax(118px, 1fr));
    min-width: 1520px; align-items: center; padding: 5px 0;
    border-bottom: 1px solid rgba(148,163,184,.08);
    background: transparent; transition: background 0.15s ease;
}
.vt-row.vt-fc {
    grid-template-columns: 340px repeat(13, minmax(118px, 1fr));
    min-width: 1856px;
}
.vt-row > span {
    text-align: right; white-space: nowrap; overflow: visible;
    padding: 2px 6px; border-left: 1px solid rgba(148,163,184,.06);
    font-variant-numeric: tabular-nums; color: #e2e8f0;
}
.vt-row > span:first-child {
    text-align: left; border-left: none; font-weight: 500;
    overflow: hidden; text-overflow: ellipsis;
    position: sticky; left: 0; z-index: 1;
    background: #0f172a; box-shadow: 2px 0 8px -2px rgba(0,0,0,.35);
}
.vt-head > span:first-child { background: #1e293b !important; z-index: 3; }
.vt-tot > span:first-child { background: #1e293b !important; }
.vt-nvl0 > span:first-child { background: #111827; }
details[open] > summary .vt-row > span:first-child,
.vt-row.vt-open > span:first-child { background: #0c4a6e !important; }
.vt-tree summary:hover .vt-row > span:first-child,
.vt-row:hover > span:first-child { background: #1e293b !important; }
.vt-head {
    position: sticky; top: 0; background: #1e293b !important;
    color: #7dd3fc !important; font-weight: 700 !important; z-index: 2;
    border-bottom: 2px solid rgba(56,189,248,.35) !important;
}
.vt-head > span {
    font-weight: 700 !important; text-align: center !important; color: #7dd3fc !important;
}
.vt-head > span:first-child { text-align: left !important; }
.vt-r26-head {
    background-color: rgba(56,189,248,.22) !important;
    color: #e0f2fe !important; font-weight: 700 !important;
}
.vt-r26 {
    background-color: rgba(56,189,248,.08) !important; font-weight: 600;
}
.vt-var-zero { color: #94a3b8; }
.vt-var-pos { color: #fb7185 !important; font-weight: 600 !important; background: transparent !important; }
.vt-var-neg { color: #34d399 !important; font-weight: 600 !important; background: transparent !important; }
.vt-pos { color: #fb7185; font-weight: 600; }
.vt-neg { color: #e2e8f0; }

.vt-tot {
    background: #1e293b !important; font-weight: 700 !important;
    border-top: 2px solid rgba(56,189,248,.3) !important;
    position: sticky; bottom: 0;
}

.vt-nvl0 { background: rgba(15,23,42,.5); font-size: 14px; }
.vt-nvl0 > span:first-child { font-weight: 700 !important; color: #f8fafc; }
.vt-nvl1 > span:first-child { font-weight: 600 !important; color: #e2e8f0; }
.vt-nvl2 > span:first-child { font-weight: 600; color: #cbd5e1; }
.vt-nvl3 > span:first-child { font-weight: 500; color: #94a3b8; }
.vt-nvl4 > span:first-child { font-weight: 500; color: #94a3b8; }
.vt-nvl5 > span:first-child { font-weight: 400; color: #64748b; }

details[open] > summary .vt-row,
.vt-row.vt-open { background: rgba(14,116,144,.25) !important; }
details[open] > summary .vt-row > span:first-child,
.vt-row.vt-open > span:first-child {
    font-weight: 700 !important; color: #7dd3fc !important;
}
.vt-tree summary:hover .vt-row,
.vt-row:hover { background: rgba(30,41,59,.7) !important; }

.vt-tree summary { list-style:none; cursor:pointer; }
.vt-tree summary::-webkit-details-marker { display:none; }
.vt-caret::before {
    content:'+'; display:inline-flex; align-items:center; justify-content:center;
    width:18px; height:18px; margin-right:7px; vertical-align:-4px;
    border:1px solid rgba(56,189,248,.35); border-radius:6px;
    background:rgba(15,23,42,.9);
    font-size:13px; font-weight:800; color:#7dd3fc; line-height:1;
}
details[open] > summary .vt-caret::before { content:'−'; }
.vt-tree summary:hover .vt-caret::before {
    background:#38bdf8; border-color:#38bdf8; color:#0f172a;
}
.vt-hoja::before { content:''; display:inline-block; width:18px; margin-right:7px; }
.vt-tab { display:inline-block; width:12px; height:16px; position:relative; vertical-align:-3px; }
.vt-tab::before { content:''; position:absolute; left:5px; top:0; bottom:0; border-left:1.5px solid rgba(148,163,184,.25); }
.vt-tab:last-child::after { content:''; position:absolute; left:5px; top:50%; width:6px; border-top:1.5px solid rgba(148,163,184,.25); }

[class*="st-key-_btn_cerrado_"] button, [class*="st-key-_btn_abierto_"] button {
    border:1px solid rgba(56,189,248,.35) !important;
    background:rgba(15,23,42,.9) !important;
    border-radius:6px !important; padding:0 !important; min-height:0 !important;
    width:26px !important; height:26px !important; min-width:26px !important;
    position:relative; overflow:hidden;
}
[class*="st-key-_btn_cerrado_"] button p, [class*="st-key-_btn_abierto_"] button p,
[class*="st-key-_btn_cerrado_"] button div, [class*="st-key-_btn_abierto_"] button div {
    display:none !important;
}
[class*="st-key-_btn_cerrado_"] button::after,
[class*="st-key-_btn_abierto_"] button::after {
    position:absolute; inset:0; display:flex; align-items:center;
    justify-content:center; font-size:17px; font-weight:800;
    color:#7dd3fc; line-height:1;
}
[class*="st-key-_btn_cerrado_"] button::after { content:'+'; }
[class*="st-key-_btn_abierto_"] button::after { content:'−'; }
[class*="st-key-_btn_cerrado_"] button:hover, [class*="st-key-_btn_abierto_"] button:hover {
    background:#38bdf8 !important; border-color:#38bdf8 !important;
}
[class*="st-key-_btn_cerrado_"] button:hover::after,
[class*="st-key-_btn_abierto_"] button:hover::after { color:#0f172a; }
div[data-testid="stHorizontalBlock"]:has(.vt-row) { margin-top:-0.35rem; align-items:center; }
div[data-testid="stHorizontalBlock"]:has(.vt-row) div[data-testid="stElementContainer"] { margin:0; }

</style>"""

def _get_arbol_head(inc_fc: bool = False) -> str:
    # mismo ORDEN OFICIAL que fmt_tabla (ver _VS_COMBINADAS). Nvo Plan y vs Nvo
    # Plan/% se OCULTAN en pantalla (2026-07-27, ver _COLS_NVO_PLAN) — el
    # árbol perezoso y arbol_multicol siguen sin llamarlos, así que basta con
    # no listarlos aquí y en _fila_arbol()/_fila_total_arbol().
    headers = [
        "Real 2025", "% del Total", "Plan 2026", "% del Total",
        ("Real 2026", True), ("% del Total", True),
        "vs Plan", "vs Plan %", "vs AA", "vs AA %"
    ]
    if inc_fc:
        headers.extend(["Forecast", "vs Forecast", "vs Forecast %"])
    
    spans = []
    for item in headers:
        if isinstance(item, tuple):
            h, _ = item
            spans.append(f"<span class='vt-r26-head'>{h}</span>")
        else:
            spans.append(f"<span>{item}</span>")
    
    return f"<div class='vt-row {'vt-fc' if inc_fc else ''} vt-head'><span>Concepto</span>" + "".join(spans) + "</div>"

def _sp(v, fmt=MC) -> str:
    cls = "vt-pos" if (v is not None and not pd.isna(v) and v > 0) else "vt-neg"
    return f"<span class='{cls}'>{fmt(v)}</span>"

def _spv(monto, pct) -> str:
    if monto is None or pd.isna(monto) or monto == 0:
        cls = "vt-var-zero"
    elif monto > 0:
        cls = "vt-var-pos"
    else:
        cls = "vt-var-neg"
    return f"<span class='{cls}'>{MC(monto)}</span><span class='{cls}'>{PC(pct)}</span>"

def _fila_arbol(r, col: str, nivel: int, hoja: bool, sin_caret: bool = False, inc_fc: bool = False, abierto: bool = False) -> str:
    import html as _h
    caret = "" if sin_caret else ("vt-hoja" if hoja else "vt-caret")
    etiqueta = con_id(r[col]) if col == "cat_PDC" else r[col]
    pad = 6
    tabs = "<span class='vt-tab'></span>" * nivel
    pct_r25 = r.get("Pct_r25", 0.0)
    pct_p26 = r.get("Pct_p26", 0.0)
    pct_r26 = r.get("Pct_r26", 0.0)

    r26_val = r.get("Real_2026", 0.0)
    cls_r26 = "vt-pos" if (r26_val is not None and not pd.isna(r26_val) and r26_val > 0) else "vt-neg"
    cls_open = " vt-open" if abierto else ""

    # Nvo Plan y vs Nvo Plan/% OCULTOS en pantalla (2026-07-27) — se siguen
    # calculando en r (ver _COLS_NVO_PLAN) y salen completos en el CSV.
    html = (
        f"<div class='vt-row {'vt-fc' if inc_fc else ''} vt-nvl{nivel}{cls_open}'>"
        f"<span style='padding-left:{pad}px'>"
        f"{tabs}<span class='{caret}'></span>{_h.escape(str(etiqueta))}</span>"
        f"{_sp(r['Real_2025'])}<span>{pct_r25:.1f}%</span>"
        f"{_sp(r['Plan_2026'])}<span>{pct_p26:.1f}%</span>"
        f"<span class='vt-r26 {cls_r26}'>{MC(r26_val)}</span><span class='vt-r26 {cls_r26}'>{pct_r26:.1f}%</span>"
        f"{_spv(r['Vs_Plan_Abs'], r['Vs_Plan_Pct'])}"
        f"{_spv(r['Vs_AA_Abs'], r['Vs_AA_Pct'])}"
    )
    if inc_fc:
        fc_val = r.get('Forecast_Cierre', r26_val)
        vs_fc_abs = r.get('Vs_Forecast_Abs', 0.0)
        vs_fc_pct = r.get('Vs_Forecast_Pct', 0.0)
        html += f"{_sp(fc_val)}{_spv(vs_fc_abs, vs_fc_pct)}"
    html += "</div>"
    return html

def _fila_total_arbol(df0: pd.DataFrame, inc_fc: bool = False) -> str:
    if df0.empty:
        return ""
    r25, p26, r26 = (df0["Real_2025"].sum(), df0["Plan_2026"].sum(), df0["Real_2026"].sum())
    vs_plan, vs_aa = r26 - p26, r26 - r25
    pct_plan = (vs_plan / abs(p26) * 100) if p26 else 0.0
    pct_aa = (vs_aa / abs(r25) * 100) if r25 else 0.0

    cls_r26 = "vt-pos" if r26 > 0 else "vt-neg"

    # Nvo Plan OCULTO en pantalla (2026-07-27) — ver _COLS_NVO_PLAN.
    tot_html = (
        f"<div class='vt-row {'vt-fc' if inc_fc else ''} vt-tot'><span>TOTAL</span>"
        f"{_sp(r25)}<span>100.0%</span>"
        f"{_sp(p26)}<span>100.0%</span>"
        f"<span class='vt-r26 {cls_r26}'>{MC(r26)}</span><span class='vt-r26 {cls_r26}'>100.0%</span>"
        f"{_spv(vs_plan, pct_plan)}"
        f"{_spv(vs_aa, pct_aa)}"
    )
    if inc_fc:
        fc = df0["Forecast_Cierre"].sum() if "Forecast_Cierre" in df0 else r26
        vs_fc = r26 - fc
        pct_fc = (vs_fc / abs(fc) * 100) if fc else 0.0
        tot_html += f"{_sp(fc)}{_spv(vs_fc, pct_fc)}"
    tot_html += "</div>"
    return tot_html

@st.cache_data(show_spinner=False)
def _nivel_arbol_perezoso(fuente: str, cols: tuple, prof: int, valores_padre: tuple,
                          si: int, sf: int, ind: bool, filtros: tuple = (),
                          mes=None, sems: tuple = None) -> pd.DataFrame:
    df = AGG.get(fuente) if fuente in AGG else EXTRA.get(fuente)
    if df is None:
        fp = AGGS / f"{fuente}.parquet"
        if not fp.exists():
            return pd.DataFrame()
        df = pl.scan_parquet(fp)
    for k, v in filtros:
        if k in df.columns:
            df = df.filter(pl.col(k).is_in(list(v)) if isinstance(v, tuple) else pl.col(k) == v)
    for c, val in zip(cols[:prof], valores_padre):
        if c in df.columns:
            df = df.filter(pl.col(c) == val)
    if isinstance(df, pl.LazyFrame):
        df = df.collect()
    return _agregar_df(df, [cols[prof]], si, sf, ind, mes, sems)

def _toggle_rama(k: str) -> None:
    st.session_state[k] = not st.session_state.get(k, False)

def _arbol_html(niveles: list, inc_fc: bool = False) -> str:
    nodos = [None] * len(niveles)
    for i in range(1, len(niveles)):
        df, _ = niveles[i]
        pcols = [niveles[j][1] for j in range(i)]
        d = {}
        for _, r in df.iterrows():
            d.setdefault(tuple(r[c] for c in pcols), []).append(r)
        nodos[i] = d

    def rec(i, clave_padre):
        df, col = niveles[i]
        filas = ([r for _, r in df.iterrows()] if i == 0
                 else nodos[i].get(clave_padre, []))
        out = []
        for r in filas:
            clave = clave_padre + (r[col],)
            hijos = i + 1 < len(niveles) and clave in nodos[i + 1]
            if hijos:
                out.append(f"<details><summary>{_fila_arbol(r, col, i, False, inc_fc=inc_fc)}</summary>"
                           f"{rec(i + 1, clave)}</details>")
            else:
                out.append(_fila_arbol(r, col, i, True, inc_fc=inc_fc))
        return "".join(out)

    total = _fila_total_arbol(niveles[0][0], inc_fc=inc_fc) if niveles else ""
    head = _get_arbol_head(inc_fc=inc_fc)
    return _ARBOL_CSS + "<div class='vt-tree-wrap'><div class='vt-tree'>" + head + rec(0, ()) + total + "</div></div>"

@st.cache_data(show_spinner="Armando árbol…")
def arbol_jerarquia(hier: tuple, si: int, sf: int, ind: bool,
                    filtros: tuple = (), mes=None, sems: tuple = None,
                    inc_fc: bool = False, sin_padres: tuple = ()) -> str:
    """`sin_padres`: columnas de PARENTS que NO son niveles visibles de este
    árbol y por tanto no deben partir la agregación. Sin esto, quitar una rama
    de la jerarquía (p.ej. Agrupa 1 en "Por Cuenta") dejaría el agrupado por
    esa columna y un mismo Grupo/Cuenta saldría duplicado en varias filas."""
    niveles = [(compute(a, si, sf, ind, filtros, mes=mes, sems=sems,
                        excluir_padres=sin_padres), DIMS[a][1])
               for a in hier]
    return _arbol_html(niveles, inc_fc=inc_fc)

def _idtxt_pdc(p: str) -> str:
    """nombre de PDC -> su Eco PDV (id) como texto, o '' si no está en catálogo."""
    x = PDC_IDS.get(p)
    if x is None:
        return ""
    if isinstance(x, float) and x.is_integer():
        return str(int(x))
    return str(x).strip()

def buscador_pdc_ceco(key_prefix: str, label: str = "Buscar PDC o ID Centro de Costos") -> tuple | None:
    """Buscador único por nombre/ID de PDC y por ID_CENTRO_COSTOS, con selección
    múltiple. Devuelve una tupla de cat_PDC para usar como filtro
    (("cat_PDC", tuple(...)),) o None si el usuario no escribió nada (sin filtro).
    Si escribió pero no hay resultados, ya emite el warning y detiene la pestaña."""
    import hashlib as _hashlib

    q = st.text_input(
        label,
        placeholder="nombre o ID del PDC, o ID_CENTRO_COSTOS…",
        key=f"{key_prefix}_q"
    ).strip()

    if not q:
        return None

    texto_busqueda = q.casefold()
    lista_pdc = sorted({p for p in _CECO_A_PDC.values()} | set(PDC_IDS.keys()))

    _por_pdc = {
        p for p in lista_pdc
        if texto_busqueda in str(p).casefold() or texto_busqueda in _idtxt_pdc(p).casefold()
    }
    _por_ceco = {_CECO_A_PDC[c] for c in _CECO_A_PDC if texto_busqueda in c.casefold()}
    encontrados = sorted(_por_pdc | _por_ceco,
                         key=lambda p: (str(p).casefold(), _idtxt_pdc(p).casefold()))

    if not encontrados:
        st.warning(f"No se encontraron PDC/CECO que coincidan con «{q}».")
        st.stop()

    clave = _hashlib.md5(texto_busqueda.encode("utf-8")).hexdigest()[:10]
    seleccionados = st.multiselect(
        "Selecciona los PDC/CECO que quieres incluir",
        options=encontrados,
        default=encontrados,
        format_func=lambda p: f"{_idtxt_pdc(p)} · {p}" if _idtxt_pdc(p) else str(p),
        key=f"{key_prefix}_matches_{clave}",
        placeholder="Selecciona uno o varios PDC/CECO"
    )
    st.caption(f"Se encontraron **{len(encontrados):,}** coincidencias para «{q}» "
              f"y seleccionaste **{len(seleccionados):,}**.")

    if not seleccionados:
        st.info("Selecciona al menos un PDC/CECO para mostrar la información.")
        st.stop()

    return (("cat_PDC", tuple(seleccionados)),)

def arbol_perezoso(fuente: str, cols: tuple, si: int, sf: int, ind: bool,
                   filtros: tuple = (), mes=None, sems: tuple = None, key: str = "arbol",
                   permitir_expandir: bool = False, inc_fc: bool = False) -> None:
    st.markdown(_ARBOL_CSS, unsafe_allow_html=True)
    raiz = _nivel_arbol_perezoso(fuente, cols, 0, (), si, sf, ind, filtros, mes, sems)
    if raiz.empty:
        st.info("Sin datos para esta selección.")
        return

    def fila(html_str: str):
        c1, c2 = st.columns([2, 40])
        with c2:
            st.markdown(html_str, unsafe_allow_html=True)
        return c1

    if permitir_expandir:
        _raiz_keys = [f"_abierto_{key}__{v}__{i}" for i, v in enumerate(raiz[cols[0]])]
        _algo_abierto = any(st.session_state.get(k, False) for k in _raiz_keys)

        def _set_raiz(valor: bool):
            for k in _raiz_keys:
                st.session_state[k] = valor

        b1, _ = st.columns([1, 5])
        with b1:
            if _algo_abierto:
                st.button("Colapsar todo", key=f"_colapsar_{key}",
                          on_click=_set_raiz, args=(False,))
            else:
                st.button("Expandir todo (1er nivel)", key=f"_expandir_{key}",
                          on_click=_set_raiz, args=(True,))

    # nota de uso: el control es un cuadro sin texto, así que se explica aquí
    # (2026-07-27, pedido del usuario). Aplica a todos los drill-downs.
    st.caption("Usa el botón **+** de la izquierda para desglosar una fila, "
               "y **−** para cerrarla.")

    st.markdown("<div class='vt-tree-wrap'><div class='vt-tree'>", unsafe_allow_html=True)
    fila(_get_arbol_head(inc_fc=inc_fc))

    def render_nivel(prof: int, valores_padre: tuple, df_nivel: pd.DataFrame):
        for i, (_, r) in enumerate(df_nivel.iterrows()):
            valores = valores_padre + (r[cols[prof]],)
            es_hoja = prof + 1 >= len(cols)
            if es_hoja:
                fila(_fila_arbol(r, cols[prof], prof, True, inc_fc=inc_fc))
                continue
            rama_key = f"{key}__{'__'.join(str(v) for v in valores)}__{i}"
            estado_key = f"_abierto_{rama_key}"
            abierto = st.session_state.get(estado_key, False)
            c1, c2 = st.columns([2, 40])
            with c1:
                # 2026-07-27: el botón va SIN TEXTO — la columna es angosta y
                # cualquier label se comprimía hasta volverse ilegible. Es un
                # cuadro de tamaño fijo y el signo +/− se dibuja con CSS
                # (::after en _ARBOL_CSS). Streamlit expone la key del widget
                # como .st-key-<key> en el DOM, así que el estado abierto se
                # distingue por el sufijo de la key, sin envolver nada.
                st.button("", key=f"_btn_{'abierto' if abierto else 'cerrado'}_{rama_key}",
                          help="Cerrar el desglose" if abierto else "Abrir el desglose",
                          on_click=_toggle_rama, args=(estado_key,))
            with c2:
                st.markdown(_fila_arbol(r, cols[prof], prof, False, sin_caret=True, inc_fc=inc_fc, abierto=abierto),
                            unsafe_allow_html=True)
            if abierto:
                hijos = _nivel_arbol_perezoso(fuente, cols, prof + 1, valores,
                                              si, sf, ind, filtros, mes, sems)
                if hijos.empty:
                    st.caption("Sin desglose adicional.")
                else:
                    render_nivel(prof + 1, valores, hijos)

    render_nivel(0, (), raiz)
    fila(_fila_total_arbol(raiz, inc_fc=inc_fc))
    st.markdown("</div></div>", unsafe_allow_html=True)

def estilizar(d: pd.DataFrame, money: list, pct: list = (), fila_total: dict = None):
    """Formato contable sobre un DF con nombres finales de columna.
    fila_total: dict {columna: valor_crudo} a anexar como fila resaltada al
    final (el caller decide qué sumar, ya que aquí no se conoce la semántica
    de cada tabla — p.ej. algunas tienen columna de texto como 1a col)."""
    d = d.copy()
    if fila_total is not None:
        d = pd.concat([d, pd.DataFrame([fila_total])], ignore_index=True)
    css = pd.DataFrame("", index=d.index, columns=d.columns)
    for c in money:
        if c in d.columns:
            css[c] = d[c].map(_css_gasto)
            d[c] = d[c].map(MC)
    for c in pct:
        if c in d.columns:
            css[c] = d[c].map(_css_gasto)
            d[c] = d[c].map(PC)
    if fila_total is not None:
        css.loc[d.index[-1], :] = css.loc[d.index[-1], :] + ";background-color:#eef0f3;font-weight:700"
    return d.style.apply(lambda _: css, axis=None)

_DL_N = {"i": 0}

def descargar(df: pd.DataFrame, nombre: str):
    """Botón de descarga con números completos (sin resumir)."""
    _DL_N["i"] += 1
    d_exp = df.copy()
    if "cat_PDC" in d_exp.columns:
        d_exp["cat_PDC"] = d_exp["cat_PDC"].map(con_id)
    csv_bytes = d_exp.to_csv(index=False, float_format="%.2f").encode("utf-8-sig")
    st.download_button("Descargar CSV (datos completos)", csv_bytes,
                       f"{nombre}.csv", "text/csv", type="secondary",
                       key=f"dl_{_DL_N['i']}")

def tabla(df: pd.DataFrame, label_cols: list[str], nombre: str = "", height=None, total: bool = True, inc_fc: bool = False):
    kw = {"height": height} if height else {}
    sty = fmt_tabla(df, label_cols, total=total, inc_fc=inc_fc)
    # 2026-07-24: st.dataframe (canvas) recorta el texto por su cuenta según
    # el ancho de columna que calcula solo — ningún CSS lo toca (ver nota en
    # fmt_tabla). column_config con ancho fijo evita el corte sin agrandar
    # la tabla entera (width='stretch', NO 'content' — ese la hacía enorme).
    n_label = sum(1 for c in label_cols if c in df.columns)
    col_cfg = {c: st.column_config.TextColumn(width=125)
              for i, c in enumerate(sty.data.columns) if i >= n_label}
    # 2026-07-27: la primera columna (el concepto/nombre) arranca fija al
    # deslizar horizontalmente — el usuario puede desfijarla desde el propio
    # menú de la columna en la tabla (clic derecho en el encabezado).
    if sty.data.columns.size:
        col_cfg[sty.data.columns[0]] = st.column_config.TextColumn(pinned=True)
    st.dataframe(sty, width='stretch', hide_index=True, column_config=col_cfg, **kw)

# =========================================================== KPIs de Gasto
# 2026-07-27: los KPIs de gasto llevan la variación DEBAJO, dentro de la misma
# tarjeta, como señal (flecha + monto + porcentaje) — no en un cuadro aparte.
# Convención contable de siempre: gasto por ARRIBA de la referencia = rojo con
# flecha ▲; por DEBAJO (ahorro) = verde con flecha ▼.
_KPI_CSS = """<style>
.vt-kpis { display:grid; grid-template-columns:repeat(auto-fit,minmax(210px,1fr));
           gap:12px; margin:4px 0 6px 0; }
.vt-kpi { border:1px solid #dcdfe4; border-radius:10px; background:#fff;
          padding:12px 14px; box-shadow:0 1px 3px rgba(0,0,0,.03); }
.vt-kpi-tit { font-size:.72rem; letter-spacing:.08em; text-transform:uppercase;
              color:#5b6470; font-weight:700; margin-bottom:4px; }
.vt-kpi-val { font-size:1.55rem; font-weight:800; color:#2b2b2b;
              line-height:1.15; font-variant-numeric:tabular-nums; }
.vt-kpi-sig { margin-top:8px; padding-top:7px; border-top:1px solid #eef0f4;
              font-size:.86rem; font-weight:700; font-variant-numeric:tabular-nums; }
.vt-kpi-sig .vt-kpi-et { display:block; font-size:.68rem; font-weight:600;
              color:#7a828d; text-transform:uppercase; letter-spacing:.05em;
              margin-bottom:2px; }
.vt-sig-alza { color:#fb7185; }
.vt-sig-baja { color:#34d399; }
.vt-sig-cero { color:#94a3b8; }
</style>"""

def _senal(etiqueta: str, monto, pct) -> str:
    """Bloque de señal: flecha + monto en MDP + porcentaje. Gasto arriba de la
    referencia (monto > 0) = rojo ▲; abajo = verde ▼."""
    if monto is None or pd.isna(monto) or monto == 0:
        cls, flecha = "vt-sig-cero", "="
    elif monto > 0:
        cls, flecha = "vt-sig-alza", "▲"
    else:
        cls, flecha = "vt-sig-baja", "▼"
    p = "—" if pct is None or pd.isna(pct) else f"{abs(pct):,.1f}%"
    return (f"<div class='vt-kpi-sig {cls}'><span class='vt-kpi-et'>{etiqueta}</span>"
            f"{flecha} {M(abs(monto) if monto is not None and not pd.isna(monto) else monto)}"
            f" &nbsp;·&nbsp; {p}</div>")

# --- KPIs como TABLA estilizada (2026-07-27): el usuario pidió no ver los
# indicadores como cuadritos sueltos y separados sino juntos, en una sola
# tabla compacta. Una fila de etiquetas sobre una fila de valores.
_KPITBL_CSS = """<style>
.vt-kpitbl { width:100%; border-collapse:collapse; margin:2px 0 12px 0;
             border:1px solid #dcdfe4; border-radius:10px; overflow:hidden;
             font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif; }
.vt-kpitbl th { background:#f1f3f6; color:#111827; font-weight:800;
                font-size:.70rem; letter-spacing:.06em; text-transform:uppercase;
                padding:7px 12px; text-align:center; border-right:1px solid #e3e6ea;
                border-bottom:2px solid #cfd4dc; white-space:nowrap; }
.vt-kpitbl td { padding:9px 12px; text-align:center; border-right:1px solid #eef0f4;
                background:#fff; font-variant-numeric:tabular-nums; }
.vt-kpitbl th:last-child, .vt-kpitbl td:last-child { border-right:none; }
.vt-kpitbl .vt-kv { font-size:1.15rem; font-weight:800; color:#2b2b2b;
                    line-height:1.2; display:block; }
.vt-kpitbl .vt-ks { font-size:.74rem; font-weight:600; color:#5b6470;
                    display:block; margin-top:2px; }
.vt-kpitbl .vt-kv.alza, .vt-kpitbl .vt-ks.alza { color:#fb7185; }
.vt-kpitbl .vt-kv.baja, .vt-kpitbl .vt-ks.baja { color:#34d399; }
</style>"""

def kpis_tabla(items: list) -> str:
    """items: lista de (etiqueta, valor, subtexto, tono) donde tono es
    '' | 'alza' | 'baja'. Devuelve el HTML de una tabla compacta de KPIs."""
    ths = "".join(f"<th>{et}</th>" for et, _v, _s, _t in items)
    tds = ""
    for _et, val, sub, tono in items:
        t = f" {tono}" if tono else ""
        sub_html = f"<span class='vt-ks{t}'>{sub}</span>" if sub else ""
        tds += f"<td><span class='vt-kv{t}'>{val}</span>{sub_html}</td>"
    return (_KPITBL_CSS + "<table class='vt-kpitbl'><thead><tr>" + ths +
            "</tr></thead><tbody><tr>" + tds + "</tr></tbody></table>")

def _kpi(titulo: str, valor: str, senal_html: str = "") -> str:
    return (f"<div class='vt-kpi'><div class='vt-kpi-tit'>{titulo}</div>"
            f"<div class='vt-kpi-val'>{valor}</div>{senal_html}</div>")

def kpis_gasto_html(x) -> str:
    """KPIs de gasto (Real 2026 / Real 2025 / Plan 2026 / Forecast Cierre),
    cada uno con su variación en monto y en % debajo. `x` es la fila de
    compute('global'). 2026-07-27: se quitó la tarjeta de "Nvo Plan" — ya no
    se muestra en ninguna vista, solo se descarga (ver _COLS_NVO_PLAN)."""
    tarjetas = [
        _kpi("Real 2026", M(x.Real_2026)),
        _kpi("Real 2025", M(x.Real_2025),
             _senal("Real 2026 vs AA", x.Vs_AA_Abs, x.Vs_AA_Pct)),
        _kpi("Plan 2026", M(x.Plan_2026),
             _senal("Real 2026 vs Plan", x.Vs_Plan_Abs, x.Vs_Plan_Pct)),
        _kpi("Forecast Cierre", M(x.Forecast_Cierre),
             _senal("Forecast vs Plan", x.Forecast_Cierre - x.Plan_2026,
                    ((x.Forecast_Cierre - x.Plan_2026) / abs(x.Plan_2026) * 100)
                    if x.Plan_2026 else 0.0)),
    ]
    return _KPI_CSS + "<div class='vt-kpis'>" + "".join(tarjetas) + "</div>"

# =========================================================== tablas de variaciones
# Vienen del Word 'Riesgos y Oportunidades.docx' que Planeación edita y que
# build_data.py convierte a aggs/riesgos.json. Dentro de cada lista, el propio
# Word trae separadores literales "— Variaciones Negativas vs AA (YTD) —" que
# se parsean con regex para clasificar cada nota en una de las 4 tablas
# (positiva/negativa × AA/Plan). Estaban en la pestaña "Riesgos & Oport."
# (eliminada 2026-07-27); ahora se muestran al final del Resumen.
_NOTA_RE = re.compile(
    r"^\d+\.\s*(?P<cuenta>.+?)\s+(?P<monto>\(?[\d.,]+\s*mdp\)?)\s*/\s*"
    r"(?P<pct>\(?[\d.,]+%\)?)\s*(?:·\s*(?P<com>.*))?$")

def _split_subsecciones(lista: list) -> dict:
    bloques: dict = {}
    actual = None
    for t_ in lista:
        m = re.match(r"^[—-]+\s*(.+?)\s*[—-]+$", t_.strip())
        if m:
            actual = m.group(1).strip()
            bloques.setdefault(actual, [])
        elif actual:
            bloques.setdefault(actual, []).append(t_)
        else:
            bloques.setdefault("(sin sección)", []).append(t_)
    return bloques

def _parsear_notas(notas: list) -> pd.DataFrame:
    filas = []
    for t_ in notas:
        m = _NOTA_RE.match(t_.strip())
        if m:
            filas.append({"Cuenta": m.group("cuenta").strip(),
                          "Monto": m.group("monto").replace(" ", ""),
                          "%": m.group("pct"),
                          "Comentario": (m.group("com") or "").strip()})
        else:
            filas.append({"Cuenta": t_.strip(), "Monto": "", "%": "", "Comentario": ""})
    return pd.DataFrame(filas)

def _leer_riesgos() -> dict:
    rj = AGGS / "riesgos.json"
    if not rj.exists():
        return {}
    try:
        return json.loads(rj.read_text(encoding="utf-8"))
    except Exception:
        return {}

def tablas_variaciones() -> None:
    """Las 4 tablas de variaciones (positivas/negativas × vs AA/vs Plan)."""
    rdoc = _leer_riesgos()
    if not (rdoc.get("riesgos") or rdoc.get("oportunidades")):
        st.caption("Para agregar los comentarios de variaciones: edita "
                   "`Riesgos y Oportunidades.docx` en la carpeta del dashboard "
                   "y corre `python build_data.py --aggs-only`.")
        return

    st.subheader("Variaciones · Riesgos y Oportunidades")
    st.caption(f"Comentarios de Planeación (Word) · actualizado {rdoc.get('actualizado', '—')}")
    neg = _split_subsecciones(rdoc.get("riesgos", []))
    pos = _split_subsecciones(rdoc.get("oportunidades", []))
    ROW_H, HEAD_H = 34, 45

    def _tabla_notas(titulo: str, notas: list, positivo: bool, alto: int):
        st.markdown(f"**{titulo}**")
        if not notas:
            st.caption("Sin comentarios.")
            return
        d = _parsear_notas(notas)
        bg, fg = ("rgba(16,185,129,.12)", "#34d399") if positivo else ("rgba(244,63,94,.12)", "#fb7185")
        css = pd.DataFrame("", index=d.index, columns=d.columns)
        css["Cuenta"] = f"font-weight:600;color:{fg}"
        css[["Monto", "%"]] = f"color:{fg};font-weight:600"
        sty = (d.style.apply(lambda _: css, axis=None)
                .set_properties(**{"background-color": bg}, subset=["Cuenta", "Monto", "%"]))
        st.dataframe(sty, width='stretch', hide_index=True, height=alto)

    def _bloque(clave: str, encabezado: str):
        st.markdown(f"##### {encabezado}")
        notas_pos = next((v for k, v in pos.items() if clave in k), [])
        notas_neg = next((v for k, v in neg.items() if clave in k), [])
        # misma altura para ambas tablas del par (scroll interno si excede)
        n_max = max(len(notas_pos), len(notas_neg), 1)
        alto = min(HEAD_H + ROW_H * n_max, 430)
        ca, cb = st.columns(2)
        with ca:
            _tabla_notas("Variaciones positivas", notas_pos, True, alto)
        with cb:
            _tabla_notas("Variaciones negativas", notas_neg, False, alto)

    _bloque("vs AA", "Variaciones vs Año Anterior (YTD)")
    st.divider()
    _bloque("vs Plan", "Variaciones vs Plan (YTD)")

# =========================================================== PDF del Resumen
# 2026-07-27: el usuario pidió poder descargar TODA la pestaña de Resumen.
# Se genera un HTML autocontenido (logo embebido en base64, CSS inline, sin
# recursos externos) que abre el diálogo de impresión al cargarse: el usuario
# elige "Guardar como PDF". Se reutilizan las mismas funciones de formato del
# dashboard (kpis_gasto_html, fmt_tabla, _arbol_html) para que el documento
# salga con el orden de columnas, la paleta y los millones idénticos.
_PDF_CSS = """
@page { size: A4 landscape; margin: 12mm; }
body { font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,sans-serif;
       color:#2b2b2b; margin:0; padding:14px 18px; }
h1 { font-size:1.5rem; margin:0 0 2px 0; color:#2b2b2b; }
h2 { font-size:1.05rem; margin:20px 0 8px 0; color:#2b2b2b;
     border-bottom:2px solid rgba(56,189,248,.35); padding-bottom:4px; }
h3 { font-size:.9rem; margin:12px 0 6px 0; color:#5b6470; }
.enc { display:flex; align-items:center; gap:16px;
       border-bottom:3px solid rgba(56,189,248,.35); padding-bottom:10px; margin-bottom:6px; }
.enc img { height:58px; }
.kicker { font-size:.72rem; letter-spacing:.14em; font-weight:700; color:#0284c7; }
.sub { font-size:.85rem; color:#5b6470; }
table.dat { border-collapse:collapse; width:100%; font-size:.66rem; margin-bottom:10px; }
table.dat th { background:#f1f3f6; color:#111827; font-weight:800; padding:5px 6px;
               border:1px solid #dcdfe4; text-align:center; white-space:nowrap; }
table.dat td { padding:4px 6px; border:1px solid #eef0f4; text-align:right;
               white-space:nowrap; font-variant-numeric:tabular-nums; }
table.dat td:first-child, table.dat th:first-child { text-align:left; }
.vt-tree-wrap { overflow:visible !important; }
.vt-row { min-width:0 !important; font-size:.62rem; }
.nofilt { font-size:.72rem; color:#7a828d; margin:4px 0 10px 0; }
.pie { margin-top:16px; font-size:.66rem; color:#8a93a0;
       border-top:1px solid #e3e6ea; padding-top:6px; }
@media print { .noprint { display:none !important; } }
"""

def _tabla_html_simple(sty) -> str:
    """Styler de fmt_tabla -> <table class='dat'> con los colores ya aplicados."""
    try:
        return (sty.hide(axis="index")
                   .set_table_attributes("class='dat'")
                   .to_html())
    except Exception:
        return sty.data.to_html(index=False, classes="dat")

def html_resumen_pdf(x, gc_, arbol_html: str, periodo: str) -> str:
    """Documento imprimible de la pestaña Resumen."""
    logo_html = ""
    filtro = _desc_filtros()
    nota = (f"<div class='nofilt'>Vista filtrada — solo incluye: <b>{filtro}</b></div>"
            if filtro else "")

    tab_gc = _tabla_html_simple(fmt_tabla(gc_, ["cat_Grupo_de_Cuentas"],
                                          total=True, inc_fc=True))

    # las 4 tablas de variaciones (mismo origen que en pantalla)
    rdoc = _leer_riesgos()
    bloques_var = ""
    if rdoc.get("riesgos") or rdoc.get("oportunidades"):
        neg = _split_subsecciones(rdoc.get("riesgos", []))
        pos = _split_subsecciones(rdoc.get("oportunidades", []))
        for clave, titulo in (("vs AA", "Variaciones vs Año Anterior (YTD)"),
                              ("vs Plan", "Variaciones vs Plan (YTD)")):
            n_pos = next((v for k, v in pos.items() if clave in k), [])
            n_neg = next((v for k, v in neg.items() if clave in k), [])
            bloques_var += f"<h3>{titulo}</h3>"
            for etiqueta, notas in (("Variaciones positivas", n_pos),
                                    ("Variaciones negativas", n_neg)):
                bloques_var += f"<h3>{etiqueta}</h3>"
                d_ = _parsear_notas(notas)
                bloques_var += (d_.to_html(index=False, classes="dat", escape=True)
                                if not d_.empty else "<p class='nofilt'>Sin comentarios.</p>")

    return f"""<!doctype html>
<html lang="es"><head><meta charset="utf-8">
<title>Resumen · Vista Territorio</title>
<style>{_PDF_CSS}{_KPI_CSS.replace('<style>', '').replace('</style>', '')}
{_ARBOL_CSS.replace('<style>', '').replace('</style>', '')}</style></head>
<body onload="window.print()">
<div class="enc">{logo_html}
  <div><div class="kicker">VISTA TERRITORIO · DASHBOARD FINANCIERO</div>
    <h1>Resumen de gasto territorial</h1>
    <div class="sub">Resumen &nbsp;·&nbsp; {periodo} &nbsp;·&nbsp; cifras en millones de pesos</div>
  </div></div>
{nota}
<h2>Gasto</h2>
{kpis_gasto_html(x).replace('<style>', '<style>')}
<h2>Desglose completo · Grupo de Cuentas → Cuenta</h2>
{arbol_html}
<h2>Detalle por Grupo de Cuentas</h2>
{tab_gc}
{('<h2>Variaciones · Riesgos y Oportunidades</h2>' + bloques_var) if bloques_var else ''}
<div class="pie">Generado el {datetime.now():%d/%m/%Y %H:%M} · Datos del dashboard Vista Territorio.
  Para guardarlo como PDF elige "Destino: Guardar como PDF" en el diálogo de impresión.</div>
</body></html>"""

# =========================================================== acceso público (sin login)
if "user" not in st.session_state or st.session_state.user is None:
    st.session_state.user = {"username": "visitante", "role": "viewer"}

# =========================================================== sidebar
with st.sidebar:
    st.markdown(
        """<div class='vt-side-brand'>
          <div class='t'>◈ Vista Territorio</div>
          <div class='s'>Tableros · alertas · demo pública</div>
        </div>""",
        unsafe_allow_html=True,
    )
    st.caption(f"Real hasta sem **{SMR}** · Plan a sem **{SMP}**")
    modo = st.radio("Filtrar por", ["Rango de semanas", "Mes", "Trimestre", "Semana individual"],
                    key="modo")
    ind = modo == "Semana individual"
    MES = None    # tupla de meses seleccionados (prorrateo diario) o None
    SEMS = None   # tupla de semanas no contiguas (trimestres) o None

    if ind:
        sf = st.slider("Semana", 1, SMP, SMR, key="sem_ind")
        si = sf
    elif modo == "Mes":
        _p26 = pesos_mes(2026)
        msel = st.multiselect(
            "Mes(es)", list(range(1, 13)),
            format_func=lambda m: f"{NOM_MES[m]} (sem {rango_mes(2026, m)[0]}–{rango_mes(2026, m)[1]})",
            default=[max(_p26.get(SMR, {12: 1.0}).items(), key=lambda kv: kv[1])[0]],
            key="mes_multi")
        if not msel:
            st.info("Elige al menos un mes.")
            st.stop()
        MES = tuple(sorted(msel))
        si = min(rango_mes(2026, m)[0] for m in MES)
        sf = max(rango_mes(2026, m)[1] for m in MES)
        parciales = [s for s in range(si, sf + 1)
                     if 0 < sum(_p26.get(s, {}).get(m, 0) for m in MES) < 1]
        if parciales:
            det = " · ".join(
                f"sem {s}: {sum(_p26[s].get(m, 0) for m in MES)*7:.0f}/7 días" for s in parciales)
            st.caption(f"Prorrateo diario — {det}")
    elif modo == "Trimestre":
        qsel = st.multiselect("Trimestre(s)", ["Q1", "Q2", "Q3", "Q4"],
                              default=[f"Q{min(4, (SMR - 1) // 13 + 1)}"], key="tri_multi")
        if not qsel:
            st.info("Elige al menos un trimestre.")
            st.stop()
        QRANGO = {"Q1": (1, 13), "Q2": (14, 26), "Q3": (27, 39), "Q4": (40, 53)}
        semanas = sorted({s for q in qsel for s in range(QRANGO[q][0], QRANGO[q][1] + 1)})
        si, sf = min(semanas), max(semanas)
        if len(semanas) < sf - si + 1:   # trimestres no contiguos (p.ej. Q1+Q3)
            SEMS = tuple(semanas)
    else:
        si, sf = st.select_slider("Rango de semanas", options=list(range(1, SMP + 1)),
                                  value=(1, SMR), key="sem_rango")

    if not ind and si > SMR:
        st.warning(f"El rango empieza después del corte real (sem {SMR}): "
                   "Real 2026 aparecerá en cero.")
    elif not ind and sf > SMR:
        st.caption(f"Real 2026 solo tiene datos hasta la sem {SMR}.")
    st.caption(f"Analizando semanas **{si}–{sf}**"
               + (f" · {', '.join(NOM_MES[m] for m in MES)} prorrateado" if MES else "")
               + (" · trimestres no contiguos" if SEMS else ""))

    # ---------------------------------------- filtro global (todas las pestañas)
    st.divider()
    st.markdown("**Filtro global**")
    fg_resp = st.multiselect(
        "Responsable Interno",
        sorted(RESPONSABLES_GPO),
        key="fg_resp",
        placeholder="Todos los responsables"
    )

    _gpos_resp = None
    if fg_resp:
        _gpos_resp = {
            grupo
            for responsable in fg_resp
            for grupo in RESPONSABLES_GPO.get(responsable, ())
        }

    _gpos = sorted(
        (x for x in AGG["grupo_cuentas"]["cat_Grupo_de_Cuentas"].unique().to_list()
         if x and (_gpos_resp is None or x in _gpos_resp)),
        key=_orden_gpo_key
    )
    fg_gpo = st.multiselect("Grupo de Cuentas", _gpos, key="fg_gpo",
                            placeholder="Todos los grupos")

    _gpos_filtro = tuple(fg_gpo) if fg_gpo else (
        tuple(sorted(_gpos_resp)) if _gpos_resp is not None else ()
    )
    _ctas_src = AGG["cuentas"]
    if _gpos_filtro:
        _ctas_src = _ctas_src.filter(pl.col("cat_Grupo_de_Cuentas").is_in(_gpos_filtro))
    _ctas = sorted(x for x in _ctas_src["cat_Cuentas"].unique().to_list() if x)
    fg_cta = st.multiselect("Cuenta", _ctas, key="fg_cta",
                            placeholder="Todas las cuentas")

    # cat_Subtipo: baja cardinalidad (28 valores), vive en global/division/
    # grupo_cuentas/cuentas/trimestre/conciliacion (FILTROS_LIGEROS en
    # build_data.py) pero NO en pdc.parquet -> no afecta Div/Terr, Detalle
    # PDC ni Detalle Cuenta (esas pestañas simplemente ignoran el filtro,
    # mismo comportamiento que ya tienen otros filtros con columnas ausentes).
    _subtipos = sorted(x for x in AGG["global"]["cat_Subtipo"].unique().to_list() if x)
    fg_subtipo = st.multiselect("Subtipo", _subtipos, key="fg_subtipo",
                                placeholder="Todos los subtipos")

    FILT = []
    if _gpos_filtro:
        FILT.append(("cat_Grupo_de_Cuentas", _gpos_filtro))
    if fg_cta:
        # los aggs sin cat_Cuentas (global, división…) se filtran por el ID niv3
        _ids = tuple(AGG["cuentas"].filter(pl.col("cat_Cuentas").is_in(fg_cta))
                     ["ID_CONCEPTO_CUENTA_NIV3"].unique().to_list())
        FILT.append(("cat_Cuentas", tuple(fg_cta)))
        FILT.append(("ID_CONCEPTO_CUENTA_NIV3", _ids))
    if fg_subtipo:
        FILT.append(("cat_Subtipo", tuple(fg_subtipo)))
    FILT = tuple(FILT)
    # base para nota_filtro(); cada pestaña que agrega filtros locales
    # (División, PosPre, PDC…) la sobrescribe con su propia tupla.
    st.session_state["_filtros_vista"] = FILT
    if FILT:
        # cat_Subtipo NO existe en pdc.parquet ni en pdc_pospre.parquet
        # (FILTROS_LIGEROS en build_data.py: meterlo ahí multiplicaba el tamaño
        # del archivo). Las vistas que salen de esas fuentes lo ignoran, así
        # que se avisa explícitamente en vez de fallar en silencio.
        st.caption("Filtro activo en todas las pestañas."
                   + (" Cierres solo aplica el de Grupo." if fg_cta else "")
                   + (" Subtipo no aplica en Detalle PDC, PDC & Calor ni en el"
                      " árbol de Detalle Cuenta (esas vistas no traen la columna)."
                      if fg_subtipo else ""))

    # ---------------------------------------- actualización de datos
    st.divider()
    est = leer_estado()
    corriendo = bool(est and est.get("corriendo"))

    gen = META.get("generado", "")
    st.caption(f"Datos del **{gen[:10]}** {gen[11:16]}" if gen else "Datos: fecha desconocida")

    if corriendo:
        st.progress(est["pct"] / 100, text=f"{est['fase']} · {est['pct']}%")
        st.caption(est.get("detalle", ""))
        if st.button("Refrescar estado", key="btn_refresh_est"):
            st.rerun()
        st.caption("Puedes seguir usando el dashboard mientras tanto.")
    else:
        if est and est.get("error"):
            st.error(f"Última actualización falló: {est['error'][:120]}")
        elif est and est.get("fase") == "listo":
            st.success(f"Actualizado {est['actualizado'][11:16]}")

        estado_sql = hay_datos_nuevos(VERSION)
        # El aviso de "datos nuevos" solo se muestra los jueves (día en que
        # Planeación actualiza el dashboard) — pedido del usuario, para no
        # generar ruido el resto de la semana aunque SQL Server ya tenga
        # cambios pendientes de un ciclo anterior.
        if estado_sql == "nuevos" and date.today().weekday() == 3:
            st.warning("Hay datos nuevos en SQL Server")
        elif estado_sql == "desconocido":
            st.caption("No se pudo consultar SQL Server (¿VPN?)")

        if st.button("Recargar datos", key="btn_reload", help="Vuelve a leer aggs/ del disco"):
            st.cache_resource.clear()
            st.cache_data.clear()
            st.rerun()
        st.caption("Modo demo/público: sin conexión a SQL Server.")
    st.divider()
    st.caption(f"Datos generados: {META.get('generado', '—')[:16].replace('T', ' ')}")

R = dict(si=si, sf=sf, individual=ind)

_demo = bool(META.get("demo"))
st.markdown(f"""
<div class="vt-hero">
  <div class="vt-hero-kicker">Vista Territorio · Analytics</div>
  <div class="vt-hero-title">Dashboard de gasto territorial</div>
  <p class="vt-hero-sub">
    Drill-down · alertas · forecast &nbsp;·&nbsp; Semana {SMR} / 2026
    &nbsp;·&nbsp; cifras en millones de pesos
  </p>
  {"<span class='vt-badge'>Datos demo sintéticos</span>" if _demo else "<span class='vt-badge'>Datos locales</span>"}
</div>""", unsafe_allow_html=True)

SECCIONES = ["Resumen", "Temporalidad", "Div / Terr", "Detalle Cuenta", "Detalle PDC",
             "PDC & Calor", "Cierres", "Trimestres", "Comparación de Versiones"]
# 2026-07-24: "Plan + Real" oculta del menú por pedido del usuario (temporal)
# — el código y los datos siguen intactos, solo se quitó de la navegación.
# 2026-07-27: "Sem / Mes" movida justo después de Resumen. "Riesgos & Oport."
# ELIMINADA: sus 4 tablas de variaciones (positivas/negativas × vs AA/vs Plan)
# viven ahora al final del Resumen; las alertas automáticas se descartaron.
SEC = st.segmented_control("Sección", SECCIONES, key="nav_sec",
                           default="Resumen", label_visibility="collapsed")
if SEC is None:
    SEC = "Resumen"
st.divider()
st.header(SEC)

# ----------------------------------------------------------- 1. Resumen
if SEC == "Resumen":
    g = compute("global", si, sf, ind, FILT, mes=MES, sems=SEMS)
    if g.empty:
        st.warning("Sin datos en el rango.")
    else:
        x = g.iloc[0]
        # 2026-07-27: antes se llamaba a _agregar_df sin filtros -> las dos
        # gráficas del Resumen ignoraban el filtro global (Subtipo incluido) y
        # no cuadraban con los KPIs. compute() sí aplica FILT.
        gc_ = compute("grupo_cuentas", si, sf, ind, FILT, mes=MES, sems=SEMS)
        dv = compute("division", si, sf, ind, FILT, mes=MES, sems=SEMS)
        # --- KPIs de Gasto con la variación como SEÑAL debajo de cada uno
        # (2026-07-27): monto Y porcentaje juntos, con flecha ▲/▼. Se quitaron
        # el bloque separado de "Variaciones" y el KPI de "Operación".
        st.markdown("##### Gasto Acumulado")
        st.markdown(kpis_gasto_html(x), unsafe_allow_html=True)

        st.divider()
        # 2026-07-28: gráficas movidas de debajo del árbol a justo debajo de
        # los KPIs (pedido del usuario). Nvo Plan fuera de la gráfica.
        # Fix mismo día: compute() trae una fila por Grupo de Cuentas POR
        # CADA cat_Agrupa1 en que aparece (para permitir drill-down) — al
        # graficar el crudo, la línea punteada de Plan conectaba 2 puntos por
        # categoría en zigzag ("línea duplicada" que reportó el usuario con
        # captura). Aquí es un resumen de alto nivel, sin drill-down: se
        # colapsa por cat_Grupo_de_Cuentas/cat_Direccion_Division antes de
        # graficar (mismo patrón para ambas, aunque "division" hoy no repite).
        series_res = ["Real_2025", "Real_2026", "Plan_2026", "Forecast_Cierre"]
        gc_plot = (gc_.groupby("cat_Grupo_de_Cuentas", as_index=False)[MEDIDAS + ["Forecast_Cierre"]]
                      .sum().pipe(lambda d: d.reindex(
                          d["Real_2026"].abs().sort_values(ascending=False).index)))
        dv_plot = (dv.groupby("cat_Direccion_Division", as_index=False)[MEDIDAS + ["Forecast_Cierre"]]
                     .sum().pipe(lambda d: d.reindex(
                         d["Real_2026"].abs().sort_values(ascending=False).index)))
        a, b = st.columns(2)
        with a:
            chart(barras(gc_plot, "cat_Grupo_de_Cuentas", series=series_res,
                         titulo="Grupo de Cuentas acumulado"))
        with b:
            chart(barras(dv_plot, "cat_Direccion_Division", series=series_res,
                         titulo="Agrupador acumulado"))

        st.divider()
        st.subheader("Desglose completo")
        st.caption("Grupo de Cuentas → Cuenta. Respeta el filtro global del menú lateral.")
        # fuente "cuentas" (no "pdc_pospre"): es la única de las dos que trae
        # cat_Subtipo, así que el árbol respeta el MISMO filtro global que los
        # KPIs de arriba (con pdc_pospre el Subtipo se ignoraba y los números
        # del árbol no cuadraban con los KPIs). De paso es ~500x más chica.
        arbol_perezoso("cuentas", ("cat_Grupo_de_Cuentas", "cat_Cuentas"),
                       si, sf, ind, FILT, MES, SEMS, key="resumen_gc",
                       permitir_expandir=True, inc_fc=True)
        descargar(gc_, f"grupo_cuentas_sem{si}-{sf}")

        # --- 4 tablas de variaciones (antes en la pestaña Riesgos & Oport.,
        # eliminada el 2026-07-27). Solo las tablas con formato: las alertas
        # automáticas y las listas de riesgos/oportunidades se descartaron.
        st.divider()
        tablas_variaciones()

        # --- descarga de TODA la pestaña como documento imprimible
        st.divider()
        st.markdown("##### Descargar el Resumen")
        st.caption("Descarga el archivo y ábrelo: se abre solo el diálogo de "
                   "impresión — elige «Guardar como PDF».")
        _periodo = (f"Semana {sf} — 2026" if ind else f"Semanas {si}–{sf} — 2026")
        if MES:
            _periodo = f"{', '.join(NOM_MES[m] for m in MES)} — 2026"
        if st.button("Preparar Resumen para PDF", key="btn_pdf_resumen"):
            # árbol estático (todos los niveles ya desplegados) — el perezoso
            # de pantalla depende de clicks, que no existen en el documento.
            _arbol_pdf = _arbol_html(
                [(gc_, "cat_Grupo_de_Cuentas"),
                 (_agregar_df(AGG["cuentas"],
                              ["cat_Grupo_de_Cuentas", "cat_Cuentas"],
                              si, sf, ind, MES, SEMS), "cat_Cuentas")],
                inc_fc=True)
            _doc = html_resumen_pdf(x, gc_, _arbol_pdf, _periodo)
            st.download_button(
                "Descargar Resumen (HTML → PDF)",
                _doc.encode("utf-8"),
                f"resumen_gasto_red_sem{si}-{sf}.html",
                "text/html", type="primary", key="dl_pdf_resumen")

# ----------------------------------------------------------- 2. Div/Terr drill-down
def _bajar(path_key: str, col: str, widget_key: str):
    """callback: agrega un nivel al path si el usuario eligió algo distinto de '—'."""
    v = st.session_state.get(widget_key)
    if v and v != "—":
        st.session_state[path_key] = st.session_state[path_key] + [(col, v)]

def _subir(path_key: str, n: int):
    st.session_state[path_key] = st.session_state[path_key][:n]

if SEC == "Div / Terr":
    ag1_df = AGG.get("agrupa1")
    ag1_opts = ["(Todas)"] + sorted([str(x) for x in ag1_df["cat_Agrupa1"].drop_nulls().unique() if str(x) != "nan" and str(x) != "None"]) if ag1_df is not None else ["(Todas)"]

    div_df = AGG.get("division")
    pfx_opts = ["(Todos)"] + sorted([str(x) for x in div_df["cat_Prefijo"].drop_nulls().unique() if str(x) != "nan" and str(x) != "None"], key=lambda x: (len(x), x)) if div_df is not None and "cat_Prefijo" in div_df.columns else ["(Todos)"]

    col_a, col_b, col_c = st.columns([1, 1, 1])
    # 2026-07-27: "Territorial" -> "General" y "Contable" -> "Por Cuenta".
    # "Por Cuenta" arranca en Grupo de Cuentas: se quitó la rama Agrupa 1
    # porque ya existe el selector "Filtrar por Agrupa 1" de aquí al lado
    # (queda Grupo de Cuentas -> Cuentas -> PosPre). "General" no se tocó.
    jer = col_a.radio("Jerarquía", ["General", "Por Cuenta"], horizontal=True, key="jer")
    f_ag1_dt = col_b.selectbox("Filtrar por Agrupa 1", ag1_opts, key="dt_ag1_sel")
    f_pfx_dt = col_c.selectbox("Filtrar por Prefijo", pfx_opts, key="dt_pfx_sel")

    HIER_POR_CUENTA = [n for n in HIER_CTA if n != "agrupa1"]
    hier = {"General": HIER_TERR, "Por Cuenta": HIER_POR_CUENTA}[jer]
    # Agrupa 1 ya no es nivel visible en "Por Cuenta": si siguiera agrupando,
    # un Grupo/Cuenta que exista bajo dos Agrupa 1 saldría en dos filas.
    sin_padres = ("cat_Agrupa1",) if jer == "Por Cuenta" else ()

    filt_dt = FILT
    if f_ag1_dt != "(Todas)":
        filt_dt = filt_dt + (("cat_Agrupa1", f_ag1_dt),)
    if f_pfx_dt != "(Todos)":
        filt_dt = filt_dt + (("cat_Prefijo", f_pfx_dt),)
    st.session_state["_filtros_vista"] = filt_dt

    if jer == "Por Cuenta":
        st.caption("Nota: 12 cuentas tienen 2 posiciones presupuestales (PosPre) en el "
                   "catálogo de origen; el movimiento no distingue a cuál pertenece cada "
                   "transacción, así que aparece con el mismo monto completo bajo AMBAS "
                   "posiciones (no se divide).")

    d0 = compute(hier[0], si, sf, ind, filt_dt, mes=MES, sems=SEMS,
                 excluir_padres=sin_padres)
    if d0.empty:
        st.warning("Sin datos en el rango.")
    else:
        st.caption("Usa el botón **+** de cada fila para desglosarla, "
                   "y **−** para cerrarla.")
        st.markdown(arbol_jerarquia(tuple(hier), si, sf, ind, filt_dt, MES, SEMS,
                                    sin_padres=sin_padres),
                    unsafe_allow_html=True)
        descargar(compute(hier[-1], si, sf, ind, filt_dt, mes=MES, sems=SEMS,
                          excluir_padres=sin_padres),
                  f"{hier[-1]}_sem{si}-{sf}")
        st.divider()
        chart(barras(d0, DIMS[hier[0]][1],
                     titulo=f"{DIMS[hier[0]][0]} acumulado"))

# ----------------------------------------------------------- 3. Detalle Cuenta
if SEC == "Detalle Cuenta":
    ag1_df = AGG.get("agrupa1")
    ag1_opts = ["(Todas)"] + sorted([str(x) for x in ag1_df["cat_Agrupa1"].drop_nulls().unique() if str(x) != "nan" and str(x) != "None"]) if ag1_df is not None else ["(Todas)"]

    c1, c2 = st.columns(2)

    q = c1.text_input(
        "Buscar cuenta",
        placeholder="nombre de la cuenta…",
        key="cta_q"
    )

    f_ag1_dc = c2.selectbox(
        "Filtrar por Agrupa 1",
        ag1_opts,
        key="dc_ag1_sel"
    )

    cta_base = FILT
    if f_ag1_dc != "(Todas)":
        cta_base = cta_base + (("cat_Agrupa1", f_ag1_dc),)

    d = compute(
        "cuentas",
        si,
        sf,
        ind,
        cta_base,
        mes=MES,
        sems=SEMS
    )

    if d.empty:
        st.warning("Sin datos para esa búsqueda.")

    else:
        v = d

        if q:
            v = v[
                v["cat_Cuentas"].str.contains(
                    q,
                    case=False,
                    na=False,
                    regex=False
                )
            ]

        st.subheader(
            f"{v['cat_Cuentas'].nunique():,} cuentas"
        )

        cta_filt = cta_base

        if q and not v.empty:
            _ctas_q = tuple(
                v["cat_Cuentas"]
                .dropna()
                .unique()
                .tolist()
            )

            cta_filt = cta_filt + (
                ("cat_Cuentas", _ctas_q),
            )

        # 2026-07-27: se eliminó el radio de vistas (División/Territorio/Zona/
        # Región) — queda solo el desglose completo hasta Región, sin notas.
        vista_filt = cta_filt
        st.session_state["_filtros_vista"] = vista_filt

        # 2026-07-27: la última rama pasó de PosPre a Agrupador de Reales
        # (pedido del usuario) — requiere que pdc_pospre.parquet incluya
        # cat_Agrupador_Reales (build_data.py, build_pdc_pospre) y que los
        # aggs se hayan regenerado con ese cambio.
        cols_arbol = (
            "cat_Agrupa1",
            "cat_Grupo_de_Cuentas",
            "cat_Cuentas",
            "cat_Direccion_Division",
            "cat_Subdireccion_Territorio",
            "cat_Subdireccion_Zona",
            "cat_Subdireccion_Region",
            "cat_Agrupador_Reales",
        )

        st.subheader("Desglose completo")
        # el árbol sale de pdc_pospre (única fuente con PosPre + territorio),
        # que no trae cat_Subtipo -> se avisa para no dar cifras engañosas.
        aviso_sin_subtipo("el árbol de Detalle Cuenta")

        arbol_perezoso(
            "pdc_pospre",
            cols_arbol,
            si,
            sf,
            ind,
            vista_filt,
            MES,
            SEMS,
            key="dc_region",
            permitir_expandir=len(fg_gpo) == 1
        )

        descargar(
            v,
            f"cuentas_sem{si}-{sf}"
        )

        if not v.empty:
            st.divider()
            chart(
                barras(
                    v,
                    "cat_Cuentas",
                    titulo="Cuentas acumulado"
                )
            )

# ----------------------------------------------------------- 4. Detalle PDC
if SEC == "Detalle PDC":
    # 2026-07-23: 2 botones — "General" (vista de siempre, tabla plana) y
    # "Detalle" (árbol Grupo de Cuentas -> Cuenta -> PDC/CECO). pdc.parquet YA
    # trae cat_Grupo_de_Cuentas + cat_Cuentas junto con la jerarquía
    # territorial (confirmado con el schema real del parquet) -> el árbol de
    # Detalle se arma sin tocar build_data.py ni regenerar nada.
    aviso_sin_subtipo("Detalle PDC")
    vista_pdc = st.radio("Vista", ["General", "Detalle"], horizontal=True, key="pdc_vista")

    if vista_pdc == "General":
        # TODOS los registros de PDC/CECO (el conteo oficial vive en PDC & Calor).
        # Los oficiales llevan su ID (Eco PDV); los CECOs y demás van con '—'.
        d = compute("pdc", si, sf, ind, FILT, mes=MES, sems=SEMS)
        d = d.copy()
        # 2026-07-27: el ID ya no va en columna aparte — se antepone al nombre
        # ("4821 · Mega Cd Juarez Azt") con con_id(), igual que en los árboles,
        # y las filas se ordenan por ese ID para que salgan juntas y en orden.
        # el nombre crudo se conserva aparte: los filtros (oficiales, buscador)
        # comparan contra el catálogo, que va sin el prefijo del ID.
        d["_pdc_raw"] = d["cat_PDC"]
        d["cat_PDC"] = d["cat_PDC"].map(con_id)
        _ord_id = d["cat_PDC"].str.extract(r"^(\d+) · ", expand=False)
        d = (d.assign(_id_num=pd.to_numeric(_ord_id, errors="coerce"))
               .sort_values(["_id_num", "cat_PDC"], na_position="last")
               .drop(columns="_id_num")
               .reset_index(drop=True))
        if not PDC_IDS:
            st.warning("Falta `aggs/pdc_ids.json` (catálogo oficial de PDCs). "
                       "Corre `python build_data.py`.")
        st.subheader("Registro de Gastos por PDC o CECO")

        f1, f2 = st.columns([2, 2])
        divs = ["(Todas)"] + sorted(d["cat_Direccion_Division"].dropna().unique(), key=_orden_div_key)
        fdiv = f1.selectbox("División", divs, key="pdc_div")
        solo_of = f2.checkbox("Solo PDCs oficiales", value=False, key="pdc_of")
        filtro_pdc_gen = buscador_pdc_ceco("pdc_gen")

        _filt_vista_pdc = FILT
        v = d
        if fdiv != "(Todas)":
            v = v[v["cat_Direccion_Division"] == fdiv]
            _filt_vista_pdc = _filt_vista_pdc + (("cat_Direccion_Division", fdiv),)
        if solo_of:
            v = v[v["_pdc_raw"].isin(PDC_IDS)]
        if filtro_pdc_gen:
            _, _pdcs_sel_gen = filtro_pdc_gen[0]
            v = v[v["_pdc_raw"].isin(_pdcs_sel_gen)]
            _filt_vista_pdc = _filt_vista_pdc + filtro_pdc_gen
        st.session_state["_filtros_vista"] = _filt_vista_pdc

        st.caption(f"Mostrando **{len(v):,}** de {len(d):,} registros.")
        tabla(v.drop(columns="_pdc_raw"),
              ["cat_PDC", "cat_Direccion_Division", "cat_Subdireccion_Territorio",
               "cat_Subdireccion_Region", "cat_Subdireccion_Zona"],
              f"pdc_detalle_sem{si}-{sf}", height=600)

    else:  # "Detalle" — árbol perezoso Gpo de Cuentas -> Cuenta -> PDC/CECO
        filtro_pdc_det = buscador_pdc_ceco(
            "pdc_detalle", "Buscar punto de contacto o ID Centro de Costos (opcional)")
        det_filt = FILT + filtro_pdc_det if filtro_pdc_det else FILT
        st.session_state["_filtros_vista"] = det_filt

        # con un filtro de PDC ya acotado, expandir de golpe es seguro aunque
        # el filtro global de Grupo de Cuentas no esté acotado a 1 solo valor.
        _puede_expandir_pdc = len(fg_gpo) == 1 or bool(filtro_pdc_det)
        arbol_perezoso(
            "pdc",
            ("cat_Grupo_de_Cuentas", "cat_Cuentas", "cat_PDC"),
            si, sf, ind, det_filt, MES, SEMS,
            key="pdc_detalle",
            permitir_expandir=_puede_expandir_pdc
        )
        if not _puede_expandir_pdc:
            st.caption("Para expandir el árbol de un golpe (1er nivel), elige un solo "
                      "Grupo de Cuentas en el filtro global del menú lateral, o busca "
                      "un PDC/CECO específico arriba.")

        # Descarga del árbol completo (Gpo. Cuentas x Cuenta x PDC/CECO) en
        # CSV. Mismo criterio de seguridad que "Expandir todo": solo se ofrece
        # sin más cuando el filtro ya acotó los datos (1 Grupo de Cuentas o
        # una búsqueda de PDC) — calcular TODA la combinación de golpe sobre
        # pdc.parquet (7.9M filas) sin acotar es el mismo riesgo de colgar el
        # navegador que el árbol perezoso ya evita al no pre-renderizarse.
        if _puede_expandir_pdc:
            if st.button("Preparar descarga del árbol completo", key="pdc_detalle_prep_dl"):
                _df_pdc_dl = AGG["pdc"]
                for k, v in det_filt:
                    if k in _df_pdc_dl.columns:
                        _df_pdc_dl = _df_pdc_dl.filter(
                            pl.col(k).is_in(list(v)) if isinstance(v, tuple) else pl.col(k) == v)
                _df_dl = _agregar_df(_df_pdc_dl, ["cat_Grupo_de_Cuentas", "cat_Cuentas", "cat_PDC"],
                                     si, sf, ind, MES, SEMS)
                descargar(_df_dl, f"pdc_detalle_arbol_sem{si}-{sf}")
        else:
            st.caption("Para descargar el árbol completo, elige un solo Grupo de Cuentas "
                      "en el filtro global o busca un PDC/CECO específico arriba.")

# ----------------------------------------------------------- 5. PDC & Calor
if SEC == "PDC & Calor":
    # solo los PDCs OFICIALES del catálogo (2,543): el detalle granular de
    # todos los registros (~4.6K, incl. CECOs) vive en la pestaña Detalle PDC.
    aviso_sin_subtipo("PDC & Calor")
    filtro_pdc_calor = buscador_pdc_ceco("pdc_calor", "Buscar PDC o ID Centro de Costos")
    filt_calor = FILT + filtro_pdc_calor if filtro_pdc_calor else FILT
    st.session_state["_filtros_vista"] = filt_calor
    d = compute("pdc", si, sf, ind, filt_calor, mes=MES, sems=SEMS)
    d = d[d["cat_PDC"].isin(PDC_IDS)]
    # Un PDC oficial = un nombre del catálogo (2,543). Los homónimos bajo
    # distinta región generan filas extra (~2,663) que aquí se CONSOLIDAN
    # por nombre; el desglose granular vive en las tablas y en Detalle PDC.
    _mon = ["Real_2025", "Plan_2026", "Nvo_Plan_2026", "Real_2026", "Vs_Plan_Abs"]
    u = d.groupby("cat_PDC", as_index=False)[_mon].sum()
    vivos = u[(u["Plan_2026"] != 0) | (u["Real_2026"] != 0)]
    con_plan = vivos[vivos["Plan_2026"] != 0].copy()
    if con_plan.empty:
        st.warning("Sin PDCs con plan en el rango.")
    else:
        con_plan["Vs_Plan_Pct"] = con_plan["Vs_Plan_Abs"] / con_plan["Plan_2026"].abs() * 100
        n_sobre = int((con_plan["Vs_Plan_Abs"] > 0).sum())
        n_bajo = int((con_plan["Vs_Plan_Abs"] <= 0).sum())
        vs_tot = con_plan["Vs_Plan_Abs"].sum()
        estado = "Sobre plan" if vs_tot > 0 else "Bajo plan"

        # 2026-07-27: KPIs juntos en una tabla estilizada (no cuadros sueltos)
        _mediana = con_plan["Vs_Plan_Pct"].median()
        st.markdown(kpis_tabla([
            ("Estado general", estado, f"{MC(vs_tot)} vs plan",
             "alza" if vs_tot > 0 else "baja"),
            ("Sobre plan", f"{n_sobre:,}", f"{n_sobre / len(con_plan) * 100:.0f}%", "alza"),
            ("Bajo plan", f"{n_bajo:,}", f"{n_bajo / len(con_plan) * 100:.0f}%", "baja"),
            ("Desviación mediana", f"{_mediana:+.1f}%", "vs plan",
             "alza" if _mediana > 0 else "baja"),
        ]), unsafe_allow_html=True)

        st.divider()
        # histograma en DINERO — 2026-07-27: movido al PRINCIPIO, antes de las
        # dos tablas de Top 25.
        mdp = con_plan["Vs_Plan_Abs"] / 1e6
        import math as _math
        import numpy as _np
        tope_max = max(2, int(_math.ceil(mdp.abs().max())))
        defecto = min(max(1, int(_math.ceil(mdp.abs().quantile(0.99)))), tope_max)
        LIM = st.slider("Recorte del eje (± MDP)", 1, tope_max, defecto, key="hm_lim",
                        help="Los PDCs fuera de este rango se cuentan aparte; "
                             "sin recorte los casos extremos aplastan el histograma.")
        dentro = mdp[mdp.abs() <= LIM]

        sigma = float(dentro.std()) or 1.0
        cuentas_, bordes = _np.histogram(dentro, bins=40, range=(-LIM, LIM))
        centros = (bordes[:-1] + bordes[1:]) / 2
        ancho = bordes[1] - bordes[0]
        C_PLAN, C_1S, C_2S, C_3S = "#ccd1d9", "#87fc72", "#f5ff38", "#ff6456"

        def _color_banda(c0):
            if abs(c0) < ancho:
                return C_PLAN
            a = abs(c0) / sigma
            return C_1S if a <= 1 else (C_2S if a <= 2 else C_3S)

        fig = go.Figure(go.Bar(
            x=centros, y=cuentas_, width=ancho * .92,
            marker=dict(color=[_color_banda(c0) for c0 in centros], line=dict(width=0)),
            hovertemplate="%{x:.1f} MDP vs plan<br>%{y} PDCs<extra></extra>",
            showlegend=False))
        fig.add_vline(x=0, line_dash="dash", line_color="#000000",
                      annotation_text="Plan", annotation_position="top")
        for s_, txt in ((1, "±1σ"), (2, "±2σ")):
            if s_ * sigma < LIM:
                for lado in (-1, 1):
                    fig.add_vline(x=lado * s_ * sigma, line_dash="dot",
                                  line_color="#8a93a0", line_width=1)
        fig.update_xaxes(title="MDP vs Plan · ← bajo plan | sobre plan →",
                         range=[-LIM, LIM])
        fig.update_yaxes(title="Número de PDCs")
        fig.update_layout(bargap=0)
        chart(_layout(fig, f"Alineación de puntos de contacto · {len(dentro):,} PDCs (±{LIM} MDP)", 430))

        # Top 25: detalle granular (los homónimos se distinguen por su División)
        st.divider()
        s = (d[(d["Plan_2026"] != 0) | (d["Real_2026"] != 0)]
             .sort_values("Vs_Plan_Abs", ascending=False))
        cols_ = ["cat_PDC", "cat_Direccion_Division"]
        st.subheader("Top 25 Sobre Plan")
        tabla(s.head(25), cols_)
        st.subheader("Bottom 25 Sobre Plan")
        tabla(s.tail(25).iloc[::-1], cols_)

# ------------------------------------------------------ 5b. Forecast
# aggs/forecast_ceco.parquet ya viene agregado por Grupo de Cuentas x Cuenta
# x semana (no por CECO — decisión del usuario: la pestaña es por cuenta).
@st.cache_data(show_spinner="Calculando forecast…")
def _forecast_rango(si: int, sf: int, individual: bool) -> pd.DataFrame:
    """Real_2026 y Forecast acumulados por Grupo de Cuentas/Cuenta en el rango
    de semanas (sin prorrateo de mes — el CSV solo trae semana entera). Real
    2026 se topa en SMR igual que en compute()."""
    df = EXTRA.get("forecast_ceco")
    if df is None:
        return pd.DataFrame()
    lo = sf if individual else si
    real_hi = min(sf, SMR)
    g = (df.filter((pl.col("sem") >= lo) & (pl.col("sem") <= sf))
           .with_columns([
               pl.when((pl.col("sem") >= lo) & (pl.col("sem") <= real_hi))
                 .then(pl.col("Real_2026")).otherwise(0.0).alias("_r26"),
               pl.col("Forecast").alias("_fc"),
           ])
           .group_by(["cat_Grupo_de_Cuentas", "cat_Cuentas"])
           .agg([pl.col("_r26").sum().alias("Real_2026"), pl.col("_fc").sum().alias("Forecast")])
           .to_pandas())
    g["Vs_Forecast_Abs"] = g["Real_2026"] - g["Forecast"]
    g["Vs_Forecast_Pct"] = g["Vs_Forecast_Abs"] / g["Forecast"].abs().replace(0, pd.NA) * 100
    return g

@st.cache_data(show_spinner=False)
def _forecast_semanal() -> pd.DataFrame:
    """Serie semanal 1–53 (todas las cuentas): Real 2026 topado en SMR."""
    df = EXTRA.get("forecast_ceco")
    if df is None:
        return pd.DataFrame()
    w = (df.group_by("sem")
           .agg([pl.col("Real_2026").sum(), pl.col("Forecast").sum()])
           .sort("sem").to_pandas())
    w.loc[w["sem"] > SMR, "Real_2026"] = None
    return w

if SEC == "Plan + Real":
    fce = EXTRA.get("forecast_ceco")
    if fce is None:
        st.error("Falta `aggs/forecast_ceco.parquet`. Corre `python build_data.py` "
                 "con el CSV de forecast en Downloads.")
    else:
        d = _forecast_rango(si, sf, ind)
        vivos = d[(d["Forecast"] != 0) | (d["Real_2026"] != 0)]
        con_fc = vivos[vivos["Forecast"] != 0].copy()
        if con_fc.empty:
            st.warning("Sin cuentas con forecast en el rango.")
        else:
            tot_r26, tot_fc = con_fc["Real_2026"].sum(), con_fc["Forecast"].sum()
            vs_tot = tot_r26 - tot_fc
            estado = "Sobre forecast" if vs_tot > 0 else "Bajo forecast"
            n_sobre = int((con_fc["Vs_Forecast_Abs"] > 0).sum())
            n_bajo = int((con_fc["Vs_Forecast_Abs"] <= 0).sum())
            cumpl = (tot_r26 / tot_fc * 100) if tot_fc else 0.0

            c = st.columns(5)
            c[0].metric("Real 2026", M(tot_r26), f"{cumpl:.1f}% del forecast", border=True)
            c[1].metric("Forecast", M(tot_fc), "acumulado en el rango", delta_color="off", border=True)
            c[2].metric("Estado general", estado, MC(vs_tot) + " vs forecast",
                        delta_color="inverse" if vs_tot > 0 else "normal", border=True)
            c[3].metric("Cuentas sobre forecast", f"{n_sobre} de {len(con_fc)}",
                        f"{n_sobre / len(con_fc) * 100:.0f}%", delta_color="inverse", border=True)
            c[4].metric("Cuentas bajo forecast", f"{n_bajo} de {len(con_fc)}",
                        f"{n_bajo / len(con_fc) * 100:.0f}%", delta_color="off", border=True)

            gc_fc = (con_fc.groupby("cat_Grupo_de_Cuentas", as_index=False)
                     [["Real_2026", "Forecast"]].sum())
            gc_fc = gc_fc.reindex(gc_fc["Real_2026"].abs().sort_values(ascending=False).index)

            # Tree and download FIRST
            st.subheader("Detalle por Grupo de Cuentas / Cuenta (Plan + Real)")
            def _fila_fc(r, col, nivel, hoja):
                import html as _h
                v = r["Real_2026"] - r["Forecast"]
                caret = "vt-hoja" if hoja else "vt-caret"
                peso = "font-weight:600;" if nivel == 0 else ""
                pct = (v / abs(r["Forecast"]) * 100) if r["Forecast"] else None
                return (f"<div class='vt-row-fc' style='{peso}'>"
                        f"<span style='padding-left:{10 + nivel * 20}px'>"
                        f"<span class='{caret}'></span>{_h.escape(str(r[col]))}</span>"
                        f"{_sp(r['Real_2026'])}{_sp(r['Forecast'])}"
                        f"{_spv(v, pct)}</div>")

            _FC_HEAD = ("<div class='vt-row-fc vt-head'><span></span>"
                       + "".join(f"<span>{h}</span>" for h in
                                 ["Real 2026", "Forecast", "vs Forecast", "vs Fcst %"]) + "</div>")
            nodos_fc = {}
            for _, r in con_fc.iterrows():
                nodos_fc.setdefault(r["cat_Grupo_de_Cuentas"], []).append(r)
            html_fc = []
            for _, gr in gc_fc.iterrows():
                g_ = gr["cat_Grupo_de_Cuentas"]
                hijos = sorted(nodos_fc.get(g_, []), key=lambda r: -abs(r["Real_2026"]))
                html_fc.append(f"<details><summary>{_fila_fc(gr, 'cat_Grupo_de_Cuentas', 0, False)}</summary>"
                               + "".join(_fila_fc(r, "cat_Cuentas", 1, True) for r in hijos)
                               + "</details>")
            tot_fila = (f"<div class='vt-row-fc vt-tot'><span>TOTAL</span>"
                       f"{_sp(tot_r26)}{_sp(tot_fc)}"
                       f"{_spv(vs_tot, (vs_tot/abs(tot_fc)*100) if tot_fc else 0)}</div>")
            st.markdown(_ARBOL_CSS + "<div class='vt-tree'>" + _FC_HEAD
                       + "".join(html_fc) + tot_fila + "</div>", unsafe_allow_html=True)
            descargar(con_fc, f"plan_mas_real_sem{si}-{sf}")

            # Charts BELOW tree
            st.divider()
            a, b = st.columns(2)
            with a:
                fig1 = go.Figure()
                fig1.add_bar(name="Real 2026", x=gc_fc["cat_Grupo_de_Cuentas"], y=gc_fc["Real_2026"],
                            marker_color=COLOR["Real_2026"],
                            hovertemplate="%{x}<br>%{y:,.0f}<extra></extra>")
                fig1.add_scatter(name="Forecast", x=gc_fc["cat_Grupo_de_Cuentas"], y=gc_fc["Forecast"],
                                mode="lines+markers",
                                line=dict(color=COLOR["Plan_2026"], width=2.5, dash="dash"),
                                marker=dict(size=7), hovertemplate="%{x}<br>%{y:,.0f}<extra></extra>")
                fig1.update_layout(hovermode="x unified")
                fig1.update_xaxes(tickangle=-35)
                chart(_layout(fig1, "Real 2026 vs Forecast · por Grupo de Cuentas", 400))
            with b:
                wk = _forecast_semanal()
                fig2 = go.Figure()
                fig2.add_bar(name="Real 2026", x=wk["sem"], y=wk["Real_2026"],
                            marker_color=COLOR["Real_2026"], opacity=.9,
                            hovertemplate="sem %{x}<br>%{y:,.0f}<extra></extra>")
                fig2.add_scatter(name="Forecast", x=wk["sem"], y=wk["Forecast"], mode="lines+markers",
                                line=dict(color=COLOR["Plan_2026"], width=2.5, dash="dash"),
                                marker=dict(size=5), hovertemplate="sem %{x}<br>%{y:,.0f}<extra></extra>")
                fig2.update_layout(hovermode="x unified", barmode="group")
                fig2.update_xaxes(title="Semana")
                chart(_layout(fig2, f"Serie semanal 1–53 · Real 2026 vs Forecast (corte sem {SMR})", 400))

            top_dev = con_fc.reindex(con_fc["Vs_Forecast_Abs"].abs()
                                     .sort_values(ascending=False).index).head(10)
            fig3 = go.Figure(go.Bar(
                y=top_dev["cat_Cuentas"].iloc[::-1], x=top_dev["Vs_Forecast_Abs"].iloc[::-1],
                orientation="h",
                marker=dict(color=[ROJO if v > 0 else CARBON for v in top_dev["Vs_Forecast_Abs"].iloc[::-1]]),
                hovertemplate="%{y}<br>%{x:,.0f}<extra></extra>", showlegend=False))
            fig3.add_vline(x=0, line_color="#000000", line_width=1)
            fig3.update_xaxes(title="Δ vs Forecast (Real − Forecast)")
            chart(_layout(fig3, "Top 10 cuentas · mayor desviación absoluta vs forecast",
                         max(360, 60 + 32 * len(top_dev))))

# ----------------------------------------------------------- 5. Temporalidad
if SEC == "Temporalidad":
    # Filtro por PDC / Centro de Costos (2026-07-22, pedido del usuario): global.parquet
    # no tiene ID_CENTRO_COSTOS ni cat_PDC, así que se resuelve el CECO a su(s)
    # PDC(s) (vía _CECO_A_PDC, ya usado en Detalle PDC) y se agrega ese PDC como
    # filtro sobre pdc.parquet en vez de global.parquet, sin tocar ningún agg.
    filtro_pdc_sm = buscador_pdc_ceco("sm_pdc", "Buscar PDC o ID Centro de Costos")
    filt_sm = FILT
    alias_sm = "global"
    if filtro_pdc_sm:
        filt_sm = FILT + filtro_pdc_sm
        alias_sm = "pdc"
    st.session_state["_filtros_vista"] = filt_sm

    w = weekly(alias=alias_sm, filtros=filt_sm)
    gran = st.radio("Granularidad", ["Semanal", "Mensual", "Trimestral"], horizontal=True, key="sem_gran")

    # Cálculos base para todas las granularidades
    t = w.copy()
    t["Vs_AA_Abs"] = t["Real_2026"] - t["Real_2025"]
    t["Vs_AA_Pct"] = t["Vs_AA_Abs"] / t["Real_2025"].abs().replace(0, pd.NA) * 100
    t["Vs_Plan_Abs"] = t["Real_2026"] - t["Plan_2026"]
    t["Vs_Plan_Pct"] = t["Vs_Plan_Abs"] / t["Plan_2026"].abs().replace(0, pd.NA) * 100

    if gran == "Mensual":
        base = semanal_a_mensual(t)
        keycol, keylbl = "Mes", "Mes"
    elif gran == "Trimestral":
        # Mapeo semana -> Q1..Q4 con los MISMOS rangos fijos que la pestaña
        # Trimestres (QRANGO_TRI, año EKT de 53 semanas: Q4 = semanas 40-53).
        # Antes se calculaba con (sem-1)//13+1, que genera un Q5 espurio de
        # una sola semana (la 53) porque 53 no es múltiplo de 13 — reportado
        # por el usuario 2026-07-29 ("el quinto trimestre está raro"), fix
        # aplicado usando el mismo mapeo que ya existía en Trimestres.
        _sem_a_q = {s: q for q, (lo, hi) in QRANGO_TRI.items() if q != "FY"
                   for s in range(lo, hi + 1)}
        t["Trimestre"] = t["sem"].map(_sem_a_q)
        # Forecast de cierre de año (mismo criterio que Forecast_Cierre de
        # compute()/_medidas(): Real 2026 acumulado de TODO el año hasta SMR +
        # Plan 2026 de las semanas aún no ejecutadas). Es un valor único de
        # "cierre proyectado", igual en cada fila — no es un forecast por
        # trimestre ni se debe sumar entre trimestres.
        forecast_cierre = (t.loc[t["sem"] <= SMR, "Real_2026"].sum()
                          + t.loc[t["sem"] > SMR, "Plan_2026"].sum())
        base = (t.groupby("Trimestre")[MEDIDAS + ["Vs_AA_Abs", "Vs_AA_Pct", "Vs_Plan_Abs", "Vs_Plan_Pct"]]
                 .sum().reindex(["Q1", "Q2", "Q3", "Q4"]).reset_index())
        base["Forecast_Cierre"] = forecast_cierre
        keycol, keylbl = "Trimestre", "Trimestre"
    else:
        base = t
        keycol, keylbl = "sem", "Semana"

    st.subheader(f"Tabla de gasto {'trimestral' if gran == 'Trimestral' else 'mensual' if gran == 'Mensual' else 'semanal'}")
    t = base.copy()
    if gran != "Semanal" and gran != "Trimestral":
        t["Vs_AA_Abs"] = t["Real_2026"] - t["Real_2025"]
        t["Vs_AA_Pct"] = t["Vs_AA_Abs"] / t["Real_2025"].abs().replace(0, pd.NA) * 100
        t["Vs_Plan_Abs"] = t["Real_2026"] - t["Plan_2026"]
        t["Vs_Plan_Pct"] = t["Vs_Plan_Abs"] / t["Plan_2026"].abs().replace(0, pd.NA) * 100
    t["Acum. Real 2026"] = t["Real_2026"].cumsum()
    t["Acum. Plan 2026"] = t["Plan_2026"].cumsum()
    t["Vs AA"] = t["Vs_AA_Abs"]
    t["Vs AA %"] = t["Vs_AA_Pct"]
    t["Vs Plan"] = t["Vs_Plan_Abs"]
    t["Vs Plan %"] = t["Vs_Plan_Pct"]

    money = [m for m in MEDIDAS if m != "Nvo_Plan_2026"]
    vs_money = ["Vs Plan", "Vs AA"]
    vs_pct = ["Vs Plan %", "Vs AA %"]
    acum = ["Acum. Real 2026", "Acum. Plan 2026"]
    fc_cols = []
    if gran == "Trimestral":
        # Forecast_Cierre es el mismo valor en todas las filas (proyección de
        # cierre de año, no un dato por trimestre) — NO se suma en el TOTAL,
        # se toma tal cual (igual que cualquier fila).
        t["Vs Forecast"] = t["Forecast_Cierre"] - t["Plan_2026"]
        t["Vs Forecast %"] = t["Vs Forecast"] / t["Plan_2026"].abs().replace(0, pd.NA) * 100
        fc_cols = ["Forecast_Cierre", "Vs Forecast", "Vs Forecast %"]
    v = t[[keycol] + money + vs_money + vs_pct + acum + fc_cols].copy()
    v = v.rename(columns={**NOMBRE, keycol: keylbl})
    v[keylbl] = v[keylbl].astype(str)
    tot = {keylbl: "TOTAL"}
    for c_ in money:
        tot[NOMBRE.get(c_, c_)] = v[NOMBRE.get(c_, c_)].sum()
    for c_ in ("Acum. Real 2026", "Acum. Plan 2026"):
        tot[c_] = v[c_].iloc[-1] if not v.empty else 0.0
    _aa_t, _p_t = t["Vs_AA_Abs"].sum(), t["Vs_Plan_Abs"].sum()
    _r25_t, _p26_t = t["Real_2025"].sum(), t["Plan_2026"].sum()
    tot["Vs AA"] = _aa_t
    tot["Vs AA %"] = (_aa_t / abs(_r25_t) * 100) if _r25_t else 0.0
    tot["Vs Plan"] = _p_t
    tot["Vs Plan %"] = (_p_t / abs(_p26_t) * 100) if _p26_t else 0.0
    if gran == "Trimestral" and not v.empty:
        _fc_t = t["Forecast_Cierre"].iloc[-1]
        tot["Forecast"] = _fc_t
        tot["Vs Forecast"] = _fc_t - _p26_t
        tot["Vs Forecast %"] = (tot["Vs Forecast"] / abs(_p26_t) * 100) if _p26_t else 0.0
    _cfg_semmes = {c: st.column_config.TextColumn(width=125) for c in v.columns[1:]}
    if v.columns.size:
        _cfg_semmes[v.columns[0]] = st.column_config.TextColumn(pinned=True)
    _money_cols = [NOMBRE.get(c_, c_) for c_ in money] + vs_money + acum + (["Forecast", "Vs Forecast"] if gran == "Trimestral" else [])
    _pct_cols = vs_pct + (["Vs Forecast %"] if gran == "Trimestral" else [])
    st.dataframe(estilizar(v, _money_cols, pct=_pct_cols, fila_total=tot),
                 width='stretch', hide_index=True, height=460, column_config=_cfg_semmes)
    descargar(t, f"gasto_{'mensual' if gran == 'Mensual' else 'semanal'}_sem{si}-{sf}")

    st.divider()
    if gran == "Mensual":
        wm = semanal_a_mensual(w)
        chart(lineas(wm, titulo="Mensual acumulado", xcol="Mes", xtitle="Mes"))
        base = wm
        keycol, keylbl = "Mes", "Mes"
        p26 = pesos_mes(2026)
        partidas = [(s, d) for s, d in sorted(p26.items()) if len(d) > 1]
        with st.expander(f"Cómo se reparten las {len(partidas)} semanas que cruzan dos meses"):
            st.caption("Cada semana aporta a cada mes en proporción a sus días. "
                       "Real 2025 usa el calendario de 2025; las series de 2026, el de 2026.")
            st.dataframe(pd.DataFrame([
                {"Semana": s,
                 "Reparto": " · ".join(f"{NOM_MES[m]}: {f*7:.0f}/7 días ({f*100:.0f}%)"
                                       for m, f in sorted(d.items()))}
                for s, d in partidas]), width='stretch', hide_index=True)
    elif gran == "Trimestral":
        chart(lineas(t, titulo="Trimestral acumulado", xcol="Trimestre", xtitle="Trimestre"))
        a, b = st.columns(2)
        with a:
            chart(lineas(t, ["Real_2025", "Real_2026"], "Real 2026 vs Real 2025 acumulado",
                         370, xcol="Trimestre", xtitle="Trimestre"))
        with b:
            chart(lineas(t, ["Real_2026", "Plan_2026"], "Real 2026 vs Plan 2026 acumulado",
                         370, xcol="Trimestre", xtitle="Trimestre"))
        base = t
    else:
        chart(lineas(w, titulo="Semanal acumulado"))
        base = w
        keycol, keylbl = "sem", "Semana"

    if gran != "Trimestral":
        a, b = st.columns(2)
        with a:
            chart(lineas(base, ["Real_2025", "Real_2026"], "Real 2026 vs Real 2025 acumulado",
                         370, xcol=keycol, xtitle=keylbl))
        with b:
            # 2026-07-27: era "Real 2026 vs Nvo Plan" — Nvo Plan fuera de las
            # vistas, se cambió al par Real 2026 vs Plan 2026.
            chart(lineas(base, ["Real_2026", "Plan_2026"], "Real 2026 vs Plan 2026 acumulado",
                         370, xcol=keycol, xtitle=keylbl))

# ----------------------------------------------------------- 7. Cierres
import os
import openpyxl
from glob import glob

def _excel_seguimiento_mas_reciente():
    """Ruta + mtime del Excel más reciente (mtime se usa como clave de caché:
    si el archivo se vuelve a subir/editar, invalida _leer_seguimiento_excel)."""
    base = r"E:\Usuarios\112665\OneDrive - Onuris Tenant\Archivos de Karlo De Jesus Juarez Ramirez - Planeación Financiera\1. Capex y Herramientas\Master Cierres"
    archivos = glob(os.path.join(base, "Seguimiento Expansión - semana *.xlsx"))
    if not archivos:
        return None, None
    archivo = max(archivos, key=os.path.getmtime)
    return archivo, os.path.getmtime(archivo)

@st.cache_data(show_spinner="Leyendo Excel de seguimiento…")
def _leer_seguimiento_excel(archivo: str, _mtime: float):
    """Lee el Excel de Seguimiento Expansión más reciente del Master Cierres.
    Cacheado por (archivo, mtime) — evita releer/parsear el .xlsx con
    openpyxl en cada rerun de Streamlit (era el cuello de botella de la
    pestaña Cierres: el script completo se re-ejecuta con cada interacción)."""
    if archivo is None:
        return None, None, None
    wb = openpyxl.load_workbook(archivo, data_only=True)
    ws = wb["Seguimiento semanal"]

    tabla_resumen = {}
    tabla_detalle = []

    def _a_num(v):
        # Celdas de Plan/Sem26/Sem_Actual llegan con tipos mixtos (int/float/str)
        # desde el Excel — Arrow no serializa una columna object así.
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return 0

    for row_idx in range(6, 20):
        proyecto = ws[f"B{row_idx}"].value
        if not proyecto or proyecto in ("Total", None):
            continue
        tabla_resumen[proyecto] = {
            "Plan": _a_num(ws[f"I{row_idx}"].value),
            "Sem26": _a_num(ws[f"J{row_idx}"].value),
            "Sem_Actual": _a_num(ws[f"K{row_idx}"].value),
        }

    # Encabezados reales en fila 21: B=Proyecto C=Especialidad D=Formato
    # agrupado E=Formato F=Eco (id numérico) G=PDC (nombre) H=Comentario
    for row_idx in range(22, 530):
        proyecto = ws[f"B{row_idx}"].value
        if not proyecto:
            continue
        tabla_detalle.append({
            "Proyecto": proyecto,
            "Especialidad": ws[f"C{row_idx}"].value,
            "Formato_Agrupado": ws[f"D{row_idx}"].value,
            "Formato": ws[f"E{row_idx}"].value,
            "ECO": ws[f"F{row_idx}"].value,
            "PDC": ws[f"G{row_idx}"].value,
            "Comentario": ws[f"H{row_idx}"].value,
            "Plan": _a_num(ws[f"I{row_idx}"].value),
            "Sem26": _a_num(ws[f"J{row_idx}"].value),
            "Sem_Actual": _a_num(ws[f"K{row_idx}"].value),
        })

    return tabla_resumen, tabla_detalle, archivo

if SEC == "Cierres":
    _archivo_seg, _mtime_seg = _excel_seguimiento_mas_reciente()
    tabla_resumen, tabla_detalle, archivo = _leer_seguimiento_excel(_archivo_seg, _mtime_seg)

    if tabla_resumen is None:
        st.error("No se encontró Excel de 'Seguimiento Expansión'. Verifica la carpeta Master Cierres.")
    else:
        nom_archivo = os.path.basename(archivo)
        st.caption(f"📄 {nom_archivo}")

        st.markdown("### Resumen por tipo de cierre")
        _items_resumen = []
        for proyecto, valores in tabla_resumen.items():
            sem_actual = int(valores["Sem_Actual"])
            plan = int(valores["Plan"])
            delta = sem_actual - plan
            _items_resumen.append((proyecto, sem_actual, "",
                                  "alza" if delta > 0 else "baja"))
        st.markdown(kpis_tabla(_items_resumen), unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("### Tablas de seguimiento")

        df_detalle = pd.DataFrame(tabla_detalle)

        @st.cache_data(show_spinner=False)
        def _cons_pdc_pivot():
            """Gasto por PDC (Real25/Plan26/Real26) pivotado desde
            _consolidado.parquet. Cacheado: no cambia entre reruns del
            script, solo cuando corre build_data.py."""
            cons_pdc = (pl.scan_parquet(AGGS / "_consolidado.parquet")
                        .group_by(["cat_PDC", "Serie"]).agg(pl.col("monto").sum())
                        .collect().to_pandas()
                        .pivot(index="cat_PDC", columns="Serie", values="monto")
                        .fillna(0.0).reset_index())
            for _s in ("R25", "P26", "R26"):
                if _s not in cons_pdc.columns:
                    cons_pdc[_s] = 0.0
            return cons_pdc.rename(columns={
                "R26": "Gasto_Real_2026", "P26": "Gasto_Plan_2026", "R25": "Gasto_Real_2025"})

        cons_pdc = _cons_pdc_pivot()

        # El "ECO" del Excel es el Eco PDV (id) del PDC. PDC_IDS mapea
        # nombre_pdc -> id; se invierte para cruzar id -> nombre y así
        # agregar el gasto real/plan desde _consolidado.parquet (por cat_PDC).
        _ID_A_PDC = {}
        for _nombre, _pid in PDC_IDS.items():
            try:
                _ID_A_PDC[int(float(_pid))] = _nombre
            except (TypeError, ValueError):
                continue

        if not df_detalle.empty:
            # Vectorizado con merge (antes: .apply fila-por-fila + búsqueda
            # lineal en cons_pdc por cada fila — O(n²) y el principal cuello
            # de botella de la pestaña Cierres).
            def _eco_a_int(v):
                try:
                    return int(float(v))
                except (TypeError, ValueError):
                    return None
            _pdc_de_eco = df_detalle["ECO"].apply(_eco_a_int).map(_ID_A_PDC)
            _gasto = (pd.DataFrame({"cat_PDC": _pdc_de_eco})
                      .merge(cons_pdc, on="cat_PDC", how="left")
                      [["Gasto_Real_2026", "Gasto_Plan_2026", "Gasto_Real_2025"]]
                      .fillna(0.0))
            _gasto.index = df_detalle.index
            df_detalle[["Gasto_Real_2026", "Gasto_Plan_2026", "Gasto_Real_2025"]] = _gasto
            df_detalle["Δ Gasto vs Plan"] = df_detalle["Gasto_Real_2026"] - df_detalle["Gasto_Plan_2026"]

            gasto_cols = ["Gasto_Real_2025", "Gasto_Plan_2026", "Gasto_Real_2026", "Δ Gasto vs Plan"]
            _rename_gasto = {"Gasto_Real_2025": "Real 2025", "Gasto_Plan_2026": "Plan 2026",
                             "Gasto_Real_2026": "Real 2026", "Δ Gasto vs Plan": "Δ vs Plan"}

            st.markdown("**Primera tabla: Proyectos**")
            tbl1 = df_detalle[["Proyecto", "Especialidad", "Formato", "ECO", "PDC", "Comentario",
                              "Plan", "Sem26", "Sem_Actual"] + gasto_cols].copy()
            # ECO viene del Excel con tipos mixtos por fila (int/float/str) —
            # Arrow no serializa una columna object así; se homogeniza a texto.
            tbl1["ECO"] = tbl1["ECO"].apply(
                lambda v: "" if pd.isna(v) else
                (str(int(v)) if isinstance(v, float) and v.is_integer() else str(v)))
            tbl1 = tbl1.rename(columns=_rename_gasto)
            _cfg1 = {c: st.column_config.TextColumn(width=125) for c in list(_rename_gasto.values())}
            st.dataframe(estilizar(tbl1, ["Real 2025", "Plan 2026", "Real 2026", "Δ vs Plan"]),
                        width='stretch', hide_index=True, height=400, column_config=_cfg1)

            st.markdown("**Seguimiento a PDC**")
            def _eco_a_pdc_seguro(e):
                # 2026-07-30: fila de encabezado ("Eco") colada como dato en
                # la columna ECO del Excel hacía tronar float("Eco") — mismo
                # try/except que ya usa _eco_a_gasto() arriba (línea ~2904).
                try:
                    return _ID_A_PDC.get(int(float(e))) if pd.notna(e) else None
                except (TypeError, ValueError):
                    return None
            _c_buscar, _c_vista = st.columns([3, 1])
            with _c_buscar:
                _q_arbol = st.text_input(
                    "Buscar en el árbol (Grupo de Cuenta, Cuenta, Agrupador o Punto de Contacto)",
                    key="ci_arbol_q")
            with _c_vista:
                _vista_arbol = st.segmented_control(
                    "Vista", ["Version Plan", "Versión Semana Actual", "Todos"],
                    default="Todos", key="ci_arbol_vista")

            if _vista_arbol == "Version Plan":
                _df_detalle_vista = df_detalle[df_detalle["Plan"] == 1]
            elif _vista_arbol == "Versión Semana Actual":
                _df_detalle_vista = df_detalle[df_detalle["Plan"] == 0]
            else:
                _df_detalle_vista = df_detalle

            _pdc_tbl1 = tuple(_df_detalle_vista["ECO"].apply(_eco_a_pdc_seguro).dropna().unique())
            if not _pdc_tbl1:
                st.info("No se pudo mapear ningún ECO de esta tabla a un PDC del catálogo.")
            else:
                _cols_arbol_eco = ("cat_Grupo_de_Cuentas", "cat_Cuentas", "cat_Agrupador_Reales", "cat_PDC")
                _arbol_filt = (("cat_PDC", _pdc_tbl1),)
                if _q_arbol:
                    _q_low = _q_arbol.strip().lower()
                    _df_ca = EXTRA["cierres_arbol"].filter(pl.col("cat_PDC").is_in(list(_pdc_tbl1)))
                    _cond = pl.lit(False)
                    for _c in _cols_arbol_eco:
                        _cond = _cond | pl.col(_c).str.to_lowercase().str.contains(_q_low, literal=True)
                    _pdc_match = set(_df_ca.filter(_cond).select("cat_PDC").unique()
                                     .to_series().to_list())
                    # el buscador también debe encontrar por ID (Eco PDV) del PDC,
                    # que no aparece como texto dentro de cat_PDC (ese trae el
                    # nombre) — comparación EXACTA de id contra PDC_IDS (no
                    # substring, para no matchear 276 contra 1276/2765/etc.).
                    _pdc_match |= {nombre for nombre in _pdc_tbl1
                                  if _idtxt_pdc(nombre) == _q_low}
                    _pdc_match = list(_pdc_match)
                    if not _pdc_match:
                        st.info("Sin resultados para esa búsqueda.")
                        _pdc_match = []
                    _arbol_filt = (("cat_PDC", tuple(_pdc_match)),)
                if not _q_arbol or _pdc_match:
                    st.markdown(arbol_multicol("cierres_arbol", _cols_arbol_eco,
                                              si, sf, ind, _arbol_filt, MES, SEMS),
                               unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("### Árbol de cierres por estructura organizacional")

        det = cierres_rango(si, sf, ind, MES, SEMS, FILT)
        if det.empty:
            st.warning("Falta `aggs/cierres_expansion.parquet`. Corre `python build_data.py`.")
        else:
            st.caption("Haz clic en **+** para expandir cada rama.")

            f1, f2, f3 = st.columns(3)
            tipos_cie = ["(Todos)"] + sorted(det["Categoria"].unique())
            tsel = f1.selectbox("Categoría", tipos_cie, key="ci_tipo")
            tipo_col = None if tsel == "(Todos)" else tsel

            divs_ci = ["(Todas)"] + sorted(det["cat_Direccion_Division"].dropna().unique(), key=_orden_div_key)
            fdiv_ci = f2.selectbox("División", divs_ci, key="ci_div")

            terrs_ci = ["(Todos)"] + sorted(
                det[det["cat_Direccion_Division"] == fdiv_ci]["cat_Subdireccion_Territorio"].dropna().unique()
                if fdiv_ci != "(Todas)" else det["cat_Subdireccion_Territorio"].dropna().unique())
            fterr_ci = f3.selectbox("Territorio", terrs_ci, key="ci_terr")

            ci_filt = FILT
            if tipo_col is not None:
                ci_filt = ci_filt + (("Categoria", tipo_col),)
            if fdiv_ci != "(Todas)":
                ci_filt = ci_filt + (("cat_Direccion_Division", fdiv_ci),)
            if fterr_ci != "(Todos)":
                ci_filt = ci_filt + (("cat_Subdireccion_Territorio", fterr_ci),)

            if tipo_col is None and fdiv_ci == "(Todas)":
                st.info("Elige una Categoría o División para ver el árbol completo.")
            else:
                st.markdown(arbol_multicol("cierres_expansion",
                                          ("Categoria", "cat_Grupo_de_Cuentas", "cat_CtaMayor_Nombre",
                                           "cat_Direccion_Division", "cat_Subdireccion_Territorio",
                                           "cat_Subdireccion_Region"),
                                          si, sf, ind, ci_filt, MES, SEMS),
                           unsafe_allow_html=True)


# ----------------------------------------------------------- 8. Trimestres
@st.cache_data(show_spinner="Calculando trimestre…")
def _tri_arbol(q: str, filtros: tuple) -> str:
    """Árbol Grupo de Cuentas → Cuenta para el trimestre q, siempre sobre TODO
    el año de ese Q (independiente del filtro de semanas del sidebar — cada
    matriz elige su propio período, como en el HTML de referencia)."""
    lo, hi = QRANGO_TRI[q]
    return arbol_multicol("pdc", ("cat_Grupo_de_Cuentas", "cat_Cuentas"),
                          lo, hi, False, filtros)

if SEC == "Trimestres":
    def _pdc_filtrado():
        d = AGG["pdc"]
        for k, v_ in FILT:
            if k in d.columns:
                d = d.filter(pl.col(k).is_in(list(v_)) if isinstance(v_, tuple) else pl.col(k) == v_)
        return d

    st.caption("Usa el botón **+** de cada fila para desglosarla, "
               "y **−** para cerrarla.")
    ca, cb = st.columns(2)
    with ca:
        qa = st.selectbox("Matriz A", ["Q1", "Q2", "Q3", "Q4", "FY"], index=0, key="tri_qa")
        st.markdown(_tri_arbol(qa, FILT), unsafe_allow_html=True)
        descargar(_agregar_df(_pdc_filtrado(), ["cat_Grupo_de_Cuentas", "cat_Cuentas"],
                              *QRANGO_TRI[qa], False), f"trimestre_{qa}_A")
    with cb:
        qb = st.selectbox("Matriz B", ["Q1", "Q2", "Q3", "Q4", "FY"], index=4, key="tri_qb")
        st.markdown(_tri_arbol(qb, FILT), unsafe_allow_html=True)
        descargar(_agregar_df(_pdc_filtrado(), ["cat_Grupo_de_Cuentas", "cat_Cuentas"],
                              *QRANGO_TRI[qb], False), f"trimestre_{qb}_B")

# ----------------------------------------------------- 9. Comparación de Versiones
# 2026-07-30: pestaña independiente para comparar snapshots semanales COMPLETOS
# de la fuente cruda (r26_semXX.parquet / p26_semXX.parquet), no del agg activo
# en aggs/ — por eso NO usa AGG/compute()/arbol_multicol. No toca build_data.py.
CAT_CUENTAS_XLSX = BASE / "Catálogo grupo cuentas.xlsx"
COMPARATIVO_DIR = BASE / "aggs_versiones"

def _semanas_r26_disponibles() -> list[int]:
    """Semanas para las que existen AMBOS r26_semXX.parquet y p26_semXX.parquet
    en la raíz del proyecto, de mayor a menor."""
    sems = set()
    for fp in BASE.glob("r26_sem*.parquet"):
        m = re.match(r"r26_sem(\d+)\.parquet$", fp.name)
        if m and (BASE / f"p26_sem{m.group(1)}.parquet").exists():
            sems.add(int(m.group(1)))
    return sorted(sems, reverse=True)

@st.cache_data(show_spinner=False)
def _catalogo_cuentas_comparativo() -> pd.DataFrame:
    """Cta Mayor -> (Grupo de Cuentas, Cuentas), igual mapeo que build_data.py
    (COLS_CUENTAS) pero solo estas 2 columnas — no se replica el split de las
    12 cuentas CUENTAS_SPLIT_POSPRE porque este comparativo no baja a nivel
    PosPre, solo Grupo Cuenta -> Cuenta (colisión sin impacto en el total)."""
    # pl.read_excel: la firma de kwargs varía entre versiones de polars
    # (mismo problema que read_xlsx() en build_data.py) — se prueban las 3
    # variantes conocidas antes de caer al default sin forzar tipos.
    cta = None
    for kw in ({"infer_schema_length": 0},
               {"read_csv_options": {"infer_schema_length": 0}},
               {"read_options": {"infer_schema_length": 0}}):
        try:
            cta = pl.read_excel(CAT_CUENTAS_XLSX, **kw)
            break
        except TypeError:
            continue
    if cta is None:
        cta = pl.read_excel(CAT_CUENTAS_XLSX)
    cta = cta.rename({"Cta Mayor": "ID_CONCEPTO_CUENTA_NIV3"})
    cta = cta.with_columns([
        pl.col("ID_CONCEPTO_CUENTA_NIV3").cast(pl.Utf8).str.strip_chars(),
        pl.when(pl.col("Grupo de Cuentas").cast(pl.Utf8).str.strip_chars().fill_null("") == "")
          .then(pl.lit("(Sin dato)"))
          .otherwise(pl.col("Grupo de Cuentas").cast(pl.Utf8).str.strip_chars())
          .alias("cat_Grupo_de_Cuentas"),
        pl.when(pl.col("Cuentas").cast(pl.Utf8).str.strip_chars().fill_null("") == "")
          .then(pl.lit("(Sin dato)"))
          .otherwise(pl.col("Cuentas").cast(pl.Utf8).str.strip_chars())
          .alias("cat_Cuentas"),
    ]).unique(subset=["ID_CONCEPTO_CUENTA_NIV3"], keep="first")
    return cta.select(["ID_CONCEPTO_CUENTA_NIV3", "cat_Grupo_de_Cuentas", "cat_Cuentas"]).to_pandas()

def _cuentas_comparativo_semana_raw(sem: int) -> pd.DataFrame:
    """Real/Plan 2026 agregados a (SEMANA_EKT, Grupo de Cuentas, Cuentas) para
    el snapshot r26_sem{sem}/p26_sem{sem} — SIN colapsar semanas, para poder
    filtrar después tanto 'año completo' como 'una SEMANA_EKT específica'
    (caso: comparar cómo luce la semana 29 dentro del snapshot r29 vs cómo
    luce esa MISMA semana 29 dentro del snapshot r30, para detectar si al
    resubir datos cambió retroactivamente una semana ya cerrada).
    Cachea a disco en aggs_versiones/_comparativo/ — el .parquet fuente tiene
    ~16M filas y no vale la pena volver a agregarlo en cada visita a la
    pestaña; solo se recalcula si el archivo fuente cambia (mtime)."""
    cache_dir = COMPARATIVO_DIR / "_comparativo"
    cache_fp = cache_dir / f"cuentas_comparativo_sem{sem}_raw.parquet"
    r26_fp = BASE / f"r26_sem{sem}.parquet"
    p26_fp = BASE / f"p26_sem{sem}.parquet"
    if not r26_fp.exists() or not p26_fp.exists():
        return pd.DataFrame(columns=["SEMANA_EKT", "cat_Grupo_de_Cuentas", "cat_Cuentas",
                                     "Real_2026", "Plan_2026"])

    src_mtime = max(r26_fp.stat().st_mtime, p26_fp.stat().st_mtime)
    if cache_fp.exists() and cache_fp.stat().st_mtime >= src_mtime:
        return pd.read_parquet(cache_fp)

    cat = pl.from_pandas(_catalogo_cuentas_comparativo())
    monto_col = "SUM_MONTO_MNX_CONS_(SUMA)"
    gcols = ["SEMANA_EKT", "cat_Grupo_de_Cuentas", "cat_Cuentas"]

    def _agg_una(fp: Path, col_out: str) -> pl.LazyFrame:
        return (pl.scan_parquet(fp)
                 .with_columns(pl.col("ID_CONCEPTO_CUENTA_NIV3").cast(pl.Utf8).str.strip_chars())
                 .join(cat.lazy(), on="ID_CONCEPTO_CUENTA_NIV3", how="left")
                 .with_columns([
                     pl.col("cat_Grupo_de_Cuentas").fill_null("(Sin dato)"),
                     pl.col("cat_Cuentas").fill_null("(Sin dato)"),
                 ])
                 .group_by(gcols)
                 .agg(pl.col(monto_col).sum().alias(col_out)))

    real = _agg_una(r26_fp, "Real_2026")
    plan = _agg_una(p26_fp, "Plan_2026")
    out = (real.join(plan, on=gcols, how="outer")
              .with_columns([pl.col("Real_2026").fill_null(0.0), pl.col("Plan_2026").fill_null(0.0)])
              .sort(gcols)
              .collect().to_pandas())

    cache_dir.mkdir(parents=True, exist_ok=True)
    out.to_parquet(cache_fp, index=False)
    return out

def _cuentas_comparativo_semana(sem: int, semana_ekt: int | None = None) -> pd.DataFrame:
    """Real/Plan 2026 agregados a (Grupo de Cuentas, Cuentas) para el snapshot
    `sem`. Si `semana_ekt` es None, suma TODO el año (comportamiento previo);
    si se da, filtra solo esa SEMANA_EKT antes de agregar."""
    raw = _cuentas_comparativo_semana_raw(sem)
    if raw.empty:
        return pd.DataFrame(columns=["cat_Grupo_de_Cuentas", "cat_Cuentas", "Real_2026", "Plan_2026"])
    if semana_ekt is not None:
        raw = raw[raw["SEMANA_EKT"] == semana_ekt]
    return (raw.groupby(["cat_Grupo_de_Cuentas", "cat_Cuentas"], as_index=False)
               [["Real_2026", "Plan_2026"]].sum())

_COMP_CSS = """<style>
.vtc-row { display: grid; grid-template-columns: 340px repeat(5, minmax(130px, 1fr));
           min-width: 900px; align-items: center; padding: 4px 0;
           border-bottom: 1px solid #eef0f4; background: #ffffff; }
.vtc-row > span { text-align: right; white-space: nowrap; padding: 2px 6px;
                  border-left: 1px solid #f2f4f7; font-variant-numeric: tabular-nums; }
.vtc-row > span:first-child { text-align: left; border-left: none; font-weight: 500;
                              position: sticky; left: 0; background: inherit; z-index: 1; }
.vtc-head { position: sticky; top: 0; background: #f1f3f6 !important; color: #111827 !important;
            font-weight: 800 !important; z-index: 2; border-bottom: 2px solid #cfd4dc !important; }
.vtc-head > span { text-align: center !important; }
.vtc-head > span:first-child { text-align: left !important; }
.vtc-r26-head { background-color: #fff59d !important; }
.vtc-r26 { background-color: #fffde7 !important; font-weight: 600; }
.vtc-var-pos { color: #78281f !important; font-weight: 600 !important; }
.vtc-var-neg { color: #145a32 !important; font-weight: 600 !important; }
.vtc-var-zero { color: #2b2b2b; }
.vtc-nvl0 { background: #f8fafc; font-size: 14px; }
.vtc-nvl0 > span:first-child { font-weight: 700 !important; color: #0f172a; }
.vtc-nvl1 > span:first-child { font-weight: 500; color: #475569; }
.vtc-tot { background: #f1f3f6 !important; font-weight: 800 !important;
           border-top: 2px solid #cfd4dc !important; position: sticky; bottom: 0; }
.vtc-tree summary { list-style:none; cursor:pointer; }
.vtc-tree summary::-webkit-details-marker { display:none; }
.vtc-caret::before { content:'+'; display:inline-flex; align-items:center; justify-content:center;
    width:18px; height:18px; margin-right:7px; vertical-align:-4px; border:1px solid #b8bfc8;
    border-radius:4px; background:#f1f3f6; font-size:13px; font-weight:800; color:#2b2b2b; line-height:1; }
details[open] > summary .vtc-caret::before { content:'−'; }
.vtc-hoja::before { content:''; display:inline-block; width:18px; margin-right:7px; }
</style>"""

def _comp_fila(etiqueta: str, plan26, real_act, real_cmp, nivel: int, hoja: bool) -> str:
    var_abs = (real_act or 0.0) - (real_cmp or 0.0)
    var_pct = (var_abs / abs(real_cmp) * 100) if real_cmp else 0.0
    cls_var = "vtc-var-zero" if var_abs == 0 else ("vtc-var-pos" if var_abs > 0 else "vtc-var-neg")
    caret = "vtc-hoja" if hoja else "vtc-caret"
    import html as _h
    return (
        f"<div class='vtc-row vtc-nvl{nivel}'>"
        f"<span style='padding-left:{nivel*16+6}px'><span class='{caret}'></span>{_h.escape(str(etiqueta))}</span>"
        f"<span>{M(plan26)}</span>"
        f"<span class='vtc-r26'>{M(real_act)}</span>"
        f"<span class='vtc-r26'>{M(real_cmp)}</span>"
        f"<span class='{cls_var}'>{M(var_abs)}</span>"
        f"<span class='{cls_var}'>{P(var_pct)}</span>"
        f"</div>")

def _comp_arbol_html(d_act: pd.DataFrame, d_cmp: pd.DataFrame, lbl_act: str, lbl_cmp: str) -> str:
    """Árbol Grupo de Cuentas -> Cuentas comparando dos snapshots. `d_act` trae
    Plan_2026 (fijo, no cambia entre snapshots — se toma del snapshot actual)
    y Real_2026 de la semana actual; d_cmp aporta el Real_2026 de la semana
    comparada. Merge por (Grupo, Cuenta), no por _agregar_df/AGG (datos
    ajenos al agg activo del dashboard)."""
    m = (d_act.merge(d_cmp[["cat_Grupo_de_Cuentas", "cat_Cuentas", "Real_2026"]],
                     on=["cat_Grupo_de_Cuentas", "cat_Cuentas"], how="outer",
                     suffixes=("_act", "_cmp"))
              .fillna(0.0))
    head = (f"<div class='vtc-row vtc-head'><span>Grupo de Cuentas / Cuenta</span>"
            f"<span>Plan 2026</span>"
            f"<span class='vtc-r26-head'>Real {lbl_act}</span>"
            f"<span class='vtc-r26-head'>Real {lbl_cmp}</span>"
            f"<span>Variación</span><span>Variación %</span></div>")

    grupos = m.groupby("cat_Grupo_de_Cuentas", as_index=False).agg(
        Plan_2026=("Plan_2026", "sum"), Real_2026_act=("Real_2026_act", "sum"),
        Real_2026_cmp=("Real_2026_cmp", "sum"))
    grupos = grupos.reindex(grupos["Real_2026_act"].abs().sort_values(ascending=False).index)

    out = []
    for _, g in grupos.iterrows():
        hijos = m[m["cat_Grupo_de_Cuentas"] == g["cat_Grupo_de_Cuentas"]]
        hijos = hijos.reindex(hijos["Real_2026_act"].abs().sort_values(ascending=False).index)
        filas_hijas = "".join(
            _comp_fila(h["cat_Cuentas"], h["Plan_2026"], h["Real_2026_act"], h["Real_2026_cmp"], 1, True)
            for _, h in hijos.iterrows())
        fila_grupo = _comp_fila(g["cat_Grupo_de_Cuentas"], g["Plan_2026"], g["Real_2026_act"],
                               g["Real_2026_cmp"], 0, False)
        out.append(f"<details><summary>{fila_grupo}</summary>{filas_hijas}</details>")

    tot_plan, tot_act, tot_cmp = m["Plan_2026"].sum(), m["Real_2026_act"].sum(), m["Real_2026_cmp"].sum()
    tot_var = tot_act - tot_cmp
    tot_pct = (tot_var / abs(tot_cmp) * 100) if tot_cmp else 0.0
    cls_tot = "vtc-var-zero" if tot_var == 0 else ("vtc-var-pos" if tot_var > 0 else "vtc-var-neg")
    total_html = (f"<div class='vtc-row vtc-tot'><span>TOTAL</span>"
                  f"<span>{M(tot_plan)}</span><span class='vtc-r26'>{M(tot_act)}</span>"
                  f"<span class='vtc-r26'>{M(tot_cmp)}</span>"
                  f"<span class='{cls_tot}'>{M(tot_var)}</span><span class='{cls_tot}'>{P(tot_pct)}</span></div>")

    return (_COMP_CSS + "<div class='vt-tree-wrap'><div class='vtc-tree'>"
            + head + "".join(out) + total_html + "</div></div>")

if SEC == "Comparación de Versiones":
    st.markdown("### Comparación de Versiones")
    st.caption("Compara el Real 2026 de la versión vigente del dashboard contra "
               "el de cualquier semana archivada, por Grupo de Cuentas → Cuenta. "
               "El Plan 2026 se toma de la versión actual (no varía entre snapshots).")

    _sems_disp = _semanas_r26_disponibles()
    if not _sems_disp:
        st.warning("No se encontraron pares r26_semXX.parquet / p26_semXX.parquet en el proyecto.")
    else:
        _sem_actual = _sems_disp[0]
        _opciones_cmp = [s for s in _sems_disp if s != _sem_actual]
        c1, c2 = st.columns(2)
        with c1:
            st.text_input("Versión actual (dashboard)", value=f"Semana {_sem_actual}",
                         disabled=True, key="cv_actual")
        with c2:
            if _opciones_cmp:
                _sem_cmp = st.selectbox("Comparar contra versión", _opciones_cmp,
                                        format_func=lambda s: f"Semana {s}", key="cv_sem_cmp")
            else:
                _sem_cmp = None
                st.info("No hay otra versión archivada para comparar todavía.")

        # Selector de "Semana a revisar": cada snapshot r26_semXX trae TODO el
        # año (SEMANA_EKT 1-53) — el número del nombre es solo el corte con el
        # que se generó. Filtrar por una SEMANA_EKT deja ver cómo cambió el
        # dato de ESA semana entre snapshots (p.ej. sem29 dentro de r29 vs
        # sem29 dentro de r30 — detecta datos resubidos/corregidos).
        _opts_sem_ekt = ["Todo el año (acumulado)"] + [f"Semana {i}" for i in range(1, 54)]
        _sel_sem_ekt = st.selectbox("Semana a revisar", _opts_sem_ekt, index=0, key="cv_semana_ekt")
        _semana_ekt = None if _sel_sem_ekt == "Todo el año (acumulado)" else int(_sel_sem_ekt.split()[-1])

        if _sem_cmp is not None:
            with st.spinner("Calculando comparación…"):
                d_act = _cuentas_comparativo_semana(_sem_actual, _semana_ekt)
                d_cmp = _cuentas_comparativo_semana(_sem_cmp, _semana_ekt)
            if d_act.empty and d_cmp.empty:
                st.warning("Sin datos para esta combinación de versión/semana.")
            else:
                _lbl_extra = "" if _semana_ekt is None else f" · SEMANA_EKT {_semana_ekt}"
                st.caption("Haz clic en **+** para expandir cada Grupo de Cuentas.")
                st.markdown(_comp_arbol_html(d_act, d_cmp,
                                            f"{_sem_actual}{_lbl_extra}", f"{_sem_cmp}{_lbl_extra}"),
                           unsafe_allow_html=True)

